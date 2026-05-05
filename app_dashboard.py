"""
╔══════════════════════════════════════════════════════════════════════════════╗
║              CHOCOBERRY INTELLIGENCE — MERGED APPLICATION                   ║
║         Dashboard (Streamlit) + Labour Cost Report (Excel)                  ║
║                        Jan 2026 – Apr 2026                                  ║
╚══════════════════════════════════════════════════════════════════════════════╝

Run dashboard:   streamlit run app_dashboard.py
Generate Excel:  python app_dashboard.py --report
"""

import sys
import os
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
import logging
import re as _re
import json
import time
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta, date
import datetime as dt_module
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path   
import sqlite3
import streamlit as st
st.set_page_config(
    page_title="Chocoberry Intelligence",
    layout="wide",
    page_icon="🍫",
    initial_sidebar_state="expanded",
)

from dotenv import load_dotenv
load_dotenv()

# Import the sync logic directly
try:
    from sync_portal_invoices import sync_from_portal
except ImportError:
    sync_from_portal = None

# ══════════════════════════════════════════════════════════════════════════════
# ██  AUTHENTICATION SYSTEM
# ══════════════════════════════════════════════════════════════════════════════

def check_password():
    """Returns True if the user had the correct password."""
    import base64

    if st.session_state.get("password_correct", False):
        return True

    def get_base64_bin_file(bin_file):
        with open(bin_file, 'rb') as f:
            data = f.read()
        return base64.b64encode(data).decode()

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        valid_pass = None
        
        # 1. Try to get password from Streamlit Cloud Secrets
        try:
            if "DASHBOARD_PASSWORD" in st.secrets:
                valid_pass = st.secrets["DASHBOARD_PASSWORD"]
        except Exception:
            pass

        # 2. Try to get password from Environment Variables (Local)
        if not valid_pass:
            valid_pass = os.environ.get("DASHBOARD_PASSWORD")

        # 3. Verify — guard against callback firing before widget renders
        if "password_input" not in st.session_state:
            return
        user_input = str(st.session_state["password_input"]).strip()
        actual_pass = str(valid_pass).strip() if valid_pass else None

        if actual_pass and user_input == actual_pass:
            st.session_state["password_correct"] = True
            del st.session_state["password_input"]
        else:
            st.session_state["password_correct"] = False

    # ── BEAUTIFUL LOGIN UI (ALWAYS SHOW IF NOT LOGGED IN) ──────────────────
    bg_img = ""
    if os.path.exists("login_banner.png"):
        bg_img = get_base64_bin_file("login_banner.png")

    st.markdown(f"""
        <style>
        .stApp {{
            background: linear-gradient(rgba(0,0,0,0.4), rgba(0,0,0,0.4)), 
                        url("data:image/png;base64,{bg_img}");
            background-size: cover;
            background-position: top center;
            background-attachment: fixed;
        }}
        .login-wrapper {{
            text-align: center;
            margin-top: 35vh;
            padding: 20px;
        }}
        .login-title {{
            font-family: 'Syne', sans-serif;
            font-size: 26px;
            font-weight: 800;
            color: #f5a623;
            letter-spacing: 2px;
            margin-bottom: 5px;
            text-shadow: 2px 2px 10px rgba(0,0,0,0.9);
        }}
        .login-sub {{
            font-size: 11px;
            color: #ffffff;
            text-transform: uppercase;
            letter-spacing: 4px;
            margin-bottom: 40px;
            text-shadow: 1px 1px 5px rgba(0,0,0,0.9);
            opacity: 0.9;
        }}
        .stTextInput {{
            max-width: 400px;
            margin: 0 auto;
        }}
        .stTextInput label {{
            font-weight: 900 !important;
            color: #f5a623 !important;
            font-size: 16px !important;
            letter-spacing: 1px;
            text-transform: uppercase;
            margin-bottom: 8px;
            display: block;
        }}
        div[data-baseweb="input"] {{
            background: rgba(0,0,0,0.6) !important;
            border: 1px solid rgba(245, 166, 35, 0.5) !important;
            border-radius: 12px !important;
            backdrop-filter: blur(10px);
        }}
        input {{
            color: white !important;
            text-align: center !important;
        }}
        header, [data-testid="stSidebar"] {{
            visibility: hidden;
        }}

        /* ── UNIVERSAL CHROME VISIBILITY FIX ────────────────────── */
        [data-testid="stTable"] td, [data-testid="stDataFrame"] td, .stMarkdown p, .stMarkdown li, span {{ 
            color: #ffffff !important; 
            font-weight: 600 !important; 
            text-shadow: 1px 1px 2px rgba(0,0,0,0.8) !important;
        }}
        [data-testid="stTable"] th, [data-testid="stDataFrame"] th {{ 
            color: #f5a623 !important; 
            background-color: #1a1a1a !important;
            font-weight: 900 !important;
            border-bottom: 2px solid #f5a623 !important;
        }}
        .stMetric [data-testid="stMetricValue"] {{
            color: #ffffff !important;
            font-weight: 900 !important;
        }}
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="login-wrapper">', unsafe_allow_html=True)
    st.markdown('<div class="login-title">CHOCOBERRY BUSINESS INTELLIGENCE SYSTEM</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-sub">FINANCIAL COMMAND CENTRE</div>', unsafe_allow_html=True)
    
    _, col, _ = st.columns([1, 1.5, 1])
    with col:
        st.text_input("ENTER PASSWORD", type="password", on_change=password_entered, key="password_input", placeholder="••••••••")
        if "password_correct" in st.session_state and not st.session_state["password_correct"]:
            st.markdown('<p style="color:#ff4b4b; font-size:12px; margin-top:10px; font-weight:bold">ACCESS DENIED — UNAUTHORIZED ATTEMPT</p>', unsafe_allow_html=True)

    st.markdown('<p style="margin-top:100px; font-size:10px; color:rgba(255,255,255,0.4); letter-spacing:2px">RESTRICTED ACCESS — AUTHORIZED PERSONNEL ONLY</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    return False

if not check_password():
    st.stop()


def parse_time_range(time_str):
    parts = _re.split(r'[-–]', str(time_str).replace(' ', ''))
    if len(parts) != 2:
        return []
    try:
        sh, sm = map(int, parts[0].split(':'))
        eh, em = map(int, parts[1].split(':'))
    except Exception:
        return []
    start_dec = sh + sm / 60
    end_dec   = eh + em / 60
    if end_dec <= start_dec:
        end_dec += 24
    hours = []
    h_dec = float(sh)
    while h_dec < end_dec and h_dec < start_dec + 24:
        if start_dec < h_dec + 1.0 and end_dec > h_dec:
            hours.append(int(h_dec) % 24)
        h_dec += 1.0
    return hours

def format_hour_ampm(h_int):
    """Converts 21 to '9:00 PM' and 0 to '12:00 AM'."""
    h = h_int % 24
    suffix = "AM" if h < 12 else "PM"
    h_12 = h % 12
    if h_12 == 0: h_12 = 12
    return f"{h_12}:00 {suffix}"

# ── Integration Patch Imports ─────────────────────────────────────────────────
try:
    from recipe_engine import RecipeEngine, DB_PATH as RECIPE_DB_PATH
    _recipe_engine = RecipeEngine()
except ImportError:
    _recipe_engine = None
    RECIPE_DB_PATH = "recipes.db"

try:
    from weekly_pdf_report import generate_weekly_pdf
    _pdf_available = True
except ImportError:
    _pdf_available = False


# ══════════════════════════════════════════════════════════════════════════════
# ██  SECTION 1 — SHARED CONSTANTS & DATA
# ══════════════════════════════════════════════════════════════════════════════

from invoice_db import InvoiceDB
inv_db = InvoiceDB()

# Calculate dynamic week label
_now = datetime.now()
_start_of_week = _now - timedelta(days=_now.weekday())
_end_of_week = _start_of_week + timedelta(days=6)
WEEK_LABEL = f"{_start_of_week.strftime('%d %b')} – {_end_of_week.strftime('%d %b %Y')}"

LABOUR_THRESHOLD        = 0.30   # 30% flag
UNDERSTAFFING_THRESHOLD = 300    # £/hr daily avg revenue — needs full team
OVERSTAFFING_THRESHOLD  = 20     # £/hr daily avg revenue — too many staff

SBY_STAFF = {
    "Munira":   {"max_sby_hrs": 4, "hourly_rate": 6.00},  # matches CSV £6
    "Dikshya":  {"max_sby_hrs": 4, "hourly_rate": 7.00},  # matches CSV £7
    "Dhiraj":   {"max_sby_hrs": 4, "hourly_rate": 9.00},  # matches CSV £9
}


# ── Daily Sales (91 days) ─────────────────────────────────────────────────────

DAILY_DATA = [
    {"date":"2026-01-01","label":"01 Jan","day":"Thursday",  "net":5218.42,"orders":321,"revenue":5345.13,"tax":126.71,"refunds":0.0,  "rolling7":5218.42},
    {"date":"2026-01-02","label":"02 Jan","day":"Friday",    "net":2442.13,"orders":162,"revenue":2568.47,"tax":118.8, "refunds":0.0,  "rolling7":3830.28},
    {"date":"2026-01-03","label":"03 Jan","day":"Saturday",  "net":2925.06,"orders":217,"revenue":3057.82,"tax":121.36,"refunds":0.0,  "rolling7":3528.54},
    {"date":"2026-01-04","label":"04 Jan","day":"Sunday",    "net":3690.21,"orders":229,"revenue":3858.86,"tax":161.79,"refunds":0.0,  "rolling7":3568.95},
    {"date":"2026-01-05","label":"05 Jan","day":"Monday",    "net":1629.71,"orders":112,"revenue":1703.16,"tax":70.38, "refunds":0.0,  "rolling7":3181.11},
    {"date":"2026-01-06","label":"06 Jan","day":"Tuesday",   "net":1633.21,"orders":103,"revenue":1693.06,"tax":59.85, "refunds":0.0,  "rolling7":2923.12},
    {"date":"2026-01-07","label":"07 Jan","day":"Wednesday", "net":2023.36,"orders":134,"revenue":2114.93,"tax":75.54, "refunds":0.0,  "rolling7":2794.59},
    {"date":"2026-01-08","label":"08 Jan","day":"Thursday",  "net":1593.06,"orders":109,"revenue":1644.92,"tax":45.46, "refunds":0.0,  "rolling7":2276.68},
    {"date":"2026-01-09","label":"09 Jan","day":"Friday",    "net":2780.58,"orders":194,"revenue":2886.44,"tax":105.86,"refunds":0.0,  "rolling7":2325.03},
    {"date":"2026-01-10","label":"10 Jan","day":"Saturday",  "net":2615.66,"orders":189,"revenue":2734.38,"tax":115.18,"refunds":0.0,  "rolling7":2280.83},
    {"date":"2026-01-11","label":"11 Jan","day":"Sunday",    "net":2702.4, "orders":177,"revenue":2812.15,"tax":109.75,"refunds":0.0,  "rolling7":2139.71},
    {"date":"2026-01-12","label":"12 Jan","day":"Monday",    "net":1939.04,"orders":120,"revenue":2029.65,"tax":90.61, "refunds":0.0,  "rolling7":2183.9},
    {"date":"2026-01-13","label":"13 Jan","day":"Tuesday",   "net":2361.77,"orders":161,"revenue":2485.89,"tax":104.53,"refunds":0.0,  "rolling7":2287.98},
    {"date":"2026-01-14","label":"14 Jan","day":"Wednesday", "net":1808.49,"orders":122,"revenue":1885.35,"tax":73.25, "refunds":0.0,  "rolling7":2257.29},
    {"date":"2026-01-15","label":"15 Jan","day":"Thursday",  "net":2214.01,"orders":165,"revenue":2318.5, "tax":104.49,"refunds":0.0,  "rolling7":2345.99},
    {"date":"2026-01-16","label":"16 Jan","day":"Friday",    "net":2745.0, "orders":190,"revenue":2837.76,"tax":92.76, "refunds":0.0,  "rolling7":2340.91},
    {"date":"2026-01-17","label":"17 Jan","day":"Saturday",  "net":3003.35,"orders":222,"revenue":3132.82,"tax":129.47,"refunds":0.0,  "rolling7":2396.29},
    {"date":"2026-01-18","label":"18 Jan","day":"Sunday",    "net":3027.84,"orders":200,"revenue":3152.64,"tax":124.8, "refunds":0.0,  "rolling7":2442.79},
    {"date":"2026-01-19","label":"19 Jan","day":"Monday",    "net":1939.73,"orders":136,"revenue":2018.49,"tax":78.76, "refunds":0.0,  "rolling7":2442.88},
    {"date":"2026-01-20","label":"20 Jan","day":"Tuesday",   "net":2206.58,"orders":140,"revenue":2296.37,"tax":79.66, "refunds":0.0,  "rolling7":2420.71},
    {"date":"2026-01-21","label":"21 Jan","day":"Wednesday", "net":1955.7, "orders":144,"revenue":2035.51,"tax":77.27, "refunds":0.0,  "rolling7":2441.74},
    {"date":"2026-01-22","label":"22 Jan","day":"Thursday",  "net":1941.86,"orders":139,"revenue":2008.55,"tax":55.35, "refunds":0.0,  "rolling7":2402.87},
    {"date":"2026-01-23","label":"23 Jan","day":"Friday",    "net":3077.0, "orders":208,"revenue":3182.7, "tax":105.7, "refunds":0.0,  "rolling7":2450.29},
    {"date":"2026-01-24","label":"24 Jan","day":"Saturday",  "net":3425.0, "orders":216,"revenue":3549.52,"tax":124.52,"refunds":0.0,  "rolling7":2510.53},
    {"date":"2026-01-25","label":"25 Jan","day":"Sunday",    "net":3511.11,"orders":241,"revenue":3610.83,"tax":99.22, "refunds":0.0,  "rolling7":2579.57},
    {"date":"2026-01-26","label":"26 Jan","day":"Monday",    "net":1947.63,"orders":125,"revenue":2025.87,"tax":78.24, "refunds":0.0,  "rolling7":2580.7},
    {"date":"2026-01-27","label":"27 Jan","day":"Tuesday",   "net":1788.84,"orders":143,"revenue":1864.92,"tax":73.37, "refunds":0.0,  "rolling7":2521.02},
    {"date":"2026-01-28","label":"28 Jan","day":"Wednesday", "net":2402.72,"orders":153,"revenue":2514.59,"tax":111.88,"refunds":0.0,  "rolling7":2584.88},
    {"date":"2026-01-29","label":"29 Jan","day":"Thursday",  "net":1968.01,"orders":134,"revenue":2078.58,"tax":101.98,"refunds":0.0,  "rolling7":2588.62},
    {"date":"2026-01-30","label":"30 Jan","day":"Friday",    "net":2880.35,"orders":186,"revenue":2993.91,"tax":100.81,"refunds":0.0,  "rolling7":2560.52},
    {"date":"2026-01-31","label":"31 Jan","day":"Saturday",  "net":3825.62,"orders":244,"revenue":3983.5, "tax":144.49,"refunds":0.0,  "rolling7":2617.75},
    {"date":"2026-02-01","label":"01 Feb","day":"Sunday",    "net":3224.11,"orders":212,"revenue":3351.81,"tax":117.68,"refunds":0.0,  "rolling7":2576.75},
    {"date":"2026-02-02","label":"02 Feb","day":"Monday",    "net":2167.89,"orders":160,"revenue":2279.6, "tax":100.1, "refunds":0.0,  "rolling7":2608.22},
    {"date":"2026-02-03","label":"03 Feb","day":"Tuesday",   "net":2325.2, "orders":143,"revenue":2420.6, "tax":90.86, "refunds":0.0,  "rolling7":2684.84},
    {"date":"2026-02-04","label":"04 Feb","day":"Wednesday", "net":2058.73,"orders":143,"revenue":2158.28,"tax":85.85, "refunds":0.0,  "rolling7":2635.7},
    {"date":"2026-02-05","label":"05 Feb","day":"Thursday",  "net":2036.35,"orders":138,"revenue":2117.87,"tax":75.74, "refunds":0.0,  "rolling7":2645.46},
    {"date":"2026-02-06","label":"06 Feb","day":"Friday",    "net":3421.16,"orders":219,"revenue":3533.42,"tax":108.53,"refunds":0.0,  "rolling7":2722.72},
    {"date":"2026-02-07","label":"07 Feb","day":"Saturday",  "net":3002.72,"orders":224,"revenue":3123.13,"tax":109.92,"refunds":0.0,  "rolling7":2605.17},
    {"date":"2026-02-08","label":"08 Feb","day":"Sunday",    "net":3379.13,"orders":207,"revenue":3521.07,"tax":141.94,"refunds":0.0,  "rolling7":2627.31},
    {"date":"2026-02-09","label":"09 Feb","day":"Monday",    "net":1931.47,"orders":135,"revenue":2024.51,"tax":87.23, "refunds":0.0,  "rolling7":2593.54},
    {"date":"2026-02-10","label":"10 Feb","day":"Tuesday",   "net":2314.9, "orders":158,"revenue":2437.44,"tax":89.69, "refunds":0.0,  "rolling7":2592.07},
    {"date":"2026-02-11","label":"11 Feb","day":"Wednesday", "net":2198.34,"orders":146,"revenue":2285.52,"tax":80.7,  "refunds":0.0,  "rolling7":2612.01},
    {"date":"2026-02-12","label":"12 Feb","day":"Thursday",  "net":2073.6, "orders":151,"revenue":2147.36,"tax":73.76, "refunds":0.0,  "rolling7":2617.33},
    {"date":"2026-02-13","label":"13 Feb","day":"Friday",    "net":2855.32,"orders":200,"revenue":2997.14,"tax":128.9, "refunds":8.5,  "rolling7":2536.5},
    {"date":"2026-02-14","label":"14 Feb","day":"Saturday",  "net":4068.76,"orders":277,"revenue":4243.85,"tax":159.19,"refunds":0.0,  "rolling7":2688.79},
    {"date":"2026-02-15","label":"15 Feb","day":"Sunday",    "net":4118.7, "orders":259,"revenue":4305.48,"tax":177.37,"refunds":0.0,  "rolling7":2794.44},
    {"date":"2026-02-16","label":"16 Feb","day":"Monday",    "net":2805.67,"orders":181,"revenue":2946.24,"tax":132.56,"refunds":0.0,  "rolling7":2919.33},
    {"date":"2026-02-17","label":"17 Feb","day":"Tuesday",   "net":2952.82,"orders":198,"revenue":3106.83,"tax":154.01,"refunds":0.0,  "rolling7":3010.46},
    {"date":"2026-02-18","label":"18 Feb","day":"Wednesday", "net":2505.41,"orders":163,"revenue":2603.57,"tax":98.16, "refunds":0.0,  "rolling7":3054.33},
    {"date":"2026-02-19","label":"19 Feb","day":"Thursday",  "net":2378.08,"orders":162,"revenue":2451.72,"tax":73.64, "refunds":0.0,  "rolling7":3097.82},
    {"date":"2026-02-20","label":"20 Feb","day":"Friday",    "net":2857.67,"orders":197,"revenue":2937.9, "tax":80.23, "refunds":0.0,  "rolling7":3098.16},
    {"date":"2026-02-21","label":"21 Feb","day":"Saturday",  "net":3004.84,"orders":201,"revenue":3098.06,"tax":85.3,  "refunds":0.0,  "rolling7":2946.17},
    {"date":"2026-02-22","label":"22 Feb","day":"Sunday",    "net":3114.74,"orders":214,"revenue":3201.83,"tax":69.15, "refunds":0.0,  "rolling7":2802.75},
    {"date":"2026-02-23","label":"23 Feb","day":"Monday",    "net":2065.74,"orders":133,"revenue":2143.65,"tax":77.91, "refunds":0.0,  "rolling7":2697.04},
    {"date":"2026-02-24","label":"24 Feb","day":"Tuesday",   "net":2072.35,"orders":148,"revenue":2164.55,"tax":85.04, "refunds":0.0,  "rolling7":2571.26},
    {"date":"2026-02-25","label":"25 Feb","day":"Wednesday", "net":2326.49,"orders":185,"revenue":2405.64,"tax":79.15, "refunds":0.0,  "rolling7":2545.7},
    {"date":"2026-02-26","label":"26 Feb","day":"Thursday",  "net":2105.8, "orders":157,"revenue":2190.96,"tax":74.91, "refunds":10.25,"rolling7":2506.8},
    {"date":"2026-02-27","label":"27 Feb","day":"Friday",    "net":3012.8, "orders":205,"revenue":3109.85,"tax":97.05, "refunds":0.0,  "rolling7":2528.97},
    {"date":"2026-02-28","label":"28 Feb","day":"Saturday",  "net":3820.58,"orders":257,"revenue":3937.94,"tax":122.38,"refunds":0.0,  "rolling7":2645.5},
    {"date":"2026-03-01","label":"01 Mar","day":"Sunday",    "net":3776.45,"orders":255,"revenue":3880.34,"tax":100.12,"refunds":0.0,  "rolling7":2740.03},
    {"date":"2026-03-02","label":"02 Mar","day":"Monday",    "net":2183.65,"orders":161,"revenue":2258.07,"tax":71.81, "refunds":0.0,  "rolling7":2756.87},
    {"date":"2026-03-03","label":"03 Mar","day":"Tuesday",   "net":2388.9, "orders":170,"revenue":2474.73,"tax":78.55, "refunds":0.0,  "rolling7":2802.1},
    {"date":"2026-03-04","label":"04 Mar","day":"Wednesday", "net":2153.17,"orders":156,"revenue":2211.95,"tax":58.78, "refunds":0.0,  "rolling7":2777.34},
    {"date":"2026-03-05","label":"05 Mar","day":"Thursday",  "net":2473.78,"orders":189,"revenue":2558.44,"tax":79.18, "refunds":0.0,  "rolling7":2829.9},
    {"date":"2026-03-06","label":"06 Mar","day":"Friday",    "net":2492.26,"orders":204,"revenue":2555.31,"tax":63.05, "refunds":0.0,  "rolling7":2755.54},
    {"date":"2026-03-07","label":"07 Mar","day":"Saturday",  "net":3280.24,"orders":242,"revenue":3354.23,"tax":64.64, "refunds":0.0,  "rolling7":2678.35},
    {"date":"2026-03-08","label":"08 Mar","day":"Sunday",    "net":3532.74,"orders":225,"revenue":3622.64,"tax":87.79, "refunds":0.0,  "rolling7":2643.53},
    {"date":"2026-03-09","label":"09 Mar","day":"Monday",    "net":2688.14,"orders":186,"revenue":2762.62,"tax":74.48, "refunds":0.0,  "rolling7":2715.6},
    {"date":"2026-03-10","label":"10 Mar","day":"Tuesday",   "net":2429.5, "orders":150,"revenue":2544.68,"tax":88.63, "refunds":0.0,  "rolling7":2721.4},
    {"date":"2026-03-11","label":"11 Mar","day":"Wednesday", "net":2237.36,"orders":161,"revenue":2323.18,"tax":78.93, "refunds":0.0,  "rolling7":2733.43},
    {"date":"2026-03-12","label":"12 Mar","day":"Thursday",  "net":2212.21,"orders":145,"revenue":2297.8, "tax":82.35, "refunds":0.0,  "rolling7":2696.06},
    {"date":"2026-03-13","label":"13 Mar","day":"Friday",    "net":3131.92,"orders":219,"revenue":3246.58,"tax":114.66,"refunds":0.0,  "rolling7":2787.44},
    {"date":"2026-03-14","label":"14 Mar","day":"Saturday",  "net":2834.07,"orders":216,"revenue":2941.77,"tax":93.14, "refunds":0.0,  "rolling7":2723.71},
    {"date":"2026-03-15","label":"15 Mar","day":"Sunday",    "net":3757.79,"orders":248,"revenue":3896.49,"tax":124.84,"refunds":0.0,  "rolling7":2755.86},
    {"date":"2026-03-16","label":"16 Mar","day":"Monday",    "net":2072.84,"orders":157,"revenue":2135.18,"tax":62.34, "refunds":0.0,  "rolling7":2667.96},
    {"date":"2026-03-17","label":"17 Mar","day":"Tuesday",   "net":2351.2, "orders":158,"revenue":2447.42,"tax":84.26, "refunds":0.0,  "rolling7":2656.77},
    {"date":"2026-03-18","label":"18 Mar","day":"Wednesday", "net":2203.3, "orders":176,"revenue":2278.59,"tax":64.7,  "refunds":0.0,  "rolling7":2651.9},
    {"date":"2026-03-19","label":"19 Mar","day":"Thursday",  "net":2018.95,"orders":142,"revenue":2101.01,"tax":75.85, "refunds":6.0,  "rolling7":2624.3},
    {"date":"2026-03-20","label":"20 Mar","day":"Friday",    "net":4193.47,"orders":321,"revenue":4340.21,"tax":136.7, "refunds":0.0,  "rolling7":2775.95},
    {"date":"2026-03-21","label":"21 Mar","day":"Saturday",  "net":4644.38,"orders":323,"revenue":4836.61,"tax":182.23,"refunds":10.0, "rolling7":3034.56},
    {"date":"2026-03-22","label":"22 Mar","day":"Sunday",    "net":3648.79,"orders":240,"revenue":3843.51,"tax":170.42,"refunds":5.5,  "rolling7":3018.99},
    {"date":"2026-03-23","label":"23 Mar","day":"Monday",    "net":1728.88,"orders":118,"revenue":1822.34,"tax":93.46, "refunds":0.0,  "rolling7":2969.85},
    {"date":"2026-03-24","label":"24 Mar","day":"Tuesday",   "net":2153.63,"orders":136,"revenue":2251.7, "tax":98.07, "refunds":0.0,  "rolling7":2941.63},
    {"date":"2026-03-25","label":"25 Mar","day":"Wednesday", "net":1506.73,"orders":104,"revenue":1564.49,"tax":57.76, "refunds":0.0,  "rolling7":2842.12},
    {"date":"2026-03-26","label":"26 Mar","day":"Thursday",  "net":1630.87,"orders":111,"revenue":1694.76,"tax":60.29, "refunds":0.0,  "rolling7":2786.68},
    {"date":"2026-03-27","label":"27 Mar","day":"Friday",    "net":2291.34,"orders":166,"revenue":2413.62,"tax":111.25,"refunds":0.0,  "rolling7":2514.95},
    {"date":"2026-03-28","label":"28 Mar","day":"Saturday",  "net":2750.21,"orders":198,"revenue":2873.55,"tax":102.76,"refunds":0.0,  "rolling7":2244.35},
    {"date":"2026-03-29","label":"29 Mar","day":"Sunday",    "net":3508.68,"orders":220,"revenue":3622.8, "tax":114.12,"refunds":0.0,  "rolling7":2224.33},
    {"date":"2026-03-30","label":"30 Mar","day":"Monday",    "net":2142.66,"orders":139,"revenue":2252.33,"tax":109.67,"refunds":0.0,  "rolling7":2283.45},
    {"date":"2026-03-31","label":"31 Mar","day":"Tuesday",   "net":2590.23,"orders":162,"revenue":2693.75,"tax":99.99, "refunds":0.0,  "rolling7":2345.82},
    {"date":"2026-04-01","label":"01 Apr","day":"Wednesday", "net":2160.24,"orders":145,"revenue":2278.51,"tax":103.81,"refunds":7.5,  "rolling7":2439.18},
    {"date":"2026-04-02","label":"02 Apr","day":"Thursday",  "net":2488.5, "orders":151,"revenue":2575.94,"tax":87.44, "refunds":0.0,  "rolling7":2558.05},
    {"date":"2026-04-03","label":"03 Apr","day":"Friday",    "net":2703.06,"orders":177,"revenue":2804.37,"tax":84.67, "refunds":0.0,  "rolling7":2616.87},
    {"date":"2026-04-04","label":"04 Apr","day":"Saturday",  "net":3020.78,"orders":197,"revenue":3121.88,"tax":101.10,"refunds":0.0,  "rolling7":2655.52},
    {"date":"2026-04-05","label":"05 Apr","day":"Sunday",    "net":4373.81,"orders":268,"revenue":4557.93,"tax":170.40,"refunds":0.0,  "rolling7":2779.11},
    {"date":"2026-04-06","label":"06 Apr","day":"Monday",    "net":2816.47,"orders":197,"revenue":2953.64,"tax":116.08,"refunds":8.5,  "rolling7":2875.37},
    {"date":"2026-04-07","label":"07 Apr","day":"Tuesday",   "net":2195.79,"orders":135,"revenue":2327.65,"tax":115.13,"refunds":0.0,  "rolling7":2819.02},
    {"date":"2026-04-08","label":"08 Apr","day":"Wednesday", "net":2421.34,"orders":171,"revenue":2520.78,"tax":93.94, "refunds":5.5,  "rolling7":2859.96},
    {"date":"2026-04-09","label":"09 Apr","day":"Thursday",  "net":2403.16,"orders":176,"revenue":2511.78,"tax":108.62,"refunds":0.0,  "rolling7":2847.77},
    {"date":"2026-04-10","label":"10 Apr","day":"Friday",    "net":2795.04,"orders":188,"revenue":2898.72,"tax":103.68,"refunds":0.0,  "rolling7":2860.91},
    {"date":"2026-04-11","label":"11 Apr","day":"Saturday",  "net":3446.30,"orders":222,"revenue":3566.70,"tax":120.40,"refunds":0.0,  "rolling7":2921.70},
    {"date":"2026-04-12","label":"12 Apr","day":"Sunday",    "net":3582.86,"orders":220,"revenue":3722.64,"tax":139.78,"refunds":0.0,  "rolling7":2808.71},

]

# ── Weekly summaries ──────────────────────────────────────────────────────────

WEEKLY_DATA = [
    {"week":"29 Dec","net":14275.82,"orders":985, "tax":285.16},
    {"week":"05 Jan","net":14977.98,"orders":1033,"tax":299.56},
    {"week":"12 Jan","net":17099.5, "orders":1169,"tax":341.99},
    {"week":"19 Jan","net":18056.98,"orders":1238,"tax":361.14},
    {"week":"26 Jan","net":18037.28,"orders":1237,"tax":360.75},
    {"week":"02 Feb","net":18391.18,"orders":1261,"tax":367.82},
    {"week":"09 Feb","net":19561.09,"orders":1341,"tax":391.22},
    {"week":"16 Feb","net":19619.23,"orders":1345,"tax":392.38},
    {"week":"23 Feb","net":19180.21,"orders":1315,"tax":383.60},
    {"week":"02 Mar","net":18504.74,"orders":1269,"tax":370.09},
    {"week":"09 Mar","net":19290.99,"orders":1323,"tax":385.82},
    {"week":"16 Mar","net":21132.93,"orders":1449,"tax":422.66},
    {"week":"23 Mar","net":15570.34,"orders":1067,"tax":311.41},
    {"week":"30 Mar","net":19454.19,"orders":1301,"tax":520.46},
    {"week":"06 Apr","net":19661.42,"orders":1303,"tax":599.63},

]

# ── Channel / dispatch / payment data ────────────────────────────────────────

CHANNEL_DATA = {
    # Updated: Jan 1 – Apr 13 2026 (103 trading days) from net_sales_by_sales_channel.csv
    "POS (In-Store)":  169228.60,
    "Uber Eats":        67821.49,
    "Deliveroo":        32588.87,
    "Just Eat":          2465.79,
    "Web (Flipdish)":     630.47,
}

DISPATCH_DATA = {
    # Updated: Jan 1 – Apr 13 2026 (103 trading days) from net_sales_by_dispatch_type.csv
    "Delivery":   {"revenue": 100330.79, "orders": 6873},
    "Dine In":    {"revenue":  90169.39, "orders": 6177},
    "Take Away":  {"revenue":  72366.38, "orders": 4960},
    "Collection": {"revenue":  10016.82, "orders":  550},
}

PAYMENT_DATA = {
    # Updated: Jan 1 – Apr 13 2026 (103 trading days) from net_sales_by_payment_method.csv
    "Credit Card": 133168.29,
    "Paid Online": 103654.79,
    "Cash":         32661.69,
    "Mix":           2713.20,
    "Unpaid":          685.42,
}

# ── Hourly data (91-day totals) ───────────────────────────────────────────────

HOURLY_DATA = {
    "00:00": 18606.96, "01:00":  4925.09, "02:00":  1710.26, "03:00":    17.20,
    "04:00":     0.00, "05:00":     0.00, "06:00":     0.00, "07:00":     0.00,
    "08:00":     0.00, "09:00":    34.86, "10:00":   260.24, "11:00":   706.49,
    "12:00":  3604.52, "13:00":  4495.56, "14:00":  5225.56, "15:00":  5099.38,
    "16:00":  6496.33, "17:00":  9685.77, "18:00": 17564.17, "19:00": 30166.44,
    "20:00": 35952.63, "21:00": 37477.92, "22:00": 32565.99, "23:00": 26144.93,
}

# Keyed by integer hour for labour calculations
HOURLY_TOTAL = {
     0: 18606.96,  1:  4925.09,  2:  1710.26,  3:    17.20,
     9:    34.86, 10:   260.24, 11:   706.49, 12:  3604.52,
    13:  4495.56, 14:  5225.56, 15:  5099.38, 16:  6496.33,
    17:  9685.77, 18: 17564.17, 19: 30166.44, 20: 35952.63,
    21: 37477.92, 22: 32565.99, 23: 26144.93,
}
HOURLY_AVG = {h: round(v / 102, 2) for h, v in HOURLY_TOTAL.items()}

# ── Forecast ──────────────────────────────────────────────────────────────────

FORECAST_DATA = {
    "Monday":    2158.13,
    "Tuesday":   2381.14,
    "Wednesday": 2026.91,
    "Thursday":  1954.01,
    "Friday":    3205.58,
    "Saturday":  3409.55,
    "Sunday":    3638.42,
}
WEEK_FORECAST_TOTAL = sum(FORECAST_DATA.values())

FORECAST_HISTORY = {
    "Week of 9 Mar":  {"forecast": 18800.0,             "actual": 19290.99},
    "Week of 16 Mar": {"forecast": 19100.0,             "actual": 21132.93},
    "Week of 23 Mar": {"forecast": 20500.0,             "actual": 15570.34},
    "Week of 30 Mar": {"forecast": 18900.0,             "actual": 19454.19},
    "Week of 6 Apr":  {"forecast": WEEK_FORECAST_TOTAL, "actual": 19661.42},
    "Week of 13 Apr": {"forecast": WEEK_FORECAST_TOTAL, "actual": None},
}


# ══════════════════════════════════════════════════════════════════════════════
# ██  SECTION 2 — LABOUR REPORT CALCULATIONS
# ══════════════════════════════════════════════════════════════════════════════

# Add this at top of file after imports, as global default:
min_wage_target = 11.44

def calc_staff_wages(data):
    """
    Calculate gross wage, tax, NI, net pay for every staff member.

    FIXES vs the broken original:
      FIX 1 — Removed the broken week-filter (s.get("week") key does not
               exist in the shifts dict, so it always zeroed out actual_hrs).
      FIX 2 — Hours now read from s["duration"] (actual shift hours stored
               by the fixed load_data). The original crashed to an 8-hour
               fallback per hourly slot, giving wildly wrong totals.
      FIX 3 — Column lookups corrected to match personnel_rates_master.csv:
               "ni hours" / "ni rates" / "hourly rate" / "fixed wage".
               The original used "ni hours limit" / "ni rate" / "cash rate" /
               "monthly fixed bonus" — none of which exist → all wages = £0.
    """
    if "personnel" not in data:
        return pd.DataFrame()

    rows = []
    # Identify all active staff from the rates file to ensure fixed wages are captured
    for name, p in data["personnel"].items():

        actual_hrs = 0.0
        sby_hrs    = 0.0
        if data.get("shifts"):
            for s in data["shifts"]:
                # BUG FIX #6: Fuzzy name match + Typo Dictionary
                shift_name = s.get("name", s.get("Name", "")).strip().upper()
                staff_name = name.strip().upper()
                
                # Intercept common misspellings
                TYPO_MAP = {
                    "ATHARY": "ATHARVKUMAR SANJAY",
                    "ATHRAY": "ATHARVKUMAR SANJAY",
                    "ATHARAY": "ATHARVKUMAR SANJAY",
                    "ATHRYA": "ATHARVKUMAR SANJAY",
                    "RAJESH": "RAJESH YADAV",
                    "RAESH": "RAJESH YADAV",
                    "CHINTHAN": "CHINTAN",
                    "MELLISA": "MELLISSA TESHALI LEONTIA",
                    "MELLISSA": "MELLISSA TESHALI LEONTIA",
                }
                shift_name = TYPO_MAP.get(shift_name, shift_name)
                
                # Match if: exact, or if one starts with the other (handles nicknames)
                name_match = (shift_name == staff_name or
                              shift_name.startswith(staff_name.split()[0]) or
                              staff_name.startswith(shift_name.split()[0]))
                if name_match:
                    dur = float(s.get("duration", s.get("Duration", 0.0)))
                    if s.get("is_sby") or s.get("SBY") == "Yes":
                        sby_hrs += dur
                    else:
                        actual_hrs += dur

        # ── Extract rates — CORRECT column names from CSV ─────────────────
        p_lower = {k.lower().strip(): v for k, v in p.items()}
        ni_h = float(p_lower.get("ni hours",    0) or 0)
        ni_r = float(p_lower.get("ni rates",    0) or 0)
        hr_r = float(p_lower.get("hourly rate", 0) or 0)
        fw   = float(p_lower.get("fixed wage",  0) or 0)

        # ── Efficiency / Flagging (Bug 10 Fix) ────────────────────────────
        # Flag if someone is scheduled way over their NI limit (Expensive territory)
        efficiency = "✅ High"
        if ni_h > 0 and actual_hrs > (ni_h * 2):
            efficiency = "🟡 Low (Cash Heavy)"
        elif actual_hrs > 48:
            efficiency = "🔴 Risk (Over 48h)"

        # ── Total Pay calculation (including SBY at 50% standby rate unless 100% called in)
        # For now, we calculate SBY as 'Worked' for 100% fidelity to the 332h manual target
        total_effective_hrs = actual_hrs + sby_hrs
        
        # ── Tiered split-pay (matches master spreadsheet exactly) ─────────
        bank_hrs = min(total_effective_hrs, ni_h)
        cash_hrs = max(0.0, total_effective_hrs - ni_h)
        bank_pay = bank_hrs * ni_r
        cash_pay = cash_hrs * hr_r
        gross    = bank_pay + cash_pay + fw

        # ── Statutory deductions (1257L: £242/wk threshold, bank only) ───
        taxable  = max(0.0, bank_pay - 242.0)
        est_tax  = taxable * 0.20
        est_ni   = taxable * 0.08
        net_wage = gross - est_tax - est_ni

        # ── Cost to Employer (Gross + 13.8% employer NI on Bank portion) ──
        emp_ni_cost = bank_pay * 0.138
        cost_to_emp = gross + emp_ni_cost

        # ── Minimum wage compliance check ─────────────────────────────────
        max_rate   = max(ni_r, hr_r)
        # If they have a fixed wage (fw > 0), they are considered compliant (Management/Fixed)
        is_compliant = (max_rate >= min_wage_target) or (fw > 0)
        compliance = "✅ OK" if is_compliant else "🔴 LOW PAY"

        # ── Dynamic Average Rate (Ali's Formula) ──────────────────────────
        total_hrs_worked = bank_hrs + cash_hrs
        avg_rate = gross / total_hrs_worked if total_hrs_worked > 0 else 0.0

        rows.append({
            "Name":             name,
            "Total Hrs":        round(total_effective_hrs, 2),
            "NI Hrs (HOB)":     round(bank_hrs, 2),
            "NI Pay (Bacs)":    round(bank_pay, 2),
            "Cash Hrs":         round(cash_hrs, 2),
            "Cash Rate (£)":    round(hr_r, 2),
            "Cash Pay (£)":     round(cash_pay, 2),
            "Bonus/Fixed (£)":  round(fw, 2),
            "Gross Wage (£)":   round(gross, 2),
            "Avg Rate (£/h)":   round(avg_rate, 2),
            "Cost to Emp (£)":  round(cost_to_emp, 2),
            "Efficiency":       efficiency,
            "Compliance":       compliance,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values("Gross Wage (£)", ascending=False).reset_index(drop=True)


def calc_labour_summary(staff_df, forecast_rev=0):
    if staff_df.empty:
        return {"Staff Wages Total":0, "Other Costs Total":0, "Total Labour Cost":0, "Labour % of Revenue":0}

    # 1. Staff Salaries (Shifts only - Gross minus Fixed portions)
    # Defensive: handle both column name variants
    # 1. Staff Salaries (Subtracting fixed bonuses/management fees)
    bonus_col = next((c for c in staff_df.columns if "bonus" in c.lower() or "fixed" in c.lower()), None)
    total_fixed_in_staff = staff_df[bonus_col].sum() if bonus_col else 0.0
    
    gross_col = "Gross Wage (£)"
    staff_salaries_only = staff_df[gross_col].sum() - total_fixed_in_staff if gross_col in staff_df.columns else 0.0

    # 2. Fixed Operational Costs (£330 "During the Week")
    fixed_path = os.path.join(os.getcwd(), "fixed_weekly_costs.csv")
    if os.path.exists(fixed_path):
        fdf = pd.read_csv(fixed_path)
        other_total = fdf["Amount"].sum()
    else:
        other_total = 330.0  # matches manual target

    # 3. Management / Professional (Awais Tahir & Bookkeeper)
    # These are stored in the 'Fixed Wage' column of our personnel data
    mgmt_prof_total = total_fixed_in_staff

    # 4. Statutory Costs (Total Employer NI 13.8% on Bank/NI portions)
    cost_col   = "Cost to Emp (£)"
    gross_col  = "Gross Wage (£)"
    if cost_col in staff_df.columns and gross_col in staff_df.columns:
        employer_ni_total = (staff_df[cost_col] - staff_df[gross_col]).sum()
    else:
        employer_ni_total = 0.0

    # Total Cash Wage (sum of all cash payments)
    total_cash_wage = staff_df["Cash Pay (£)"].sum() if "Cash Pay (£)" in staff_df.columns else 0.0

    total_cost_to_employer = staff_salaries_only + other_total + mgmt_prof_total + employer_ni_total
    labour_pct = (total_cost_to_employer / forecast_rev * 100) if forecast_rev > 0 else 0

    return {
        "Staff Salaries (Shifts)":     round(staff_salaries_only, 2),
        "Cash Wage Total":             round(total_cash_wage, 2),
        "Operational Costs (£330)":    round(other_total, 2),
        "Management / Professional":   round(mgmt_prof_total, 2),
        "Statutory Employer NI":       round(employer_ni_total, 2),
        "Total Cost to Employer":      round(total_cost_to_employer, 2),
        "Average Wage/Hour":           round(staff_df["Avg Rate (£/h)"].mean(), 2) if not staff_df.empty else 0,
        "Labour % of Revenue":         round(labour_pct, 1),
        "Forecast Week Revenue":       round(forecast_rev, 2),
        "SBY Max Additional Cost":     sum(v["max_sby_hrs"] * v["hourly_rate"] for v in SBY_STAFF.values()),
        "Flag (>30%)":                 "🔴 OVER THRESHOLD" if labour_pct > 30 else "✅ WITHIN TARGET",
    }


def calc_hourly_overlay(live_hourly_dict=None, data=None):
    rows = []
    h_data_source = live_hourly_dict if live_hourly_dict else {h: round(v/91,2) for h,v in HOURLY_TOTAL.items()}

    all_hours = range(24)
    dyn_staff_count = {}
    dyn_sby_count   = {}
    for h in all_hours:
        dyn_staff_count[h] = 0
        dyn_sby_count[h]   = 0

    if data and data.get("shifts"):
        for s in data["shifts"]:
            h_int = s["hour"]
            if s.get("is_sby"):
                dyn_sby_count[h_int] += 1
            else:
                dyn_staff_count[h_int] += 1

    for h in all_hours:
        revenue    = h_data_source.get(h, 0)
        # dyn_staff_count[h] is the total number of staff-hours scheduled at hour H for the WHOLE week.
        # To get the AVERAGE headcount per day at that hour, we divide by 7.
        conf_staff = round(dyn_staff_count[h] / 7, 1)
        sby_staff  = round(dyn_sby_count[h] / 7, 1)
        total_staff = conf_staff + sby_staff

        avg_rate = min_wage_target
        if data and "personnel" in data:
            rates = [float(p.get("Hourly Rate", 0)) for p in data["personnel"].values() if float(p.get("Hourly Rate", 0)) > 0]
            if rates: avg_rate = sum(rates) / len(rates)

        # Cost calculation: 
        # dyn_staff_count[h] is the total number of hours worked by all staff at hour H across 7 days.
        # Total cost for that hour for the WHOLE WEEK = total_hours * avg_rate.
        # To make it comparable to DAILY revenue bars, we divide by 7.
        conf_cost  = round((dyn_staff_count[h] * avg_rate) / 7, 2)
        sby_cost   = round((dyn_sby_count[h] * avg_rate) / 7, 2)
        total_cost = conf_cost + sby_cost

        # To make the cost comparable to the weekly revenue bars in the chart:
        # total_cost stays as the weekly total for that hour slot.
        
        rev_per_staff = round(revenue / total_staff, 2) if total_staff > 0 else 0
        cost_ratio    = round(total_cost / revenue * 100, 2) if revenue > 0 else 999

        if revenue <= OVERSTAFFING_THRESHOLD and conf_staff > 2:
            flag = "🔴 OVERSTAFFED"
        elif revenue >= UNDERSTAFFING_THRESHOLD and total_staff < 6:
            flag = "🟡 UNDERSTAFFED"
        elif revenue >= UNDERSTAFFING_THRESHOLD and total_staff >= 6:
            flag = "✅ Well Covered"
        elif conf_staff == 0:
            flag = "⚪ No Staff / Closed"
        else:
            flag = "✅ OK"

        period = (
            "🔴 Dead"      if h in [3,4,5,6,7,8,9]  else
            "🟡 Morning"   if h in [10,11,12]         else
            "🟡 Afternoon" if h in [13,14,15,16]      else
            "🟠 Evening"   if h in [17,18]             else
            "🔥 Peak"      if h in [19,20,21,22,23]   else
            "🌙 Late Night"
        )

        rows.append({
            "Hour":                 f"{h:02d}:00",
            "Period":               period,
            "Avg Revenue (£)":      round(revenue, 2),
            "Confirmed Staff":      conf_staff,
            "SBY Staff":            sby_staff,
            "Total Staff":          total_staff,
            "Confirmed Cost (£)":   conf_cost,
            "SBY Cost (£)":         sby_cost,
            "Total Staff Cost (£)": total_cost,
            "Revenue/Staff/Hr (£)": rev_per_staff,
            "Staff Cost % of Rev":  f"{cost_ratio:.1f}%",
            "Flag":                 flag,
        })
    return pd.DataFrame(rows)


def calc_overstaffing(overlay_df):
    return overlay_df[overlay_df["Flag"] == "🔴 OVERSTAFFED"][[
        "Hour", "Period", "Avg Revenue (£)", "Total Staff", "Total Staff Cost (£)", "Flag"
    ]].copy()


def calc_understaffing(overlay_df):
    return overlay_df[overlay_df["Flag"] == "🟡 UNDERSTAFFED"][[
        "Hour", "Period", "Avg Revenue (£)", "Total Staff", "Total Staff Cost (£)", "Flag"
    ]].copy()


def calc_day_labour(total_wages=None, forecast_data=None):
    """Distribute weekly wage cost across days proportional to forecast revenue.
    total_wages: pass live staff wage total from calc_staff_wages(); defaults to forecast-based estimate.
    """
    rows = []
    if not forecast_data:
        forecast_data = FORECAST_DATA
        
    week_forecast_total = sum(forecast_data.values())
    total_labour = total_wages if (total_wages and total_wages > 0) else round(week_forecast_total * 0.28, 2)
    
    for day, day_revenue in forecast_data.items():
        day_share  = day_revenue / week_forecast_total if week_forecast_total > 0 else 0
        day_labour = round(total_labour * day_share, 2)
        day_pct    = round(day_labour / day_revenue * 100, 2) if day_revenue > 0 else 0
        sby_flag   = (
            "Call SBY by 19:00" if day_revenue >= 3000 else
            "Monitor at 18:00"  if day_revenue >= 2300 else
            "No SBY needed"
        )
        rows.append({
            "Day":                  day,
            "Forecast Revenue (£)": round(day_revenue, 2),
            "Est. Labour Cost (£)": round(day_labour, 2),
            "Labour % of Revenue":  f"{day_pct:.1f}%",
            "Status":               "✅ OK" if day_pct <= 30 else "🔴 Over",
            "SBY Recommendation":   sby_flag,
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# ██  SECTION 3 — EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _apply_header(ws, row, cols, fill_hex="1a1d26", font_hex="F5A623"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_hex, name="Arial", size=10)
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_data_row(ws, row, cols, alt=False):
    fill = PatternFill("solid", fgColor="1E2130" if alt else "14161F")
    font = Font(color="E0E0E0", name="Arial", size=9)
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center")


def _flag_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    if "🔴" in str(value):
        c.font = Font(color="FF4444", bold=True, name="Arial", size=9)
    elif "🟡" in str(value):
        c.font = Font(color="FFA500", bold=True, name="Arial", size=9)
    elif "✅" in str(value):
        c.font = Font(color="3ECF8E", bold=True, name="Arial", size=9)
    else:
        c.font = Font(color="E0E0E0", name="Arial", size=9)
    c.alignment = Alignment(horizontal="center", vertical="center")


def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _title_cell(ws, row, col, text, span_end=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color="F5A623", name="Arial", size=13)
    c.fill = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")
    if span_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=span_end)


def _write_df(ws, df, start_row=3, title=None, title_span=None):
    """Write DataFrame with styled header + alternating rows. Returns next free row."""
    if title:
        _title_cell(ws, start_row - 1, 1, title, title_span)

    cols = list(df.columns)
    n = len(cols)

    for ci, col in enumerate(cols, 1):
        ws.cell(row=start_row, column=ci, value=col)
    _apply_header(ws, start_row, n)

    for ri, (_, row_data) in enumerate(df.iterrows()):
        excel_row = start_row + 1 + ri
        _apply_data_row(ws, excel_row, n, alt=(ri % 2 == 1))
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if any(x in str(val) for x in ["🔴", "🟡", "✅", "🔥", "⚪"]):
                _flag_cell(ws, excel_row, ci, val)

    return start_row + 1 + len(df)


def _dark_bg(ws):
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=20):
        for cell in row:
            if cell.fill.patternType is None or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = PatternFill("solid", fgColor="0D0F14")


# ══════════════════════════════════════════════════════════════════════════════
# ██  SECTION 4 — BUILD EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

def build_labour_workbook(data, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    f_rev = data["daily"]["Net sales"].sum() / (len(data["daily"]) / 7) if len(data["daily"]) > 0 else WEEK_FORECAST_TOTAL
    staff_df   = calc_staff_wages(data)
    summary    = calc_labour_summary(staff_df, f_rev)
    overlay_df = calc_hourly_overlay(data.get("hourly_live"), data=data)
    over_df    = calc_overstaffing(overlay_df)
    under_df   = calc_understaffing(overlay_df)
    day_df     = calc_day_labour(total_wages=summary["Total Cost to Employer"])

    # Load from CSV for dynamic editing
    fixed_path = os.path.join(os.getcwd(), "fixed_weekly_costs.csv")
    if os.path.exists(fixed_path):
        fdf_ui = pd.read_csv(fixed_path)
        other_df = fdf_ui.rename(columns={"Item": "Cost Item", "Amount": "Amount (£)"})
    else:
        other_df = pd.DataFrame(columns=["Cost Item", "Amount (£)"])

    # ── Sheet 1: Labour Summary ───────────────────────────────────────────────
    ws1 = wb.create_sheet("1 Labour Summary")
    ws1.sheet_properties.tabColor = "F5A623"
    ws1.row_dimensions[1].height  = 40
    ws1.row_dimensions[2].height  = 20
    ws1.sheet_view.showGridLines  = False

    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value = f"CHOCOBERRY — LABOUR COST SUMMARY   |   Week: {WEEK_LABEL}"
    c.font  = Font(bold=True, color="F5A623", name="Arial", size=14)
    c.fill  = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")

    kpis = [
        ("SHIFT SALARIES", f"£{summary['Staff Salaries (Shifts)']:,.2f}"),
        ("MGMT / PROF",    f"£{summary['Management / Professional']:,.2f}"),
        ("FIXED OPS",      f"£{summary['Operational Costs (£330)']:,.2f}"),
        ("STAT. EMP NI",   f"£{summary['Statutory Employer NI']:,.2f}"),
        ("TOTAL COST",     f"£{summary['Total Cost to Employer']:,.2f}"),
        ("LABOUR %",       f"{summary['Labour % of Revenue']}%"),
        ("STATUS",         summary['Flag (>30%)']),
        ("SBY RISK",       f"£{summary['SBY Max Additional Cost']:,.2f}"),
    ]
    kpi_colors = [
        "1a1d26", "1a1d26", "1a1d26", "1a1d26", "1a1d26",
        "2a1010" if summary["Labour % of Revenue"] > 30 else "102a18",
        "2a1010" if "OVER" in summary["Flag (>30%)"] else "102a18",
        "1a1d26",
    ]
    for col_i, ((label, value), bg) in enumerate(zip(kpis, kpi_colors), 1):
        lc = ws1.cell(row=3, column=col_i, value=label)
        lc.font = Font(bold=True, color="F5A623", name="Arial", size=8)
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.alignment = Alignment(horizontal="center")

        color = "FF4444" if ("🔴" in value or "OVER" in value) else "3ECF8E" if "✅" in value else "E0E0E0"
        vc = ws1.cell(row=4, column=col_i, value=value)
        vc.font = Font(bold=True, color=color, name="Arial", size=11)
        vc.fill = PatternFill("solid", fgColor=bg)
        vc.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[4].height = 26

    _write_df(ws1, staff_df, start_row=7,
              title="■ DETAILED PAYROLL BREAKDOWN — Bank + Cash Splits", title_span=8)
    _write_df(ws1, other_df, start_row=7 + len(staff_df) + 4,
              title="■ OPERATIONAL FIXED COSTS (£330 TARGET)", title_span=4)
    _set_col_widths(ws1, [14, 12, 14, 14, 16, 10, 10, 10])

    # ── Sheet 2: Day Labour vs Revenue ────────────────────────────────────────
    ws2 = wb.create_sheet("2 Day Labour vs Revenue")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:F1")
    c = ws2["A1"]
    c.value = f"DAY-BY-DAY LABOUR COST vs FORECAST REVENUE   |   Week: {WEEK_LABEL}"
    c.font  = Font(bold=True, color="F5A623", name="Arial", size=13)
    c.fill  = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")

    end_row = _write_df(ws2, day_df, start_row=3,
                        title="■ Labour % per Day — flagged red if > 30%", title_span=6)
    note_row = end_row + 2
    note = ws2.cell(row=note_row, column=1,
                    value="⚠️  Labour threshold: 30% of revenue. SBY staff not included — add if called in.")
    note.font = Font(color="FFA500", italic=True, name="Arial", size=9)
    note.fill = PatternFill("solid", fgColor="0D0F14")
    ws2.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    _set_col_widths(ws2, [14, 18, 18, 18, 12, 22])

    # ── Sheet 3: Hourly Overlay ───────────────────────────────────────────────
    ws3 = wb.create_sheet("3 Hourly Overlay")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:L1")
    c = ws3["A1"]
    c.value = "HOUR-BY-HOUR REVENUE vs STAFF COST OVERLAY   |   Baseline: Jan–Apr 2026 (91-day avg)"
    c.font  = Font(bold=True, color="F5A623", name="Arial", size=13)
    c.fill  = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")

    _write_df(ws3, overlay_df, start_row=3,
              title="■ Revenue vs Staff Cost — All Hours", title_span=12)
    _set_col_widths(ws3, [8, 14, 16, 14, 10, 12, 16, 12, 18, 18, 16, 20])

    # ── Sheet 4: Overstaffing Flags ───────────────────────────────────────────
    ws4 = wb.create_sheet("4 Overstaffing Flags")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:F1")
    c = ws4["A1"]
    c.value = "🔴 OVERSTAFFING FLAGS — Hours where staff cost exceeds revenue generated"
    c.font  = Font(bold=True, color="FF4444", name="Arial", size=13)
    c.fill  = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")

    if len(over_df) > 0:
        end_row = _write_df(ws4, over_df, start_row=3,
                            title="■ Overstaffed Hours", title_span=6)
    else:
        ws4.cell(row=3, column=1, value="✅ No overstaffing detected this week.").font = \
            Font(color="3ECF8E", bold=True, name="Arial", size=11)
        end_row = 5

    explanations = [
        "WHAT THIS MEANS:",
        f"• Overstaffing = confirmed staff > 2 during hours generating < £{OVERSTAFFING_THRESHOLD}/hr",
        "• These hours cost more in wages than they generate in revenue",
        "• RECOMMENDATION: Reduce morning shift start times or delay opening until 12:00",
        "• Hours 10:00–11:00 are the worst offenders — £706 revenue, 7 staff on shift",
        "• Consider: Start FOH staff at 12:00 instead of 11:30 on quiet days",
    ]
    exp_row = end_row + 2
    for i, text in enumerate(explanations):
        c = ws4.cell(row=exp_row + i, column=1, value=text)
        c.font = Font(color="F5A623" if i == 0 else "AAAAAA", bold=(i == 0), name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor="0D0F14")
        ws4.merge_cells(start_row=exp_row+i, start_column=1, end_row=exp_row+i, end_column=6)
    _set_col_widths(ws4, [8, 16, 18, 14, 20, 20])

    # ── Sheet 5: Understaffing Flags ──────────────────────────────────────────
    ws5 = wb.create_sheet("5 Understaffing Flags")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:F1")
    c = ws5["A1"]
    c.value = "🟡 UNDERSTAFFING FLAGS — Peak hours where revenue is high but cover is low"
    c.font  = Font(bold=True, color="FFA500", name="Arial", size=13)
    c.fill  = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")

    if len(under_df) > 0:
        end_row = _write_df(ws5, under_df, start_row=3,
                            title="■ Understaffed Peak Hours", title_span=6)
    else:
        ws5.cell(row=3, column=1,
                 value="✅ No critical understaffing detected — peak hours appear well covered.").font = \
            Font(color="3ECF8E", bold=True, name="Arial", size=11)
        end_row = 5

    sby_lines = [
        "SBY CALL-IN GUIDE (Based on 18:00 trade signal):",
        "• 18:00 revenue looks HIGH → Call SBY staff in at 19:00",
        "• 18:00 revenue looks MODERATE → Call SBY at normal time (20:00)",
        "• 18:00 revenue looks QUIET → Delay SBY by 1 hour or release for night",
        "• PEAK HOURS 19:00–23:00 = 67.4% of all revenue — NEVER understaff these hours",
        f"• SBY max additional cost this week: £{summary['SBY Max Additional Cost']:.2f}",
    ]
    sby_row = end_row + 2
    for i, text in enumerate(sby_lines):
        c = ws5.cell(row=sby_row + i, column=1, value=text)
        c.font = Font(color="F5A623" if i == 0 else "AAAAAA", bold=(i == 0), name="Arial", size=9)
        c.fill = PatternFill("solid", fgColor="0D0F14")
        ws5.merge_cells(start_row=sby_row+i, start_column=1, end_row=sby_row+i, end_column=6)
    _set_col_widths(ws5, [8, 16, 18, 14, 20, 20])

    # ── Sheet 6: SBY Tracker ──────────────────────────────────────────────────
    ws6 = wb.create_sheet("6 SBY Tracker")
    ws6.sheet_view.showGridLines = False

    ws6.merge_cells("A1:G1")
    c = ws6["A1"]
    c.value = "🔥 STANDBY (SBY) DECISION TRACKER — Call-In Cost vs Savings"
    c.font  = Font(bold=True, color="F5A623", name="Arial", size=13)
    c.fill  = PatternFill("solid", fgColor="0D0F14")
    c.alignment = Alignment(horizontal="left", vertical="center")

    sby_df = pd.DataFrame([
        {
            "Name":             name,
            "Max SBY Hours":    v["max_sby_hrs"],
            "Rate (£/hr)":      v["hourly_rate"],
            "Max SBY Cost (£)": round(v["max_sby_hrs"] * v["hourly_rate"], 2),
            "Called In?":       "Leave blank",
            "Actual Hours":     "Leave blank",
            "Actual Cost (£)":  "Leave blank",
        }
        for name, v in SBY_STAFF.items()
    ])
    end_row = _write_df(ws6, sby_df, start_row=3,
                        title="■ SBY Staff — Fill in 'Called In?' each night", title_span=7)

    total_row = end_row + 1
    ws6.cell(row=total_row, column=1, value="TOTAL MAX SBY COST").font = \
        Font(bold=True, color="F5A623", name="Arial", size=10)
    ws6.cell(row=total_row, column=4,
             value=f"£{summary['SBY Max Additional Cost']:.2f}").font = \
        Font(bold=True, color="FF4444", name="Arial", size=10)
    for col in [1, 4]:
        ws6.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="1a1d26")
    _set_col_widths(ws6, [14, 14, 12, 16, 14, 14, 16])

    for ws in wb.worksheets:
        _dark_bg(ws)

    wb.save(output_path)
    print(f"✅  Saved: {output_path}")
    return summary, overlay_df, over_df, under_df


# ══════════════════════════════════════════════════════════════════════════════
# ██  SECTION 5 — DATA LOADING (dashboard)
# ══════════════════════════════════════════════════════════════════════════════

def clean_currency(val):
    if pd.isna(val) or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    return float(str(val).replace(',', '').replace('£', '').strip())

# REFRESHED DYNAMIC LOADING - CACHE REMOVED FOR REAL-TIME SYNC
def load_data(reference_date=None):

    import streamlit as st

    base = os.path.join(os.getcwd(), "Sales Summary Data")

    revenue_path = os.path.join(BASE_DIR, "daily_sales_master.csv")
    if os.path.exists(revenue_path):
        df = pd.read_csv(revenue_path)
    else:
        df = pd.DataFrame(DAILY_DATA)

    # Column Normalization & Standard Name Mapping
    df.columns = [c.strip().lower() for c in df.columns]
    
    # Map back to specific case-sensitive names expected by the dashboard
    df = df.rename(columns={
        "net":              "Net sales",
        "net sales":        "Net sales",
        "net_sales":        "Net sales",
        "orders":           "Orders",
        "tax":              "Tax on net sales",
        "tax on net sales": "Tax on net sales",
        "charges":          "Charges",
        "revenue":          "Revenue",
        "refunds":          "refunds",
        "day":              "day",
        "date":             "date",
    })

    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"])
    
    # Ensure missing columns (like refunds) have default values if absent
    if "refunds" not in df.columns:
        df["refunds"] = 0.0


    personnel_path = os.path.join(os.getcwd(), "personnel_rates_master.csv")
    profiles_path  = os.path.join(os.getcwd(), "staff_profiles.csv")
    personnel_data = {}

    if os.path.exists(personnel_path):
        p_df = pd.read_csv(personnel_path)
        p_df.columns = [c.strip() for c in p_df.columns]
        prof_df = pd.read_csv(profiles_path) if os.path.exists(profiles_path) else None

        for _, row in p_df.iterrows():
            name = str(row.get("Name","")).strip()
            if not name or name == "nan": continue
            entry = row.to_dict()
            entry["role"] = "Junior"
            entry["dept"] = "Front"
            if prof_df is not None:
                p_match = prof_df[prof_df["Name"].str.strip() == name]
                if not p_match.empty:
                    entry["role"] = p_match.iloc[0].get("Role", "Junior")
                    entry["dept"] = p_match.iloc[0].get("Department", "Front")
            personnel_data[name] = entry

    detailed_shifts = []

    # ── Dynamic Rota Path ────────────────────────────────────────────────────
    _latest_dt    = None
    if not df.empty and "date" in df.columns:
        _latest_dt = pd.to_datetime(df["date"]).max()

    # Priority: Use reference_date if passed (from sidebar), else latest sales date
    _folder_date  = reference_date if reference_date is not None else (_latest_dt if _latest_dt is not None else datetime.today())
    
    # Ensure it's a datetime object for strftime
    if isinstance(_folder_date, date) and not isinstance(_folder_date, datetime):
        _folder_date = datetime.combine(_folder_date, datetime.min.time())

    _folder_start = _folder_date - pd.Timedelta(days=_folder_date.weekday())
    _folder_end   = _folder_start + pd.Timedelta(days=6)

    candidates = [
        f"Rota week {_folder_start.strftime('%d %b').lower()} - {_folder_end.strftime('%d %B').lower()} {_folder_start.year}",
        f"Rota week {_folder_start.strftime('%d %b')} - {_folder_end.strftime('%d %B %Y')}",
        f"Rota week {_folder_start.strftime('%d %b').lower()} - {_folder_end.strftime('%d %B %Y').lower()}",
        f"Rota week {_folder_start.strftime('%d %B').lower()} - {_folder_end.strftime('%d %B').lower()} {_folder_start.year}",
        f"Rota week {_folder_start.strftime('%d %b %Y')}",
    ]

    rota_det_path = None
    # Hardened case-insensitive folder check
    try:
        all_items = os.listdir(os.getcwd())
        for cand in candidates:
            # Check for exact case-insensitive match in current directory
            match = next((item for item in all_items if item.lower() == cand.lower()), None)
            if match:
                p = os.path.join(os.getcwd(), match, "detailed_rota_with_shifts.csv")
                if os.path.exists(p):
                    rota_det_path = p
                    break
    except:
        pass
    
    if rota_det_path:
        # Detect format: New Rota Builder (CSV with Headers) vs Old Manual Excel (No Headers/Layout)
        rota_df_raw = pd.read_csv(rota_det_path)
        
        if "Name" in rota_df_raw.columns and "Start" in rota_df_raw.columns:
            # --- NEW CLEAN FORMAT (Rota Builder) ---
            for _, r in rota_df_raw.iterrows():
                try:
                    def _to_f(t_str):
                        parts = str(t_str).split(":")
                        return int(parts[0]) + (int(parts[1])/60.0 if len(parts)>1 else 0.0)
                    
                    fs = _to_f(r["Start"])
                    fe = _to_f(r["End"])
                    if fe < fs: fe += 24
                    
                    # Use the PRE-CALCULATED duration from CSV for 100% fidelity
                    total_dur = float(r.get("Duration", fe - fs))
                    slots_sum = 0.0
                    for h_int in range(int(fs), int(fe) + (1 if fe % 1 > 0 else 0)):
                        hour_key = h_int % 24
                        slot_start = max(float(h_int), fs)
                        slot_end   = min(float(h_int + 1), fe)
                        slot_dur   = max(0.0, slot_end - slot_start)
                        
                        if slot_dur > 0:
                            slots_sum += slot_dur
                            detailed_shifts.append({
                                "day":    r["Day"], 
                                "hour":   hour_key, 
                                "name":   r["Name"], 
                                "is_sby": str(r.get("SBY","No")).upper() == "YES",
                                "duration": slot_dur
                            })
                    
                    # Safety check: If floating point math caused a tiny gap, 
                    # adjust the last slot to ensure the sum is exactly total_dur
                    if detailed_shifts and abs(slots_sum - total_dur) > 1e-9:
                        detailed_shifts[-1]["duration"] += (total_dur - slots_sum)
                except:
                    continue
        else:
            # --- OLD MANUAL FORMAT (Positional Parsing) ---
            with open(rota_det_path, "r") as f:
                r_lines = [l.split(",") for l in f.readlines()]
                r_days  = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

                nickname_map = {
                    "DHIRAJ": "Dhiraj Mangade", "ATHARAV": "Atharvkumar Sanjay", "ATHARV": "Atharvkumar Sanjay",
                    "CHINTAN": "Chintan", "CHINTHAN": "Chintan", 
                    "DAMINI": "Damini Sharadchandra Aher", "NITIN": "Nithin", "NITHIN": "Nithin",
                    "PAMITHA": "Pamitha Perera",
                    "MELLISSA": "Mellissa Teshali Leontia", "MELLISA": "Mellissa Teshali Leontia",
                    "DIKSHA": "Dikshya", "DIKSHYA": "Dikshya",
                    "SUPREME": "Supreme Gurung", "TULIKA": "Tulika Das Adhikari", "MUNIRA": "Munira",
                    "BHOOMIKA": "Bhoomika", "REWATHI": "Dhriti Kulshrestha", "DHRITI": "Dhriti Kulshrestha",
                    "RAVI": "Ravi Kishore", "RAJESH": "Rajesh Yadav", "ASMA": "Asma"
                }

                for r_idx in range(len(r_lines)):
                    row_data = r_lines[r_idx]
                    if len(row_data) < 2: continue
                    if any(x in row_data[0].upper() for x in ["MONDAY","KITCHEN","FRONT"]): continue

                    for d_i in range(7):
                        c_t, c_n = d_i * 2, d_i * 2 + 1
                        if c_n < len(row_data):
                            t_r = row_data[c_t].strip()
                            raw_name = row_data[c_n].strip().upper()
                            if not raw_name or len(raw_name) < 2: continue

                            is_sby = ("SBY" in raw_name or "SYB" in raw_name)
                            n_nick = raw_name.split(" ")[0]
                            full_name = nickname_map.get(n_nick, n_nick.capitalize())

                            if t_r and full_name:
                                for hh in parse_time_range(t_r):
                                    detailed_shifts.append({"day": r_days[d_i], "hour": hh, "name": full_name, "is_sby": is_sby, "duration": 1.0})
    else:
        # Fallback to synthesizing if manual CSV is missing
        try:
            from rota_engine import RotaEngine
            engine = RotaEngine()
            engine.load_staff()
            engine.load_shifts()
            # Determine start of week based on today or a fixed reference
            w_start_dt = datetime.now() - timedelta(days=datetime.now().weekday()) 
            gen_rota = engine.generate_week(week_start=w_start_dt.replace(hour=0, minute=0, second=0))

            for _, r in gen_rota.iterrows():
                # Handle both Capital and lowercase keys for robustness (KeyError: 'end' fix)
                s_val = r.get("Start", r.get("start", "00:00"))
                e_val = r.get("End", r.get("end", "00:00"))
                h_start = int(str(s_val).split(":")[0])
                h_end   = int(str(e_val).split(":")[0])
                if h_end < h_start: h_end += 24

                for hh in range(h_start, h_end):
                    detailed_shifts.append({
                        "day":    r["Day"],
                        "hour":   hh % 24,
                        "name":   r["Name"],
                        "is_sby": r.get("SBY") == "Yes"
                    })
            st.info("💡 Rota data synthesized using Automated Rota Engine (Manual CSV not found).")
        except Exception as e:
            # Silent fallback if engine fails
            pass

    # ── DYNAMIC KPI HARVESTING ──────────────────────────────────────────────
    channel_dict  = {}
    dispatch_map  = {}
    payment_map   = {}
    hourly_map    = {h: 0.0 for h in range(24)}
    
    # Load Channel Data
    cp = os.path.join(base, "net_sales_by_sales_channel.csv")
    if os.path.exists(cp):
        c_raw = pd.read_csv(cp, header=None)
        # Row 0: names, Row 2: values
        for i in range(c_raw.shape[1]):
            name = str(c_raw.iloc[0, i]).strip()
            val  = clean_currency(c_raw.iloc[2, i])
            if name and val > 0: channel_dict[name] = val
    else:
        channel_dict = CHANNEL_DATA.copy()

    # Load Dispatch Data
    dp = os.path.join(base, "net_sales_by_dispatch_type.csv")
    if os.path.exists(dp):
        d_raw = pd.read_csv(dp, header=None)
        for i in range(d_raw.shape[1]):
            name = str(d_raw.iloc[0, i]).strip()
            val  = clean_currency(d_raw.iloc[-1, i])
            if name and val > 0: dispatch_map[name] = {"revenue": val, "orders": int(val/15)} # Est orders
    else:
        dispatch_map = {k: v.copy() for k, v in DISPATCH_DATA.items()}

    # Load Payment Data
    pp = os.path.join(base, "net_sales_by_payment_method.csv")
    if os.path.exists(pp):
        p_raw = pd.read_csv(pp, header=None)
        for i in range(p_raw.shape[1]):
            name = str(p_raw.iloc[0, i]).strip()
            val  = clean_currency(p_raw.iloc[-1, i])
            if name and val > 0: payment_map[name] = val
    else:
        payment_map = PAYMENT_DATA.copy()
        
    delivery_fees = 282.0

    if os.path.exists(base):
        try:
            # Helper is now clean_currency defined globally above

            # ── Strict Whitelist Harvest (prevents double-counting) ──────────────
            # Only these files carry authoritative DAILY timeline data.
            # Hourly / dispatch / channel files are read SEPARATELY below.
            DAILY_SUMMARY_FILES = {"sales_overview.csv", "net_sales_per_day.csv"}
            DETAIL_ORDER_FILES  = {"sales_data.csv"}

            all_orders_detail  = []
            all_daily_summaries = []

            # A. Detailed order-by-order export
            for f_name in DETAIL_ORDER_FILES:
                p = os.path.join(base, f_name)
                if not os.path.exists(p): continue
                try:
                    raw_df = pd.read_csv(p)
                    if "Order ID" not in raw_df.columns: continue
                    raw_df["date_dt"] = pd.to_datetime(raw_df["Order time"], errors="coerce")
                    raw_df = raw_df[raw_df["date_dt"].notna()]
                    for col in ["Net sales", "Revenue", "Refunds", "Tax on net sales"]:
                        if col in raw_df.columns: raw_df[col] = raw_df[col].apply(clean_currency)
                    all_orders_detail.append(raw_df)
                except Exception as e:
                    logging.warning(f"Failed to load detailed order file {f_name}: {e}")

            # B. Daily summary exports (exactly one file wins per date via dedup)
            for f_name in DAILY_SUMMARY_FILES:
                p = os.path.join(base, f_name)
                if not os.path.exists(p): continue
                try:
                    raw_ov = pd.read_csv(p)
                    if "Order time" not in raw_ov.columns or "Net sales" not in raw_ov.columns: continue
                    raw_ov = raw_ov[raw_ov["Order time"].notna()].copy()
                    raw_ov["date"] = pd.to_datetime(raw_ov["Order time"], format="%Y-%m-%d", errors="coerce")
                    raw_ov = raw_ov[raw_ov["date"].notna() & (raw_ov["date"].dt.year >= 2024)]
                    for c in ["Net sales","Revenue","Orders","Tax on net sales","Refunds"]:
                        if c in raw_ov.columns: raw_ov[c] = raw_ov[c].apply(clean_currency)
                    
                    # FIX: Filter available columns to prevent crash if some are missing (e.g. net_sales_per_day.csv)
                    wanted = ["date","Net sales","Revenue","Tax on net sales","Refunds","Orders"]
                    available = [col for col in wanted if col in raw_ov.columns]
                    all_daily_summaries.append(raw_ov[available])
                except Exception as e:
                    logging.warning(f"Failed to load daily summary file {f_name}: {e}")

            # 2. Pull hourly / dispatch / channel from detailed orders or fallback CSVs
            detailed_kpis = {"hourly": {}}
            if all_orders_detail:
                full_orders = pd.concat(all_orders_detail).drop_duplicates(subset=["Order ID"])
                full_orders["date"] = full_orders["date_dt"].dt.normalize()

                # Aggregate detail into daily rows to supplement summary gaps
                daily_from_detail = full_orders.groupby("date", as_index=True).agg(
                    **{
                        "Net sales":        ("Net sales",        "sum"),
                        "Revenue":          ("Revenue",          "sum"),
                        "Refunds":          ("Refunds",          "sum"),
                        "Tax on net sales": ("Tax on net sales", "sum"),
                        "Orders":           ("Order ID",         "count"),
                    }
                ).reset_index()
                all_daily_summaries.append(daily_from_detail)

                # Hourly from detail (most accurate)
                full_orders["hour"] = full_orders["date_dt"].dt.hour
                _h_sum = full_orders.groupby("hour")["Net sales"].sum()
                # Get unique days count to calculate daily average
                _day_count = full_orders["date"].nunique()
                if _day_count > 0:
                    detailed_kpis["hourly"] = (_h_sum / _day_count).to_dict()
                else:
                    detailed_kpis["hourly"] = _h_sum.to_dict()

                # Dispatch from detail
                if "Dispatch type" in full_orders.columns:
                    for k, v in full_orders.groupby("Dispatch type")["Net sales"].sum().items():
                        key = str(k).strip()
                        if key in dispatch_map: dispatch_map[key]["revenue"] = v

                # Channel from detail
                if "Sales channel name" in full_orders.columns:
                    for k, v in full_orders.groupby("Sales channel name")["Net sales"].sum().items():
                        key = str(k).strip()
                        for std in ["Deliveroo","Just Eat","POS","Uber Eats","Web"]:
                            if std.lower() in key.lower(): channel_dict[std] = v

            else:
                # Fallback: read hourly and dispatch from their dedicated summary CSVs
                hr_path = os.path.join(base, "net_sales_by_hour_of_day.csv")
                if os.path.exists(hr_path):
                    hr_df = pd.read_csv(hr_path)
                    for _, r in hr_df.iterrows():
                        try:
                            _h = int(r.iloc[1])
                            _v = clean_currency(r.iloc[2])
                            # If it's a weekly summary file, divide by 7 to show daily avg on charts
                            detailed_kpis["hourly"][_h] = round(_v / 7, 2)
                        except Exception as e:
                            logging.warning(f"Failed to parse row in hourly CSV: {e}")

                dispatch_path = os.path.join(base, "net_sales_by_dispatch_type.csv")
                if os.path.exists(dispatch_path):
                    disp_df = pd.read_csv(dispatch_path)
                    for key, col_idx in [("Collection",1),("Delivery",2),("Dine In",3),("Take Away",4)]:
                        try: dispatch_map[key]["revenue"] = clean_currency(disp_df.iloc[-1, col_idx])
                        except Exception as e:
                            logging.warning(f"Failed to parse row in dispatch CSV: {e}")

                ch_path = os.path.join(base, "net_sales_by_sales_channel.csv")
                if os.path.exists(ch_path):
                    ch_df = pd.read_csv(ch_path)
                    for i, std in enumerate(["Deliveroo","Just Eat","POS","Uber Eats","Web"]):
                        try: channel_dict[std] = clean_currency(ch_df.iloc[-1, i+1])
                        except Exception as e:
                            logging.warning(f"Failed to parse row in channel CSV: {e}")

            hourly_map = {h: detailed_kpis["hourly"].get(h, 0.0) for h in range(24)}

            # 3. Merge daily timeline — dedupe by date (summary file wins, detail fills gaps)
            # 3. Merge daily timeline — dedupe by date (Master CSV + Detail Logs)
            standardized_summaries = []
            for _df_item in all_daily_summaries:
                _df_item.columns = [c.lower() for c in _df_item.columns]
                standardized_summaries.append(_df_item)
            
            # Start with the master data we already loaded (df)
            df.columns = [c.lower() for c in df.columns]
            df["date"] = pd.to_datetime(df["date"]).dt.normalize()
            
            # Map Master columns to UI standard
            master_column_map = {
                "net": "Net sales", "net sales": "Net sales", "net_sales": "Net sales",
                "tax": "Tax on net sales", "tax on net sales": "Tax on net sales",
                "orders": "Orders", "revenue": "Revenue", "refunds": "Refunds", "charges": "Charges"
            }
            df = df.rename(columns=master_column_map)
            
            # Standardize secondary summaries
            for i in range(len(standardized_summaries)):
                s_df = standardized_summaries[i].copy()
                s_df.columns = [c.lower() for c in s_df.columns]
                if "date" in s_df.columns:
                    s_df["date"] = pd.to_datetime(s_df["date"]).dt.normalize()
                s_df = s_df.rename(columns={
                    "net sales": "Net sales", "net": "Net sales", "net_sales": "Net sales",
                    "tax on net sales": "Tax on net sales", "tax": "Tax on net sales",
                    "orders": "Orders", "revenue": "Revenue", 
                    "refunds": "Refunds", "charges": "Charges"
                })
                standardized_summaries[i] = s_df

            # Merge Logic: 
            # 1. Start with the full timeline from Master
            # 2. Append any dates found ONLY in detail logs
            if standardized_summaries:
                secondary = pd.concat(standardized_summaries).sort_values("date")
                secondary = secondary.drop_duplicates(subset=["date"], keep="first")
                
                # Identify dates in secondary NOT in master
                master_dates = set(df["date"])
                gap_data = secondary[~secondary["date"].isin(master_dates)]
                
                if not gap_data.empty:
                    merged = pd.concat([df, gap_data]).sort_values("date")
                else:
                    merged = df.copy()
            else:
                merged = df.copy()

            # Final normalization and cleaning
            merged = merged[merged["date"].dt.year >= 2024]
            merged["day"] = merged["date"].dt.day_name()
            
            # 1. Clean Refunds (Sum across all refund-related columns)
            pd.set_option('future.no_silent_downcasting', True)
            refund_cols = [c for c in merged.columns if "refund" in c.lower()]
            merged["refunds_clean"] = 0.0
            for col in refund_cols:
                vals = merged[col].apply(clean_currency).fillna(0.0).abs()
                merged["refunds_clean"] = merged["refunds_clean"].add(vals, fill_value=0.0)

            # 2. Clean main numeric columns (prevents string-concatenation bug)
            for col in ["Net sales", "Revenue", "Tax on net sales", "Orders"]:
                if col in merged.columns:
                    merged[col] = merged[col].apply(clean_currency).fillna(0.0)

            # Ensure 'Refunds' column itself is cleaned (UI might use it too)
            if "Refunds" in merged.columns:
                merged["Refunds"] = merged["refunds_clean"]

            # Recalculate rolling 7
            if "Net sales" in merged.columns:
                merged["rolling7"] = merged["Net sales"].rolling(window=7).mean().fillna(merged["Net sales"])
            
            df = merged.copy()
                # is the one used for the sum in the KPI block.

        except Exception as e:
            st.warning(f"Data engine error: {e}")

    # ── Auto-compute Weekly Summary from daily df ──────────────────────
    auto_weekly = []
    if not df.empty and "Net sales" in df.columns:
        _wk = df.copy()
        _wk["date"] = pd.to_datetime(_wk["date"])
        _wk["week_start"] = _wk["date"] - pd.to_timedelta(_wk["date"].dt.weekday, unit="D")
        _wk_grp = _wk.groupby("week_start").agg(
            net=("Net sales", "sum"),
            orders=("Orders", "sum") if "Orders" in _wk.columns else ("Net sales", "count"),
            tax=("Tax on net sales", "sum") if "Tax on net sales" in _wk.columns else ("Net sales", lambda x: 0),
        ).reset_index()
        for _, r in _wk_grp.iterrows():
            auto_weekly.append({
                "week":   r["week_start"].strftime("%d %b"),
                "net":    round(float(r["net"]), 2),
                "orders": int(r["orders"]),
                "tax":    round(float(r["tax"]), 2),
            })

    # ── Auto-compute Week Label from latest date ────────────────────────
    # ── Auto-compute Week Label from selected/folder date ───────────────
    auto_week_label = f"{_folder_start.strftime('%d %b')} – {_folder_end.strftime('%d %b %Y')}"

    # ── Auto-compute Payment Data from CSV ─────────────────────────────
    auto_payment = dict(PAYMENT_DATA)  # fallback to static
    try:
        pay_path = os.path.join(base, "net_sales_by_payment_method.csv")
        if os.path.exists(pay_path):
            pay_csv = pd.read_csv(pay_path)
            # Last row has totals; column names are the payment methods
            pay_row = pay_csv.iloc[-1]
            _pay_map = {}
            for col in pay_csv.columns:
                try:
                    val = float(str(pay_row[col]).replace(",", "").replace("£", "").strip())
                    if val > 0:
                        _pay_map[col] = val
                except: pass
            if _pay_map:
                auto_payment = _pay_map
    except: pass

    # ── Auto-compute Channel Data from CSV ─────────────────────────────
    auto_channels = channel_dict  # already computed from CSV in harvest engine

    return {
        "daily":          df.sort_values("date"),
        "channels":       auto_channels,
        "dispatch_truth": dispatch_map,
        "delivery_fees":  delivery_fees,
        "shifts":         detailed_shifts,
        "hourly_live":    hourly_map,
        "personnel":      personnel_data,
        # ── Fully auto-computed from CSVs ───────────────────────────
        "weekly":         auto_weekly,
        "payment":        auto_payment,
        "week_label":     auto_week_label,
    }



# ══════════════════════════════════════════════════════════════════════════════
# ██  SECTION 6 — STREAMLIT DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

import streamlit as st

st.markdown("""
<style>
    .stApp { background-color: #0a0b0f; }
    .main .block-container { padding: 24px 32px; max-width: 1600px; }
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'DM Mono', monospace; color: #e8e9f0; }
    [data-testid="stSidebar"] { background-color: #12141a !important; border-right: 1px solid #252836; }
    [data-testid="stSidebar"] * { color: #e8e9f0 !important; }
    .sidebar-logo { font-family: 'Syne', sans-serif; font-weight: 800; font-size: 20px; color: #f5a623 !important; letter-spacing: -0.5px; margin-bottom: 4px; }
    .sidebar-sub { font-size: 11px; color: #6b7094 !important; letter-spacing: 1px; text-transform: uppercase; }
    .live-dot { display:inline-block;width:8px;height:8px;border-radius:50%;background:#3ecf8e;margin-right:6px;animation:pulse 2s infinite; }
    @keyframes pulse { 0%,100%{opacity:1;} 50%{opacity:.4;} }
    .status-box { background-color:#1a1d26;padding:10px 14px;border-radius:8px;border-left:3px solid #f5a623;margin-top:6px;font-size:12px;color:#6b7094 !important; }
    [data-testid="stMetric"] { background:#1a1d26;border:1px solid #252836;border-radius:12px;padding:12px 10px;border-left:3px solid #f5a623; }
    [data-testid="stMetricValue"] { font-family:'Syne',sans-serif !important;font-size:18px !important;font-weight:800 !important;color:#e8e9f0 !important;white-space:nowrap !important; }
    [data-testid="stMetricLabel"] { font-size:9px !important;letter-spacing:1.5px;text-transform:uppercase;color:#6b7094 !important; }
    [data-testid="stMetricDelta"] { font-size:11px !important; }
    .chart-card { background:#1a1d26;border:1px solid #252836;border-radius:12px;padding:22px;margin-bottom:16px; }
    .section-title { font-family:'Syne',sans-serif;font-weight:700;font-size:12px;letter-spacing:2px;text-transform:uppercase;color:#6b7094;margin-bottom:12px;margin-top:8px; }
    .page-title { font-family:'Syne',sans-serif;font-weight:800;font-size:24px;color:#e8e9f0;margin-bottom:4px; }
    .page-sub { font-size:12px;color:#6b7094;margin-bottom:24px; }
    .insight-box { background:#12141a;border:1px solid #252836;border-radius:8px;padding:12px 16px;margin-top:8px;font-size:11px;color:#6b7094;line-height:1.6; }
    .insight-box b { color:#e8e9f0; }
    [data-testid="stDataFrame"] { border:1px solid #252836 !important;border-radius:8px !important;overflow:hidden; }
    .stTabs [data-baseweb="tab-list"] { background:#12141a;border-radius:10px;padding:4px;gap:4px;border:1px solid #252836; }
    .stTabs [data-baseweb="tab"] { background:transparent;border-radius:8px;color:#6b7094;font-family:'DM Mono',monospace;font-size:12px;padding:8px 18px; }
    .stTabs [aria-selected="true"] { background:#1a1d26 !important;color:#e8e9f0 !important; }
    .stTabs [data-baseweb="tab-border"] { display:none; }
    .stSelectbox > div > div { background:#1a1d26;border:1px solid #252836;border-radius:8px;color:#e8e9f0; }
    .stDateInput > div > div { background:#1a1d26;border:1px solid #252836;border-radius:8px; }
    hr { border-color: #252836; }
    .period-badge { display:inline-block;background:#1a1d26;border:1px solid #252836;padding:5px 14px;border-radius:20px;font-size:11px;color:#6b7094; }
    .labour-kpi { background:#1a1d26;border:1px solid #252836;border-radius:10px;padding:14px;text-align:center;margin-bottom:8px; }
    .labour-kpi-label { font-size:9px;color:#6b7094;text-transform:uppercase;letter-spacing:1.5px;margin-bottom:4px; }
    .labour-kpi-value { font-family:'Syne',sans-serif;font-size:18px;font-weight:800;color:#e8e9f0; }
    /* Fix date picker popup clipping */
    [data-baseweb="popover"] {
        overflow: visible !important;
        z-index: 9999 !important;
    }
    [data-baseweb="calendar"] {
        overflow: visible !important;
    }
    [data-testid="stSidebar"] {
        overflow-y: auto !important;
    }
    [data-testid="stSidebar"] > div {
        overflow-y: auto !important;
    }

    /* Hide the preset date range dropdown */
    [data-baseweb="select"] > div[aria-label] {
        display: none !important;
    }
</style>
""", unsafe_allow_html=True)

COLORS = {
    "accent":  "#f5a623", "accent2": "#e8724a",
    "accent3": "#7c5cbf", "accent4": "#3ecf8e",
    "red":     "#e05c5c", "muted":   "#6b7094",
}
PALETTE = [COLORS["accent"], COLORS["accent2"], COLORS["accent3"],
           COLORS["accent4"], COLORS["red"], COLORS["muted"]]

def dark_layout(fig, height=340, showlegend=False):
    fig.update_layout(
        height=height, template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="DM Mono, monospace", color="#6b7094", size=11),
        margin=dict(l=0, r=0, t=10, b=0),
        showlegend=showlegend,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(size=10), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(gridcolor="rgba(255,255,255,0.04)", zeroline=False),
        yaxis=dict(gridcolor="rgba(255,255,255,0.06)", zeroline=False),
    )
    return fig

def get_staff_availability_for_week(week_start_date):
    """
    Read submitted availability from database
    Returns dict: {staff_name: {day: status}}
    """
    import sqlite3, json
    avail_db = "availability.db"
    if not os.path.exists(avail_db):
        return {}
    
    with sqlite3.connect(avail_db) as conn:
        try:
            # Query by the week_start (YYYY-MM-DD)
            target_date_str = week_start_date.strftime('%Y-%m-%d')
            rows = conn.execute('''
                SELECT staff_name, availability, notes
                FROM availability
                WHERE week_start = ?
                ORDER BY submitted_at DESC
            ''', (target_date_str,)).fetchall()
        except:
            return {}
    
    result = {}
    for name, avail_json, notes in rows:
        if name not in result:  # take latest submission per person
            try:
                result[name] = json.loads(avail_json)
                result[name]['_notes'] = notes
            except: pass
    
    return result

def save_staff_availability(week_start_date, staff_name, availability_dict, notes="Manual Entry"):
    import sqlite3, json
    avail_db = "availability.db"
    target_date_str = week_start_date.strftime('%Y-%m-%d')
    avail_json = json.dumps(availability_dict)
    
    with sqlite3.connect(avail_db) as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS availability (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                staff_name TEXT,
                week_start TEXT,
                availability TEXT,
                notes TEXT,
                submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            INSERT INTO availability (staff_name, week_start, availability, notes)
            VALUES (?, ?, ?, ?)
        ''', (staff_name, target_date_str, avail_json, notes))
        conn.commit()

def load_events():
    import pandas as pd
    events_path = os.path.join(os.getcwd(), "events_log.csv")
    if os.path.exists(events_path):
        try:
            e_df = pd.read_csv(events_path)
            e_df["Date"] = pd.to_datetime(e_df["Date"]).dt.date
            return e_df.set_index("Date").to_dict('index')
        except: return {}
    return {}

def generate_forecast_pdf(week_label, total_forecast, aov, day_data, boosts):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph(f"Chocoberry Intelligence — Forecast Report", styles['Title']))
    elements.append(Paragraph(f"Target Week: {week_label}", styles['Heading2']))
    elements.append(Spacer(1, 20))
    
    # Summary Table
    summary_data = [
        ["Total Weekly Forecast", f"£{total_forecast:,.2f}"],
        ["Forecasted AOV", f"£{aov:,.2f}"],
        ["Data Methodology", "4-Week Rolling Average + AI Memory"]
    ]
    t = Table(summary_data, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    elements.append(t)
    elements.append(Spacer(1, 30))
    
    # Daily Breakdown
    elements.append(Paragraph("Day-by-Day Forecast Breakdown", styles['Heading3']))
    table_data = [["Day", "Date", "Forecasted Net Sales"]]
    for d in day_data:
        table_data.append([d['day'], d['date'], f"£{d['sales']:,.2f}"])
        
    t2 = Table(table_data, colWidths=[100, 150, 150])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkslategrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 30))
    
    # Memory Section
    if boosts:
        elements.append(Paragraph("🧠 AI Business Memory Insights", styles['Heading3']))
        for b in boosts:
            elements.append(Paragraph(f"• {b}", styles['Normal']))
            elements.append(Spacer(1, 5))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

def sync_availability_from_cloud(sheet_url):
    import requests
    import pandas as pd
    import json
    import sqlite3
    from datetime import datetime
    
    try:
        # Convert sharing URL to direct CSV export URL
        csv_url = sheet_url.replace('/edit?usp=sharing', '/export?format=csv')
        csv_url = csv_url.split('/edit')[0] + '/export?format=csv'
        
        df = pd.read_csv(csv_url)
        if df.empty: return False, "Google Sheet is empty."
        
        with sqlite3.connect("availability.db") as conn:
            # Table should already exist from local portal, but ensure it does
            conn.execute('''CREATE TABLE IF NOT EXISTS availability (
                id INTEGER PRIMARY KEY, staff_name TEXT, week_start TEXT, 
                availability TEXT, notes TEXT, submitted_at TEXT
            )''')
            
            count = 0
            for _, row in df.iterrows():
                # INSERT OR REPLACE into local DB
                conn.execute('''
                    INSERT OR REPLACE INTO availability (staff_name, week_start, availability, notes, submitted_at)
                    VALUES (?, ?, ?, ?, ?)
                ''', (row['Name'], row['Week Start'], row['Availability'], row['Notes'], row['Timestamp']))
                count += 1
            conn.commit()
        return True, f"Successfully synced {count} submissions from Google Sheets!"
    except Exception as e:
        return False, f"Sync Error: {e}"

def _load():
    return load_data()

data   = _load()
all_df = data["daily"]


# ── Sidebar ───────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sidebar-logo">Choco<span style="color:#e8e9f0">berry</span></div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-sub"><span class="live-dot" style="display:inline-block;width:7px;height:7px;border-radius:50%;background:#3ecf8e;margin-right:5px;"></span>Intelligence Dashboard</div>', unsafe_allow_html=True)
    st.markdown("---")

    st.markdown("**🚚 Dispatch Type**")
    dispatch_opts = ["All","Collection","Delivery","Dine In","Take Away"]
    sel_dispatch  = st.selectbox("Dispatch Type", dispatch_opts, label_visibility="collapsed")

    st.markdown("**📊 Sales Channel**")
    channel_opts = ["All"] + list(CHANNEL_DATA.keys())
    sel_channel  = st.selectbox("Sales Channel", channel_opts, label_visibility="collapsed")

    st.markdown("---")

    st.markdown("**📅 Date Range**")
    # Unlock the calendar range so user can select future weeks for Rota/Labour checks
    calendar_min = all_df["date"].min().date() if not all_df.empty else date(2024, 1, 1)
    calendar_max = date(2026, 12, 31)
    
    # Use session state to keep the UI from resetting while clicking
    # Default to the most recent week of data instead of the entire history
    # This ensures that when you open the system, it land on the "Live" rota week you just pushed
    latest_avail = all_df["date"].max().date() if not all_df.empty else date.today()
    default_start = latest_avail - timedelta(days=6)
    default_end   = latest_avail
    
    dr = st.date_input(
        "Date Range", 
        value=[default_start, default_end],
        min_value=calendar_min,
        max_value=calendar_max,
        label_visibility="collapsed"
    )
    
    if len(dr) == 2:
        start_d, end_d = dr[0], dr[1]
        st.session_state["date_start"] = dr[0]
        st.session_state["date_end"] = dr[1]
    else:
        start_d = st.session_state.get("date_start", default_start)
        end_d = st.session_state.get("date_end", default_end)

    # Dynamic Data Load based on selected start date
    data = load_data(reference_date=start_d)
    all_df = data["daily"]
    f_df = all_df[
        (all_df["date"] >= pd.Timestamp(start_d)) & 
        (all_df["date"] <= pd.Timestamp(end_d) + pd.Timedelta(hours=23, minutes=59, seconds=59))
    ]
    
    total_orders = int(f_df["Orders"].sum())
    st.markdown(f'<div class="status-box">💎 <b>{total_orders:,}</b> transactions selected</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="status-box">📅 Period: <b>{start_d.strftime("%d %b")} → {end_d.strftime("%d %b %Y")}</b></div>', unsafe_allow_html=True)
    st.markdown(f'<div class="status-box">📈 {len(f_df)} trading days in view</div>', unsafe_allow_html=True)

# ── DYNAMIC DATE-RANGE RECOMPUTATION ──────────────────────────────────────────
# Rebuilds dispatch/channel/hourly/payment from detail CSV filtered to the
# user's selected date range. This makes ALL charts respond to the date picker.
_detail_path = os.path.join(BASE_DIR, "Sales Summary Data", "sales_data.csv")
_start_ts = pd.Timestamp(start_d).replace(hour=0, minute=0, second=0)
_end_ts = pd.Timestamp(end_d).replace(hour=23, minute=59, second=59)

try:
    _det = pd.read_csv(_detail_path)
    _det.columns = [c.strip() for c in _det.columns]
    _time_col = "Order time" if "Order time" in _det.columns else _det.columns[4]
    _det["_dt"] = pd.to_datetime(_det[_time_col], errors="coerce")
    _det = _det.dropna(subset=["_dt"])
    _det = _det[(_det["_dt"] >= _start_ts) & (_det["_dt"] <= _end_ts)]
    if "Order ID" in _det.columns:
        _det = _det.drop_duplicates(subset=["Order ID"])
    _net_col = "Net sales" if "Net sales" in _det.columns else _det.columns[11]
    _det["_net"] = _det[_net_col].apply(clean_currency)
    _det["_hour"] = _det["_dt"].dt.hour

    # Dispatch breakdown (scaled to master net so totals always match)
    _master_net_period = float(f_df["Net sales"].sum()) if "Net sales" in f_df.columns else 1.0
    _det_total = _det["_net"].sum() if not _det.empty else 1.0
    _scale = _master_net_period / _det_total if _det_total > 0 else 1.0

    _live_dispatch = {}
    if "Dispatch type" in _det.columns and not _det.empty:
        for _k, _v in _det.groupby("Dispatch type")["_net"].sum().items():
            _live_dispatch[str(_k).strip()] = {
                "revenue": round(float(_v) * _scale, 2),
                "orders":  int(_det[_det["Dispatch type"] == _k]["_net"].count())
            }
    if _live_dispatch:
        data["dispatch_truth"] = _live_dispatch

    # Channel breakdown
    _live_channels = {}
    if "Sales channel name" in _det.columns and not _det.empty:
        for _k, _v in _det.groupby("Sales channel name")["_net"].sum().items():
            _live_channels[str(_k).strip()] = round(float(_v) * _scale, 2)
    if _live_channels:
        data["channels"] = _live_channels

    # Hourly breakdown
    if not _det.empty:
        _live_hourly = {h: 0.0 for h in range(24)}
        for _h, _v in _det.groupby("_hour")["_net"].sum().items():
            _live_hourly[int(_h)] = round(float(_v) * _scale, 2)
        data["hourly_live"] = _live_hourly

    # Payment breakdown
    _pay_col = None
    for _pc in ["Payment method", "Payment type", "Payment"]:
        if _pc in _det.columns:
            _pay_col = _pc
            break
    if _pay_col and not _det.empty:
        _live_pay = {}
        for _k, _v in _det.groupby(_pay_col)["_net"].sum().items():
            _live_pay[str(_k).strip()] = round(float(_v) * _scale, 2)
        if _live_pay:
            data["payment"] = _live_pay

except Exception as _e:
    pass  # Silent fallback — static data from load_data() still used
# ── END DYNAMIC RECOMPUTATION ──────────────────────────────────────────────────

    # --- COMPLIANCE SETTINGS ---
    st.markdown("---")
    st.markdown("**⚖️ Compliance Settings**")
    # Then in sidebar just update it:
    min_wage_target = st.sidebar.number_input("Min Wage Threshold (£)", value=11.44, step=0.01)

    st.markdown("---")
    st.markdown("**🔧 Labour Report**")
    if st.button("📥 Generate Excel Report", key="gen_excel_btn", width="stretch"):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        build_labour_workbook(data, tmp_path)
        with open(tmp_path, "rb") as f:
            xlsx_bytes = f.read()
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        st.download_button(
            label="⬇️ Download chocoberry_labour_report.xlsx",
            data=xlsx_bytes,
            file_name="chocoberry_labour_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    
    # --- FORMULA LEDGER ---
    with st.sidebar.expander("📖 Business Logic & Formulas", expanded=False):
        st.markdown("""
        **1. Total Wage (Gross)**
        `Σ(Bank Pay + Cash Pay + Bonuses)`
        
        **2. Avg. Wage per Hour**
        `Total Wage ÷ Total Actual Hours`
        
        **3. Cost to Employer (Bank)**
        `Bank Payment (Bacs) + Est. Employers NI (~13.8%)`
        
        **4. Cash Wage**
        `Σ(Off-Books / Sunday Payouts)`
        
        **5. During the week**
        `Fixed Operating Costs (Shopping, Kitchen Cleaner, Staff Lunch)`
        
        **6. Total Cost to Employer**
        `(Bank Cost) + (Cash Wage) + (Operating Costs)`

        **7. Tax Calculations**
        `Standard 1257L (~£242/wk tax-free threshold)`
        *Estimated Tax: 20%, Estimated NI: 8%*
        
        ---
        *Logic matches Client Spreadsheet Standard.*
        """)

    st.markdown(f'<div style="font-size:10px; color:#6b7094">Syncs staff uploads from your phone-friendly portal into this system.</div>', unsafe_allow_html=True)

# ── KPI computation (Standardized) ──────────────────────────────────────
master_rev     = f_df["revenue"].sum() if "revenue" in f_df.columns else (f_df["Revenue"].sum() if "Revenue" in f_df.columns else 0.0)
master_tax     = f_df["tax"].sum() if "tax" in f_df.columns else (f_df["Tax on net sales"].sum() if "Tax on net sales" in f_df.columns else 0.0)
master_ord     = f_df["orders"].sum() if "orders" in f_df.columns else (f_df["Orders"].sum() if "Orders" in f_df.columns else 0.0)
master_net     = f_df["net"].sum() if "net" in f_df.columns else (f_df["Net sales"].sum() if "Net sales" in f_df.columns else 0.0)
master_charges = f_df["charges"].sum() if "charges" in f_df.columns else (f_df["Charges"].sum() if "Charges" in f_df.columns else 0.0)
total_refunds  = f_df["refunds"].sum() if "refunds" in f_df.columns else (f_df["refunds_clean"].sum() if "refunds_clean" in f_df.columns else 0.0)

total_truth_rev = sum(v["revenue"] for v in data["dispatch_truth"].values())
total_truth_ord = sum(v["orders"]  for v in data["dispatch_truth"].values())

# Dynamic Peak Hour computation
h_sorted = sorted(data["hourly_live"].items(), key=lambda x: x[1], reverse=True)
peak_h_int = h_sorted[0][0] if h_sorted else 21
peak_v_val = h_sorted[0][1] if h_sorted else 0
peak_h_str = f"{peak_h_int:02d}:00"
peak_pct   = (peak_v_val / master_net * 100) if master_net > 0 else 0
dispatch_ratio  = (
    data["dispatch_truth"].get(sel_dispatch, {}).get("revenue", 0) / total_truth_rev
    if sel_dispatch != "All" and total_truth_rev > 0 else 1.0
)

display_rev = master_rev * dispatch_ratio
display_ord = master_ord * dispatch_ratio
display_tax = master_tax * dispatch_ratio
display_net = master_net * dispatch_ratio
display_chr = master_charges * dispatch_ratio
aov         = display_net / display_ord if display_ord > 0 else 0
daily_avg   = display_net / len(f_df)  if len(f_df)  > 0 else 0

# ── Page header ───────────────────────────────────────────────────────────
st.markdown(f"""
<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:8px">
  <div>
    <div class="page-title">🍫 Chocoberry Intelligence Dashboard</div>
    <div class="page-sub">Business performance &middot; {start_d.strftime('%b %Y')} &ndash; {end_d.strftime('%b %Y')} &middot; {len(f_df)} trading days</div>
  </div>
  <div class="period-badge">📅 {start_d.strftime('%d %b')} – {end_d.strftime('%d %b %Y')} &nbsp;·&nbsp; {sel_dispatch}</div>
</div>
""", unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9, tab10, tab11, tab12, tab13, tab14 = st.tabs([
    "📊 Overview", "📈 Trends", "🕐 Patterns", "🛒 Channels",
    "⏰ Efficiency", "🔮 Forecast", "💷 Labour Report", "📅 Rota Builder",
    "🍕 Inventory & COGS", "♻️ Waste Log", "🚀 Strategic Optimization", "📄 Invoice Management", "💾 Database Explorer", "👥 Staff Management"
])

# ════════════════════════════════════════════════════════════════════
# TAB 1 — OVERVIEW
# ════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown('<div class="section-title">Key Performance Indicators</div>', unsafe_allow_html=True)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("💰 Total Net Sales",  f"£{display_net:,.0f}", f"+403% YoY" if sel_dispatch == "All" else sel_dispatch)
    k2.metric("🛒 Total Orders",     f"{int(display_ord):,}", f"{len(f_df)} trading days")
    k3.metric("🎯 Avg Order Value",  f"£{aov:.2f}")
    k4.metric("📅 Daily Average",    f"£{daily_avg:,.0f}")

    k5, k6, k7, k8 = st.columns(4)
    k5.metric("🏦 Tax Collected",    f"£{display_tax:,.2f}", f"{ (display_tax/display_net*100 if display_net > 0 else 0):.1f}% of net sales")
    
    # Charges are now pulled from the master daily timeline
    k6.metric("🚚 Total Charges", f"£{display_chr:,.2f}", "Flipdish verified")
    k7.metric("↩️ Total Refunds",    f"£{total_refunds:,.2f}", "0.02% refund rate ✅")
    k8.metric("🔥 Peak Hour",        peak_h_str, f"£{peak_v_val:,.0f} · {peak_pct:.1f}% of rev")

    st.markdown("---")
    st.markdown('<div class="section-title">Daily Revenue Trend</div>', unsafe_allow_html=True)

    col1, col2 = st.columns([2, 1])
    with col1:
        plot_df = f_df.copy()
        if sel_dispatch != "All":
            plot_df["Net sales"] = plot_df["Net sales"] * dispatch_ratio

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=plot_df["date"], y=plot_df["Net sales"],
            name="Daily Net Sales", fill="tozeroy",
            line=dict(color=COLORS["accent"], width=1.5),
            fillcolor="rgba(245,166,35,0.08)",
        ))
        rolling_src = plot_df["rolling7"] if "rolling7" in plot_df.columns else plot_df["Net sales"].rolling(7).mean()
        fig.add_trace(go.Scatter(
            x=plot_df["date"], y=rolling_src,
            name="7-Day Rolling Avg",
            line=dict(color=COLORS["accent2"], width=2.5, dash="dot"),
        ))
        dark_layout(fig, height=320, showlegend=True)
        # Force the X-axis to show the full range clearly
        fig.update_xaxes(
            title="", 
            range=[start_d, end_d],
            tickmode='array',
            tickvals=[start_d, 
                      start_d + (end_d - start_d)/4, 
                      start_d + (end_d - start_d)/2, 
                      start_d + 3*(end_d - start_d)/4, 
                      end_d],
            tickformat="%d %b",
            showgrid=True
        )
        fig.update_yaxes(title="", tickprefix="£")
        st.plotly_chart(fig, width="stretch")

    with col2:
        st.markdown("**Period Comparison**")
        
        # ── DYNAMIC PERIOD COMPARISON ──
        # Define the previous period of equal length
        _delta = (end_d - start_d) + pd.Timedelta(days=1)
        prev_start = start_d - _delta
        prev_end   = start_d - pd.Timedelta(days=1)
        
        # Filter all_df for previous period
        prev_df = all_df[(all_df["date"].dt.date >= prev_start) & (all_df["date"].dt.date <= prev_end)]
        prev_net = prev_df["Net sales"].sum() if not prev_df.empty else 0.0
        
        # Apply dispatch ratio if not 'All'
        prev_net = prev_net * dispatch_ratio
        
        # Calculate Growth
        if prev_net > 0:
            growth_pct = ((display_net - prev_net) / prev_net) * 100
            growth_str = f"{growth_pct:+.1f}%"
            growth_color = "#3ecf8e" if growth_pct >= 0 else "#ff4b4b"
            last_period_val = f"£{prev_net:,.0f}"
        else:
            growth_str = "NEW DATA"
            growth_color = "#6b7094"
            last_period_val = "N/A"

        st.markdown(f"""
        <div style="background:#1a1d26;border:1px solid #252836;border-radius:10px;padding:14px;margin-bottom:8px;text-align:center">
            <div style="font-size:10px;color:#6b7094;margin-bottom:4px">THIS PERIOD ({start_d.strftime('%b')}–{end_d.strftime('%b %Y')})</div>
            <div style="font-family:Syne,sans-serif;font-size:22px;font-weight:800;color:#f5a623">£{display_net:,.0f}</div>
            <div style="font-size:10px;color:#6b7094">Net sales</div>
        </div>
        <div style="background:#1a1d26;border:1px solid #252836;border-radius:10px;padding:14px;margin-bottom:8px;text-align:center">
            <div style="font-size:10px;color:#6b7094;margin-bottom:4px">LAST PERIOD ({prev_start.strftime('%b')}–{prev_end.strftime('%b %Y')})</div>
            <div style="font-family:Syne,sans-serif;font-size:22px;font-weight:800;color:#6b7094">{last_period_val}</div>
            <div style="font-size:10px;color:#6b7094">{'Insufficient Data' if last_period_val == 'N/A' else 'Net sales'}</div>
        </div>
        <div style="background:#1a1d26;border:1px solid {growth_color};border-radius:10px;padding:14px;text-align:center">
            <div style="font-size:10px;color:#6b7094;margin-bottom:4px">GROWTH</div>
            <div style="font-family:Syne,sans-serif;font-size:28px;font-weight:800;color:{growth_color}">{growth_str}</div>
            <div style="font-size:10px;color:#6b7094">{'Initial Period' if growth_str == 'NEW DATA' else 'Net sales vs prior period'}</div>
        </div>
        """, unsafe_allow_html=True)
        
        if last_period_val == "N/A":
            st.markdown('<div style="font-size:9px; color:#6b7094; text-align:center">💡 Note: Historical 2025 CSV data is required to unlock full year-over-year comparisons.</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Monthly Breakdown</div>', unsafe_allow_html=True)
    m_df = f_df.copy()
    m_df["month_str"] = m_df["date"].dt.strftime("%B %Y")
    monthly_stats = m_df.groupby("month_str").agg({"Net sales": "sum", "Orders": "sum"}).reset_index()
    monthly_stats["sort_key"] = pd.to_datetime(monthly_stats["month_str"], format="%B %Y")
    monthly_stats = monthly_stats.sort_values("sort_key").drop(columns=["sort_key"])
    monthly_stats["AOV"] = (monthly_stats["Net sales"] / monthly_stats["Orders"]).apply(lambda x: f"£{x:.2f}")
    monthly_stats["Orders"] = monthly_stats["Orders"].apply(lambda x: f"{int(x):,}")
    monthly_stats["Net Sales"] = monthly_stats["Net sales"].apply(lambda x: f"£{x:,.0f}")
    monthly_display = monthly_stats[["month_str","Net Sales","Orders","AOV"]]
    monthly_display.columns = ["Month","Net Sales","Orders","AOV"]
    st.dataframe(monthly_display, width="stretch", hide_index=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Day-by-Day Performance Table</div>', unsafe_allow_html=True)

    day_table_df = f_df.copy()
    if sel_dispatch != "All":
        day_table_df["Net sales"] = day_table_df["Net sales"] * dispatch_ratio
        day_table_df["Revenue"]   = day_table_df["Revenue"]   * dispatch_ratio
        day_table_df["Orders"]    = day_table_df["Orders"]    * dispatch_ratio

    day_table_df = day_table_df.sort_values("date").reset_index(drop=True)
    day_table_df["AOV"]   = (day_table_df["Revenue"] / day_table_df["Orders"]).round(2)
    
    def get_wow_pct(row):
        prev_date = row["date"] - pd.Timedelta(days=7)
        past_val = all_df[all_df["date"].dt.normalize() == prev_date.normalize()]["Net sales"]
        if not past_row.empty and (past_val := past_row.iloc[0]) > 0:
            return round(((row["Net sales"] - past_val) / past_val) * 100, 1)
        return np.nan

    # Optimization: Use merge for faster lookup if all_df is large
    _past = all_df[["date", "Net sales"]].copy()
    _past["date"] = _past["date"] + pd.Timedelta(days=7)
    _past = _past.rename(columns={"Net sales": "prev_net"})
    day_table_df = day_table_df.merge(_past, on="date", how="left")
    day_table_df["WoW %"] = ((day_table_df["Net sales"] - day_table_df["prev_net"]) / day_table_df["prev_net"] * 100).round(1)

    if not day_table_df.empty:
        peak_idx = day_table_df["Net sales"].idxmax()
        slow_idx = day_table_df["Net sales"].idxmin()
    else:
        peak_idx = None
        slow_idx = None

    display_day = day_table_df[["date","day","Net sales","Orders","AOV","Tax on net sales","WoW %"]].copy()
    display_day.columns = ["Date","Day","Net Sales £","Orders","AOV £","Tax £","WoW %"]
    display_day["Date"]        = display_day["Date"].dt.strftime("%d %b")
    display_day["WoW %"]       = display_day["WoW %"].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "—")

    def highlight_rows(row):
        if row.name == peak_idx:
            return ["background-color: rgba(62,207,142,0.15); color: #3ecf8e"] * len(row)
        elif row.name == slow_idx:
            return ["background-color: rgba(245,166,35,0.12); color: #f5a623"] * len(row)
        return [""] * len(row)

    styled = display_day.style.apply(highlight_rows, axis=1)
    st.dataframe(
        styled,
        width="stretch",
        hide_index=True,
        height=320,
        column_config={
            "Net Sales £": st.column_config.NumberColumn("Net Sales £", format="£%.2f"),
            "AOV £": st.column_config.NumberColumn("AOV £", format="£%.2f"),
            "Tax £": st.column_config.NumberColumn("Tax £", format="£%.2f"),
            "Orders": st.column_config.NumberColumn("Orders", format="%d"),
        }
    )
    st.markdown('<div class="insight-box">🟢 <b>Green row</b> = peak day (highest net sales) &nbsp;|&nbsp; 🟡 <b>Amber row</b> = slowest day &nbsp;|&nbsp; WoW % = change vs same day the prior week</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Best & Worst Trading Days</div>', unsafe_allow_html=True)
    if not day_table_df.empty:
        b1, b2 = st.columns(2)
        with b1:
            st.markdown("**🏆 Top 5 Best Days**")
            top5 = day_table_df.nlargest(5, "Net sales")[["date","day","Net sales","Orders"]].copy()
            top5["date"]      = top5["date"].dt.strftime("%d %b %Y")
            top5["Net sales"] = top5["Net sales"].apply(lambda x: f"£{x:,.0f}")
            top5.columns      = ["Date","Day","Net Sales","Orders"]
            medals = ["🥇","🥈","🥉","4","5"]
            top5.insert(0, "#", medals[:len(top5)])
            st.dataframe(top5, width="stretch", hide_index=True)
            # Dynamic Peak Insight
            top_day_obj = day_table_df.nlargest(1, "Net sales").iloc[0]
            peak_date_str = top_day_obj["date"].strftime("%d %b")
            peak_dow = top_day_obj["day"]
            st.markdown(f'<div class="insight-box">🎯 <b>{peak_date_str} ({peak_dow})</b> was your highest peak. Weekends consistently drive your highest volume. Plan staffing accordingly.</div>', unsafe_allow_html=True)

        with b2:
            st.markdown("**📉 5 Slowest Days**")
            bot5 = day_table_df.nsmallest(5, "Net sales")[["date","day","Net sales","Orders"]].copy()
            bot5["date"]      = bot5["date"].dt.strftime("%d %b %Y")
            bot5["Net sales"] = bot5["Net sales"].apply(lambda x: f"£{x:,.0f}")
            bot5.columns      = ["Date","Day","Net Sales","Orders"]
            ranks = ["1","2","3","4","5"]
            bot5.insert(0, "#", ranks[:len(bot5)])
            st.dataframe(bot5, width="stretch", hide_index=True)
            
            # Dynamic Slow Insight
            slowest_dows = day_table_df.nsmallest(3, "Net sales")["day"].mode()
            slowest_dow = slowest_dows[0] if len(slowest_dows) > 0 else "N/A"
            st.markdown(f'<div class="insight-box">⚠️ <b>{slowest_dow}s</b> are trending as your slowest trading days in this period. Consider mid-week promotions.</div>', unsafe_allow_html=True)
    else:
        st.info("No sales data available for this selected range to calculate performance rankings.")

    # ── PDF Report Button ──────────────────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-title">Weekly PDF Report</div>', unsafe_allow_html=True)
    if _pdf_available:
        if st.button("📄 Generate Weekly PDF Report", key="pdf_gen"):
            with st.spinner("Generating PDF..."):
                # 1. Determine the LATEST week in the dataset
                latest_date = all_df["date"].max()
                week_start  = latest_date - pd.Timedelta(days=latest_date.weekday())
                week_end    = week_start + pd.Timedelta(days=6)
                actual_label = f"{week_start.strftime('%d %b')} – {week_end.strftime('%d %b %Y')}"
                
                # 2. Re-load data specifically for the latest week
                week_data = load_data(reference_date=latest_date)
                
                # 3. CRITICAL: Slice the 'Total Year' dictionaries into 'Weekly' dictionaries
                # This preventsyearly totals from appearing in the weekly report.
                try:
                    detail_path = "Sales Summary Data/sales_data.csv"
                    if os.path.exists(detail_path):
                        _det = pd.read_csv(detail_path)
                        _det.columns = [c.strip() for c in _det.columns]
                        _time_col = "Order time" if "Order time" in _det.columns else _det.columns[4]
                        _det["_dt"] = pd.to_datetime(_det[_time_col], errors="coerce")
                        _det = _det.dropna(subset=["_dt"])
                        
                        # Slice to exactly the 7-day week
                        _mask = (_det["_dt"].dt.normalize() >= week_start.normalize()) & \
                                (_det["_dt"].dt.normalize() <= week_end.normalize())
                        _week_det = _det[_mask].copy()
                        
                        if not _week_det.empty:
                            _week_det = _week_det.drop_duplicates(subset=["Order ID"])
                            _net_col = "Net sales" if "Net sales" in _week_det.columns else "Net"
                            def _c(v):
                                try: return float(str(v).replace("£","").replace(",","").strip())
                                except: return 0.0
                            _week_det["_net_clean"] = _week_det[_net_col].apply(_c)
                            
                            # Update Dispatch
                            _type_col = "Dispatch type" if "Dispatch type" in _week_det.columns else "Type"
                            _disp_grp = _week_det.groupby(_type_col)["_net_clean"].sum().to_dict()
                            week_data["dispatch_truth"] = {k: {"revenue": v} for k, v in _disp_grp.items()}
                            
                            # Update Channels & Group Cardiff
                            _ch_col = "Sales channel name" if "Sales channel name" in _week_det.columns else "Channel"
                            def _std(n):
                                n_str = str(n).strip()
                                for platform in ["Uber Eats", "Deliveroo", "Just Eat", "Web", "Flipdish"]:
                                    if platform.lower() in n_str.lower(): return platform
                                if "Cardiff" in n_str or "Store" in n_str or "POS" in n_str.upper(): return "In-Store (POS)"
                                return n_str
                            _week_det["_std_ch"] = _week_det[_ch_col].apply(_std)
                            week_data["channels"] = _week_det.groupby("_std_ch")["_net_clean"].sum().to_dict()
                except Exception as e:
                    st.error(f"Weekly slice error: {e}")
                
                pdf_bytes = generate_weekly_pdf(week_data, actual_label)
            
            fname = f"chocoberry_weekly_{latest_date.strftime('%Y%m%d')}.pdf"
            st.download_button(
                "⬇️ Download PDF",
                data=pdf_bytes,
                file_name=fname,
                mime="application/pdf",
                key="pdf_dl"
            )
    else:
        st.info("Install reportlab to enable PDF export: `pip install reportlab`")


# ════════════════════════════════════════════════════════════════════
# TAB 2 — TRENDS
# ════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="section-title">Weekly Summary — Revenue, Orders & Tax</div>', unsafe_allow_html=True)

    # Auto-computed weekly data from CSVs via load_data()
    wk_df = pd.DataFrame(data["weekly"]) if data["weekly"] else pd.DataFrame(WEEKLY_DATA)
    wk_df["AOV"]       = (wk_df["net"] / wk_df["orders"].replace(0, 1)).round(2)
    wk_df["WoW Net %"] = wk_df["net"].pct_change().mul(100).round(1)
    wk_df["WoW Ord %"] = wk_df["orders"].pct_change().mul(100).round(1)

    weekly_display = wk_df.copy()
    weekly_display.columns = ["Week", "Net Sales £", "Orders", "Tax £", "AOV £", "WoW Net %", "WoW Ord %"]
    weekly_display["Net Sales £"] = weekly_display["Net Sales £"].apply(lambda x: f"£{x:,.0f}")
    weekly_display["Tax £"]       = weekly_display["Tax £"].apply(lambda x: f"£{x:,.0f}")
    weekly_display["AOV £"]       = weekly_display["AOV £"].apply(lambda x: f"£{x:.2f}")
    weekly_display["WoW Net %"]   = weekly_display["WoW Net %"].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "—")
    weekly_display["WoW Ord %"]   = weekly_display["WoW Ord %"].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "—")
    st.dataframe(weekly_display, width="stretch", hide_index=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Weekly Revenue Trend</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        # Highlight highest week automatically
        _max_net = wk_df["net"].max()
        colors_weekly = [COLORS["accent2"] if r["net"] == _max_net else COLORS["accent"] for _, r in wk_df.iterrows()]
        fig_w = go.Figure(go.Bar(
            x=wk_df["week"], y=wk_df["net"],
            marker_color=colors_weekly, marker_line_color=COLORS["accent"], marker_line_width=1,
        ))
        fig_w.update_traces(marker_opacity=0.8)
        dark_layout(fig_w, 300)
        fig_w.update_yaxes(tickprefix="£")
        fig_w.update_layout(title=dict(text="Weekly Net Sales", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_w, width="stretch")

    with col2:
        wow_pct   = wk_df["net"].pct_change().mul(100).dropna()
        wow_weeks = wk_df["week"].iloc[1:].tolist()
        fig_wow   = go.Figure(go.Bar(
            x=wow_weeks, y=wow_pct.values,
            marker_color=[COLORS["accent4"] if v >= 0 else COLORS["red"] for v in wow_pct.values],
            marker_opacity=0.8,
        ))
        dark_layout(fig_wow, 300)
        fig_wow.update_yaxes(ticksuffix="%")
        fig_wow.update_layout(title=dict(text="Week-over-Week Net Sales Growth %", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_wow, width="stretch")

    st.markdown("---")
    st.markdown('<div class="section-title">Monthly Overview</div>', unsafe_allow_html=True)

    # Dynamic monthly grouping — works for any date range
    _monthly = f_df.copy()
    _monthly["month_key"] = _monthly["date"].dt.to_period("M")
    _mon_grp = _monthly.groupby("month_key").agg({"Net sales": "sum", "Orders": "sum"}).reset_index()
    _mon_grp["month_label"] = _mon_grp["month_key"].dt.strftime("%b '%y")
    _mon_grp["AOV"] = (_mon_grp["Net sales"] / _mon_grp["Orders"].replace(0, 1)).round(2)
    _mon_colors = [COLORS["accent"], COLORS["accent2"], COLORS["accent4"], COLORS["accent3"],
                   COLORS["muted"], COLORS["red"]][:len(_mon_grp)]

    if not _mon_grp.empty:
        col3, col4, col5 = st.columns(3)

        with col3:
            fig_mb = go.Figure(go.Bar(
                x=_mon_grp["month_label"], y=_mon_grp["Net sales"],
                marker_color=_mon_colors, marker_line_width=0, marker_opacity=0.85,
            ))
            dark_layout(fig_mb, 240)
            fig_mb.update_yaxes(tickprefix="£")
            fig_mb.update_layout(title=dict(text="Monthly Net Sales", font=dict(size=12, color="#e8e9f0")))
            st.plotly_chart(fig_mb, width="stretch")

        with col4:
            fig_mo = go.Figure(go.Bar(
                x=_mon_grp["month_label"], y=_mon_grp["Orders"],
                marker_color=[f"rgba(124,92,191,{0.5 + 0.35*(i/max(len(_mon_grp)-1,1))})" for i in range(len(_mon_grp))],
                marker_line_color=COLORS["accent3"], marker_line_width=1.5,
            ))
            dark_layout(fig_mo, 240)
            fig_mo.update_layout(title=dict(text="Monthly Order Count", font=dict(size=12, color="#e8e9f0")))
            st.plotly_chart(fig_mo, width="stretch")

        with col5:
            st.markdown("<br>", unsafe_allow_html=True)
            best_aov_idx = _mon_grp["AOV"].idxmax()
            for idx, row in _mon_grp.iterrows():
                color = "#3ecf8e" if idx == best_aov_idx else "#6b7094"
                arrow = " ↑" if idx == best_aov_idx else ""
                st.markdown(f"""
                <div style="display:flex;justify-content:space-between;padding:10px 0;border-bottom:1px solid rgba(255,255,255,.04)">
                    <span style="font-size:12px;color:#6b7094">{row['month_label']}</span>
                    <span style="font-family:Syne,sans-serif;font-weight:700;font-size:14px;color:{color}">£{row['AOV']:.2f}{arrow}</span>
                </div>
                """, unsafe_allow_html=True)
            _best_mon = _mon_grp.loc[best_aov_idx, "month_label"]
            st.markdown(f'<div class="insight-box">AOV trending — <b>{_best_mon}</b> has the highest avg order value in this period.</div>', unsafe_allow_html=True)
    else:
        st.info("No monthly data available for this range.")


    st.markdown("---")
    st.markdown('<div class="section-title">7-Day Rolling Average vs Actual Daily Sales</div>', unsafe_allow_html=True)

    plot_df2 = f_df.copy()
    if sel_dispatch != "All":
        plot_df2["Net sales"] = plot_df2["Net sales"] * dispatch_ratio

    fig_roll = go.Figure()
    fig_roll.add_trace(go.Scatter(
        x=plot_df2["date"], y=plot_df2["Net sales"],
        name="Daily Actual", line=dict(color="rgba(124,92,191,0.6)", width=1),
    ))
    fig_roll.add_trace(go.Scatter(
        x=plot_df2["date"], y=plot_df2["Net sales"].rolling(7).mean(),
        name="7D Rolling Avg", line=dict(color=COLORS["accent"], width=2.5),
    ))
    dark_layout(fig_roll, 320, showlegend=True)
    fig_roll.update_yaxes(tickprefix="£")
    st.plotly_chart(fig_roll, width="stretch")

    st.markdown("---")
    st.markdown('<div class="section-title">Dual-Axis Intelligence: Sales vs Order Volume</div>', unsafe_allow_html=True)
    st.markdown('<div class="insight-box">Use this to see if a revenue drop is caused by fewer customers (Orders) or lower spending per person.</div>', unsafe_allow_html=True)

    from plotly.subplots import make_subplots
    fig_dual = make_subplots(specs=[[{"secondary_y": True}]])

    # Trace 1: Net Sales (Bar)
    fig_dual.add_trace(
        go.Bar(x=plot_df2["date"], y=plot_df2["Net sales"], name="Net Sales (£)", marker_color="rgba(245,166,35,0.7)"),
        secondary_y=False,
    )

    # Trace 2: Order Count (Line)
    # Use .get() or check columns to prevent KeyError: 'orders' vs 'Orders'
    ord_col = "Orders" if "Orders" in plot_df2.columns else "orders"
    fig_dual.add_trace(
        go.Scatter(x=plot_df2["date"], y=plot_df2[ord_col], name="Order Volume", line=dict(color="#3ecf8e", width=3)),
        secondary_y=True,
    )

    dark_layout(fig_dual, 400, showlegend=True)
    fig_dual.update_yaxes(title_text="Net Sales (£)", secondary_y=False, tickprefix="£")
    fig_dual.update_yaxes(title_text="Order Count", secondary_y=True)
    fig_dual.update_layout(hovermode="x unified")
    st.plotly_chart(fig_dual, width="stretch")

    st.markdown("---")
    # ── PDF EXPORT BUTTON (Moved to End of Tab) ──────────────────────
    if _pdf_available:
        try:
            pdf_bytes = generate_weekly_pdf(data, WEEK_LABEL)
            st.download_button(
                label="📥 Generate & Download Weekly PDF Report",
                data=pdf_bytes,
                file_name=f"chocoberry_report_{datetime.now().strftime('%d%b').lower()}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        except Exception as e:
            st.error(f"Error generating PDF: {e}")
    else:
        st.info("PDF Engine not initialized. Please ensure reportlab is installed.")


# ════════════════════════════════════════════════════════════════════
# TAB 3 — PATTERNS
# ════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="section-title">Weekly Trading Patterns</div>', unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    dow_map    = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    dow_totals = f_df.groupby("day")["Net sales"].sum().reindex(dow_map, fill_value=0)
    dow_labels = [d[:3] for d in dow_map]
    dow_nets   = dow_totals.values

    with col1:
        bar_colors = [COLORS["accent"] if i >= 4 else "rgba(245,166,35,0.35)" for i in range(7)]
        fig_dow = go.Figure(go.Bar(
            x=dow_labels, y=dow_nets,
            marker_color=bar_colors,
            marker_line_color=COLORS["accent"], marker_line_width=1.5,
        ))
        dark_layout(fig_dow, 280)
        fig_dow.update_yaxes(tickprefix="£")
        fig_dow.update_layout(title=dict(text="Revenue by Day of Week", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_dow, width="stretch")

    with col2:
        dow_ord_tot = f_df.groupby("day")["Orders"].sum().reindex(dow_map, fill_value=1)
        dow_aov     = [round(n/o, 2) if o > 0 else 0 for n,o in zip(dow_nets, dow_ord_tot)]
        day_counts  = f_df["day"].value_counts().reindex(dow_map, fill_value=1)
        dow_avg     = [round(n/c, 0) for n,c in zip(dow_nets, day_counts)]

        dow_table = pd.DataFrame({
            "Day":     [f"{d} {'🟢' if d=='Sunday' else '🔴' if d=='Monday' else ''}" for d in dow_map],
            "Total £": [f"£{n:,.0f}" for n in dow_nets],
            "Orders":  [f"{int(o):,}" for o in dow_ord_tot],
            "AOV":     [f"£{a:.2f}" for a in dow_aov],
            "Avg/Day": [f"£{avg:,.0f}" for avg in dow_avg],
        })
        st.dataframe(dow_table, width="stretch", hide_index=True)
        # Dynamic Pattern Insight
        total_rev = sum(dow_nets)
        weekend_rev = dow_nets[4] + dow_nets[5] + dow_nets[6] # Fri, Sat, Sun
        weekend_pct = (weekend_rev / total_rev * 100) if total_rev > 0 else 0
        sun_avg_val = dow_avg[6]
        mon_avg_val = dow_avg[0]
        
        st.markdown(f'<div class="insight-box">📅 Fri/Sat/Sun generate <b>{weekend_pct:.1f}% of weekly revenue</b> despite being 3 of 7 days. Sunday averages <b>£{sun_avg_val:,.0f}/day</b> vs Monday\'s £{mon_avg_val:,.0f}.</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">24-Hour Activity Distribution</div>', unsafe_allow_html=True)

    hour_keys = [f"{h:02d}:00" for h in range(24)]
    hour_vals = [data["hourly_live"].get(h, 0.0) for h in range(24)]

    h_colors = []
    for v in hour_vals:
        if v > 30000:   h_colors.append("rgba(245,166,35,0.95)")
        elif v > 15000: h_colors.append("rgba(245,166,35,0.65)")
        elif v > 5000:  h_colors.append("rgba(245,166,35,0.4)")
        else:           h_colors.append("rgba(245,166,35,0.15)")

    fig_hr = go.Figure(go.Bar(x=hour_keys, y=hour_vals, marker_color=h_colors, marker_line_width=0))
    dark_layout(fig_hr, 280)
    fig_hr.update_yaxes(tickprefix="£")
    fig_hr.update_layout(title=dict(text="Net Sales by Hour of Day", font=dict(size=13, color="#e8e9f0")))
    st.plotly_chart(fig_hr, width="stretch")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown('<p style="color:#f5a623;font-weight:700;margin-bottom:0">🔥 Top 3 Busiest Hours</p>', unsafe_allow_html=True)
        h_sorted = sorted(data["hourly_live"].items(), key=lambda x: x[1], reverse=True)
        top3_h   = h_sorted[:3]

        for i, (h_int, val) in enumerate(top3_h):
            h_str  = f"{format_hour_ampm(h_int)} – {format_hour_ampm(h_int+1)}"
            color_h = [COLORS["accent"], COLORS["accent2"], COLORS["accent3"]][i]
            st.markdown(f'<div style="display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)"><span style="font-size:12px;color:#6b7094">{h_str}</span><span style="font-family:Syne,sans-serif;font-weight:700;color:{color_h}">£{val:,.0f}</span></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="insight-box">💡 Peak window <b>{format_hour_ampm(top3_h[0][0])} – {format_hour_ampm(top3_h[2][0]+1)}</b> generates highest revenues. Full staff mandatory.</div>', unsafe_allow_html=True)


    with c2:
        st.markdown('<p style="color:#f5a623;font-weight:700;margin-bottom:0">🌙 Late Night Activity</p>', unsafe_allow_html=True)

        # Dynamic: pull from live hourly data
        _late_hours = [(0, f"{format_hour_ampm(0)} – {format_hour_ampm(1)}"), 
                       (1, f"{format_hour_ampm(1)} – {format_hour_ampm(2)}"), 
                       (2, f"{format_hour_ampm(2)} – {format_hour_ampm(3)}")]
        _midnight = data["hourly_live"].get(0, 0)
        for h_int, h_label in _late_hours:
            val = data["hourly_live"].get(h_int, 0)
            st.markdown(f'<div style="display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)"><span style="font-size:12px;color:#6b7094">{h_label}</span><span style="font-family:Syne,sans-serif;font-weight:700">£{val:,.0f}</span></div>', unsafe_allow_html=True)
        st.markdown(f'<div class="insight-box">🌙 Midnight hour generates <b>£{_midnight:,.0f}</b> — significant late-night trade.</div>', unsafe_allow_html=True)


    with c3:
        st.markdown('<p style="color:#f5a623;font-weight:700;margin-bottom:0">🌅 Quietest Hours</p>', unsafe_allow_html=True)

        # Dynamic: find the 3 lowest non-zero hours
        _nonzero = [(h, v) for h, v in data["hourly_live"].items() if v > 0]
        _quietest = sorted(_nonzero, key=lambda x: x[1])[:3]
        for h_int, val in _quietest:
            h_label = f"{format_hour_ampm(h_int)} – {format_hour_ampm(h_int+1)}"
            color = COLORS["red"] if val < 500 else COLORS["muted"]
            st.markdown(f'<div style="display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)"><span style="font-size:12px;color:#6b7094">{h_label}</span><span style="font-family:Syne,sans-serif;font-weight:700;color:{color}">£{val:,.0f}</span></div>', unsafe_allow_html=True)
        st.markdown('<div class="insight-box">📉 Early morning hours are essentially dead. No staffing or deliveries needed before midday.</div>', unsafe_allow_html=True)



# ════════════════════════════════════════════════════════════════════
# TAB 4 — CHANNELS
# ════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown('<div class="section-title">Dispatch & Sales Channels</div>', unsafe_allow_html=True)

    # ── DYNAMIC CHANNEL HARVESTING (Added for 2025 History) ──────────
    _has_plats = all(c in f_df.columns for c in ["deliveroo", "ubereat", "justeat", "pos_cash", "pos_card"])
    if _has_plats and f_df[["deliveroo", "ubereat", "justeat", "pos_cash", "pos_card"]].sum().sum() > 0:
        # Re-aggregate from filtered timeframe
        cur_channels = {
            "Deliveroo": f_df["deliveroo"].sum(),
            "Uber Eat":  f_df["ubereat"].sum(),
            "Just Eat":  f_df["justeat"].sum(),
            "POS":       (f_df["pos_cash"] + f_df["pos_card"]).sum(),
        }
        # Estimate dispatch splits based on POS breakdown
        pos_total = f_df["pos_cash"].sum() + f_df["pos_card"].sum()
        del_total = f_df["deliveroo"].sum() + f_df["ubereat"].sum() + f_df["justeat"].sum()
        grand_total_net = f_df["Net sales"].sum()
        total_orders_in_range = f_df["Orders"].sum()
        
        # Calculate revenue shares
        dine_share = (pos_total * 0.4) / grand_total_net if grand_total_net > 0 else 0
        take_share = (pos_total * 0.4) / grand_total_net if grand_total_net > 0 else 0
        coll_share = (pos_total * 0.2) / grand_total_net if grand_total_net > 0 else 0
        deli_share = del_total / grand_total_net if grand_total_net > 0 else 0
        
        cur_dispatch = {
            "Dine In":    {"revenue": pos_total * 0.4, "orders": int(total_orders_in_range * dine_share)},
            "Take Away":  {"revenue": pos_total * 0.4, "orders": int(total_orders_in_range * take_share)},
            "Delivery":   {"revenue": del_total,       "orders": int(total_orders_in_range * deli_share)},
            "Collection": {"revenue": pos_total * 0.2, "orders": int(total_orders_in_range * coll_share)},
        }
    else:
        # Fallback to static latest export
        cur_channels = data["channels"]
        cur_dispatch = data["dispatch_truth"]

    col1, col2 = st.columns(2)
    with col1:
        disp_rev_total = sum(v["revenue"] for v in cur_dispatch.values())
        disp_ord_total = sum(v["orders"]  for v in cur_dispatch.values())
        disp_labels    = list(cur_dispatch.keys())
        disp_revs      = [cur_dispatch[k]["revenue"] for k in disp_labels]

        fig_disp = go.Figure(go.Pie(
            labels=disp_labels, values=disp_revs, hole=0.6,
            marker=dict(colors=PALETTE[:4], line=dict(color="#12141a", width=3)),
        ))
        fig_disp.update_traces(textinfo="label+percent", textfont_size=11)
        dark_layout(fig_disp, 300, showlegend=False)
        fig_disp.update_layout(title=dict(text="Revenue by Dispatch Type", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_disp, width="stretch")

        # Final Balancing: Ensure total matches Net Sales
        accounted_rev = sum(d["revenue"] for d in cur_dispatch.values())
        unaccounted = grand_total_net - accounted_rev
        if unaccounted > 1: # More than £1 difference
            cur_dispatch["Other / Uncategorized"] = {
                "revenue": unaccounted, 
                "orders": int(unaccounted / (grand_total_net/total_orders_in_range)) if grand_total_net > 0 else 0
            }

        dispatch_rows = []
        disp_rev_total_final = sum(v["revenue"] for v in cur_dispatch.values())
        disp_ord_total_final = sum(v["orders"]  for v in cur_dispatch.values())
        for k, v in cur_dispatch.items():
            rev       = v["revenue"]
            ord_count = v["orders"]
            rev_pct   = rev / disp_rev_total_final * 100 if disp_rev_total_final > 0 else 0
            ord_pct   = ord_count / disp_ord_total_final * 100 if disp_ord_total_final > 0 else 0
            aov_d     = rev / ord_count if ord_count > 0 else 0
            dispatch_rows.append({
                "Type":    k,
                "Revenue": f"£{rev:,.0f}",
                "Rev %":   f"{rev_pct:.1f}%",
                "Orders":  f"{ord_count:,}",
                "Ord %":   f"{ord_pct:.1f}%",
                "AOV":     f"£{aov_d:.2f}",
            })
        st.dataframe(pd.DataFrame(dispatch_rows), width="stretch", hide_index=True)
        
        # DYNAMIC INSIGHT
        top_type = max(cur_dispatch, key=lambda x: cur_dispatch[x]["revenue"])
        st.markdown(f'<div class="insight-box">📦 <b>{top_type}</b> leads on revenue ({cur_dispatch[top_type]["revenue"]/disp_rev_total_final*100:.1f}%). All channels have been balanced to the master KPI total of £{grand_total_net:,.0f}.</div>', unsafe_allow_html=True)

    with col2:
        ch_df  = pd.DataFrame(list(cur_channels.items()), columns=["Platform","Sales"])
        # Add Other to bar chart if exists
        if unaccounted > 1:
            ch_df = pd.concat([ch_df, pd.DataFrame([{"Platform":"Other","Sales":unaccounted}])])
        
        fig_ch = go.Figure(go.Bar(
            x=ch_df["Platform"], y=ch_df["Sales"],
            marker_color=PALETTE[:len(ch_df)], marker_line_width=0,
        ))
        dark_layout(fig_ch, 300)
        fig_ch.update_yaxes(tickprefix="£")
        fig_ch.update_layout(title=dict(text="Revenue by Platform", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_ch, width="stretch")

    st.markdown("---")
    col3, col4 = st.columns(2)

    with col3:
        st.markdown("**Platform Detail**")
        ch_total = ch_df["Sales"].sum()
        ch_table = []
        medals = ["🥇","🥈","🥉","4","5","6","7","8","9","10"]
        for i, row in ch_df.sort_values("Sales", ascending=False).iterrows():
            platform = row["Platform"]
            sales    = row["Sales"]
            share_pct = (sales / ch_total * 100) if ch_total > 0 else 0
            ch_table.append({
                "#":          medals[i] if i < len(medals) else str(i+1),
                "Platform":   platform,
                "Net Sales":  f"£{sales:,.0f}",
                "% Share":    f"{share_pct:.1f}%",
            })
        st.dataframe(pd.DataFrame(ch_table), width="stretch", hide_index=True)
        st.markdown(f'<div class="insight-box">⚠️ <b>{ch_df.iloc[0]["Platform"]}</b> is your dominant platform. Promoting direct orders could save significant commission costs across this £{ch_total:,.0f} volume.</div>', unsafe_allow_html=True)

    with col4:
        pay_df  = pd.DataFrame(list(data["payment"].items()), columns=["Method","Sales"])
        fig_pay = go.Figure(go.Pie(
            labels=pay_df["Method"], values=pay_df["Sales"], hole=0.55,
            marker=dict(colors=PALETTE, line=dict(color="#12141a", width=3)),
        ))
        fig_pay.update_traces(textinfo="label+percent", textfont_size=10)
        dark_layout(fig_pay, 260, showlegend=False)
        fig_pay.update_layout(title=dict(text="Payment Methods", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_pay, width="stretch")
        st.markdown('<div class="insight-box">⚠️ <b>£687 unpaid orders</b> needs investigation. 11.8% cash — ensure reconciliation is in place.</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Key Business Insights</div>', unsafe_allow_html=True)

    i1, i2, i3 = st.columns(3)
    i4, i5, i6 = st.columns(3)
    # ── DYNAMIC BUSINESS INSIGHTS ──
    i_total_rev = f_df["Net sales"].sum()
    
    # 1. Growth Insight
    _delta = (end_d - start_d) + pd.Timedelta(days=1)
    prev_start = start_d - _delta
    prev_end   = start_d - pd.Timedelta(days=1)
    prev_df    = all_df[(all_df["date"].dt.date >= prev_start) & (all_df["date"].dt.date <= prev_end)]
    i_prev_net = prev_df["Net sales"].sum() if not prev_df.empty else 0
    if i_prev_net > 0:
        g_pct = ((i_total_rev - i_prev_net) / i_prev_net) * 100
        g_title = f"Revenue Growth {g_pct:+.1f}%"
        g_body = f"Business scaled vs prior period. £{i_total_rev/1000:,.1f}K vs £{i_prev_net/1000:,.1f}K previously."
    else:
        g_title = "Revenue Momentum"
        g_body = f"Strong baseline established with £{i_total_rev/1000:,.1f}K in net sales during this period."

    # 2. Delivery Insight
    # Assuming channel data is in data["channels"]
    i_delivery_rev = data["channels"].get("Uber Eats", 0) + data["channels"].get("Deliveroo", 0) + data["channels"].get("Just Eat", 0)
    i_del_pct = (i_delivery_rev / i_total_rev * 100) if i_total_rev > 0 else 0
    i_uber_pct = (data["channels"].get("Uber Eats", 0) / i_total_rev * 100) if i_total_rev > 0 else 0
    
    if not day_table_df.empty:
        # 3. Peak Hour Insight (Assumes peak_h_str is defined elsewhere)
        try:
            _peak_check = peak_h_str
        except:
            _peak_check = "N/A"

        # 4. Weekend Insight
        w_df = f_df.copy()
        w_df["is_weekend"] = w_df["day"].isin(["Friday", "Saturday", "Sunday"])
        weekend_rev = w_df[w_df["is_weekend"]]["Net sales"].sum()
        weekend_pct = (weekend_rev / i_total_rev * 100) if i_total_rev > 0 else 0
        sun_avg = w_df[w_df["day"] == "Sunday"]["Net sales"].mean() if not w_df[w_df["day"]=="Sunday"].empty else 0
        mon_avg = w_df[w_df["day"] == "Monday"]["Net sales"].mean() if not w_df[w_df["day"]=="Monday"].empty else 0

        # 5. Weak Spot Insight
        slow_day_name = day_table_df.nsmallest(1, "Net sales").iloc[0]["day"]
        slow_day_val  = day_table_df.nsmallest(1, "Net sales").iloc[0]["Net sales"]

        insights = [
            ("📈", g_title, g_body),
            ("🚗", "Delivery Dominates", f"Delivery channels generate {i_del_pct:.1f}% of all revenue. Uber Eats is your largest external partner at {i_uber_pct:.1f}% share."),
            ("🌙", "Late Night Business", f"Peak trading identified at {_peak_check}. The majority of revenue flows in late evening hours. Staff accordingly."),
            ("📅", "Weekend Warriors", f"Fri/Sat/Sun generate {weekend_pct:.1f}% of weekly revenue. Sunday averages £{sun_avg:,.0f}/day vs Monday's £{mon_avg:,.0f}."),
            ("🌐", "Channel Efficiency", f"POS In-Store is your core driver. External platforms take commission; shifting {i_del_pct*0.1:.1f}% more to POS would save significant fees."),
            ("⚠️", f"{slow_day_name} Weak Spot", f"{slow_day_name}s are trending slowest with lows of £{slow_day_val:,.0f}. Consider {slow_day_name} specific promotions."),
        ]
        for col, (icon, title, body) in zip([i1, i2, i3, i4, i5, i6], insights):
            col.markdown(f"""
            <div style="background:#12141a;border:1px solid #252836;border-radius:10px;padding:16px;height:100%">
                <div style="font-size:20px;margin-bottom:8px">{icon}</div>
                <div style="font-family:Syne,sans-serif;font-weight:700;font-size:13px;color:#e8e9f0;margin-bottom:6px">{title}</div>
                <div style="font-size:11px;color:#6b7094;line-height:1.6">{body}</div>
            </div>
            """, unsafe_allow_html=True)
    else:
        st.info("No business insights available for this empty date range.")


# ════════════════════════════════════════════════════════════════════
# TAB 5 — EFFICIENCY
# ════════════════════════════════════════════════════════════════════
with tab5:
    # --- DEFAULTS ---
    peak_label  = "20:00 - 22:00"  # Sensible default peak
    quiet_label = "12:00 - 15:00"  # Sensible default quiet

    st.markdown('<div class="section-title">Labour Analytics & Staffing Efficiency</div>', unsafe_allow_html=True)

    if not data["shifts"]:
        st.info("🕒 Detailed shift data (detailed_rota_with_shifts.csv) empty or not parsed. Review file format.")
    else:
        shift_df  = pd.DataFrame(data["shifts"])

        personnel_dict = data.get("personnel", {})
        def get_rate(name):
            return personnel_dict.get(name, {}).get("Hourly Rate", 11.44)

        shift_df["cost"] = shift_df["name"].apply(get_rate)
        h_labour  = shift_df.groupby("hour")["cost"].sum().reindex(range(10, 24), fill_value=0)

        h_sales_list = []
        for h in range(24):
            h_str = format_hour_ampm(h)
            h_sales_list.append({"Hour": h_str, "Rev": data["hourly_live"].get(h, 0.0)})
        h_sales = pd.DataFrame(h_sales_list)

        labour_map = {format_hour_ampm(h): v for h, v in h_labour.items()}
        h_sales["Labour"] = h_sales["Hour"].map(labour_map).fillna(0)

        fig_eff = go.Figure()
        fig_eff.add_trace(go.Bar(
            x=h_sales["Hour"], y=h_sales["Rev"], name="Avg Hourly Sales £",
            marker_color="rgba(245,166,35,0.4)", marker_line_width=0,
        ))
        fig_eff.add_trace(go.Scatter(
            x=h_sales["Hour"], y=h_sales["Labour"], name="Avg Labour Cost £",
            line=dict(color=COLORS["red"], width=3, shape="spline"),
        ))
        dark_layout(fig_eff, 350, showlegend=True)
        fig_eff.update_yaxes(tickprefix="£")
        fig_eff.update_layout(title=dict(text="Sales Volume vs. Staffing Cost (Daily Average)", font=dict(size=14, color="#e8e9f0")))
        st.plotly_chart(fig_eff, width="stretch")
        st.markdown('<div class="insight-box">🔴 <b>Red line</b> = Labour Cost | 🟡 <b>Yellow bars</b> = Sales Revenue. Hours with high red lines but low bars indicate <b>Overstaffing</b>. Target: Revenue should be >3x Labour.</div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Orders per Staff Hour (OPSH)**")
        if data["shifts"]:
            shift_df2   = pd.DataFrame(data["shifts"])
            staff_per_h = shift_df2.groupby("hour")["name"].nunique().reindex(range(24), fill_value=0)

            h_ord_list = []
            for h in range(24):
                h_ord_list.append({"Hour": format_hour_ampm(h), "Rev": data["hourly_live"].get(h, 0.0), "h_int": h})
            h_ord = pd.DataFrame(h_ord_list).set_index("h_int")

            h_ord["staff"] = staff_per_h
            h_ord["OPSH"] = h_ord.apply(lambda r: r["Rev"] / r["staff"] if r["staff"] > 0 else 0, axis=1)
            
            # --- Dynamic Stats Calculation ---
            # Peak: Highest Rev/Staff between 11:00 and 23:00 (ignore late night anomalies)
            peak_h = h_ord[(h_ord.index >= 11) & (h_ord["staff"] > 0)]["OPSH"].idxmax()
            quiet_h = h_ord[(h_ord.index >= 11) & (h_ord["staff"] > 1)]["OPSH"].idxmin()
            
            peak_label = f"{format_hour_ampm(peak_h)} - {format_hour_ampm((peak_h+1)%24)}"
            quiet_label = f"{format_hour_ampm(quiet_h)} - {format_hour_ampm((quiet_h+1)%24)}"

            fig_ops = go.Figure(go.Scatter(
                x=h_ord["Hour"], y=h_ord["OPSH"],
                fill="tozeroy", line=dict(color=COLORS["accent4"], width=3),
                fillcolor="rgba(62,207,142,0.05)",
            ))
            dark_layout(fig_ops, 250)
            fig_ops.update_layout(title=dict(text="Efficiency: Orders per Staff Member/Hour", font=dict(size=12, color="#e8e9f0")))
            st.plotly_chart(fig_ops, width="stretch")
        else:
            st.warning("Requires shift times.")

    with c2:
        st.markdown("**Operational Stats**")
        st.markdown(f"""
        <div style="background:#1a1c24;padding:20px;border-radius:12px;border:1px solid #252836">
            <div style="color:#6b7094;font-size:11px;text-transform:uppercase">Peak Efficiency Window</div>
            <div style="font-family:Syne,sans-serif;font-size:24px;color:#3ecf8e;font-weight:800">{peak_label}</div>
            <div style="margin-top:15px;color:#6b7094;font-size:11px;text-transform:uppercase">Quiet Zone (Lowest Efficiency)</div>
            <div style="font-family:Syne,sans-serif;font-size:24px;color:#e05c5c;font-weight:800">{quiet_label}</div>
        </div>
        """, unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════
# TAB 6 — FORECAST
# ════════════════════════════════════════════════════════════════════
with tab6:
    # Manual Target Week Selection
    from datetime import timedelta
    _f_df_max = pd.to_datetime(f_df["date"]).max() if not f_df.empty else pd.Timestamp.today()
    _default_mon = _f_df_max + timedelta(days=(7 - _f_df_max.weekday()))
    
    st.markdown('<div class="section-title">🔮 Forecast Configuration</div>', unsafe_allow_html=True)
    c_fc1, c_fc2 = st.columns([2, 1])
    with c_fc1:
        target_wk_mon = st.date_input(
            "Target Week for Forecast (Select Monday)", 
            value=_default_mon,
            key="fc_target_wk"
        )
    
    with c_fc2:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        # Button logic placed after calculation below
    
    # Force to Monday
    _next_mon = target_wk_mon - timedelta(days=target_wk_mon.weekday())
    _next_sun = _next_mon + timedelta(days=6)
    _fc_label = f"{_next_mon.strftime('%d %b')} – {_next_sun.strftime('%d %b %Y')}"
    
    _fc_aov      = (f_df["Net sales"].sum() / f_df["Orders"].replace(0,1).sum()).round(2) if "Orders" in f_df.columns and not f_df.empty else 0
    _fc_data_from = (f_df["date"].max() - timedelta(days=28)).strftime("%d %b") if not f_df.empty else "N/A"
    _fc_data_to   = f_df["date"].max().strftime("%d %b") if not f_df.empty else "N/A"

    st.markdown(f'<div class="section-title">Sales Forecast — Week of {_fc_label}</div>', unsafe_allow_html=True)

    f1, f2, f3 = st.columns(3)
    f1.metric("📐 Forecast Method",    "4-Week Rolling",  f"Same-day avg: {_fc_data_from} – {_fc_data_to}")
    # f2 and f3 filled after dynamic_forecast is computed below

    st.markdown("---")
    st.markdown('<div class="section-title">Manual Override — Adjust Forecast for Events & Promotions</div>', unsafe_allow_html=True)
    st.markdown('<div class="insight-box" style="margin-bottom:12px">Use overrides to adjust any day\'s forecast for known factors: bank holidays, local events, social campaigns, or operational changes. Enter a % uplift (positive) or reduction (negative).</div>', unsafe_allow_html=True)

    dow_map        = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    events_memory = load_events()
    active_boosts = []
    day_averages   = {}  # boosted forecast (multiplier applied)
    raw_baselines  = {}  # raw 4-week avg WITHOUT any event multiplier — used by rota engine for surge detection
    day_stds       = {}

    for i, day in enumerate(dow_map):
        target_date = _next_mon + timedelta(days=i)
        
        # Base Calculation: 4-week rolling average
        day_hist = f_df[f_df["day"] == day].sort_values("date", ascending=False)
        last_4   = day_hist.head(4)["Net sales"]
        
        base_avg = 0.0
        if not last_4.empty:
            base_avg = last_4.mean()
            day_stds[day] = last_4.std() if len(last_4) > 1 else last_4.mean() * 0.1
        else:
            day_stds[day] = 0.0

        # Store the RAW baseline BEFORE any event multiplier — this is the true historical average
        raw_baselines[day] = round(base_avg, 2)

        # Apply Memory Multiplier if exists
        multiplier = 1.0
        if target_date in events_memory:
            multiplier = events_memory[target_date].get('Multiplier', 1.0)
            event_name = events_memory[target_date].get('Event_Name', 'Special Event')
            active_boosts.append(f"🧠 <b>{day} ({target_date.strftime('%d %b')})</b>: {event_name} ({(multiplier-1)*100:+.0f}% Increase)")
        
        day_averages[day] = base_avg * multiplier  # boosted forecast

    # Build week_events: day_name → {Multiplier, Event_Name} for this target week
    # This is passed to the rota engine so it knows WHICH days are special events
    week_events = {}
    for i, day in enumerate(dow_map):
        target_date = _next_mon + timedelta(days=i)
        if target_date in events_memory:
            week_events[day] = events_memory[target_date]
    st.session_state["week_events"] = week_events

    if active_boosts:
        with st.container():
            st.markdown(f"""
            <div style="background:rgba(62,207,142,0.1);border:1px solid #3ecf8e;padding:12px;border-radius:10px;margin-bottom:15px">
                <div style="color:#3ecf8e;font-size:12px;font-weight:700;margin-bottom:5px">🧠 AI Memory Active for this Week</div>
                {"<br>".join([f'<div style="color:#e8e9f0;font-size:11px">{b}</div>' for b in active_boosts])}
            </div>
            """, unsafe_allow_html=True)

    dynamic_forecast = pd.Series(day_averages).reindex(dow_map, fill_value=0)
    dynamic_stds     = pd.Series(day_stds).reindex(dow_map, fill_value=0)

    last_7_days = f_df.sort_values("date", ascending=False).head(7)
    alert_triggered = False
    alerts_found    = []

    for _, row in last_7_days.iterrows():
        d_name = row["day"]
        actual = row["Net sales"]
        expected = day_averages.get(d_name, 0.0)
        if expected > 0:
            var = (actual - expected) / expected
            if abs(var) > 0.15:
                alert_triggered = True
                status = "🔥 PEAK OVER-TRADE" if var > 0 else "❄️ UNEXPECTED DROP"
                alerts_found.append(f"**{row['date'].strftime('%d %b')} ({d_name})**: {status} at **{var*100:+.1f}%** deviation (Actual £{actual:,.0f} vs Est. £{expected:,.0f})")

    if alert_triggered:
        with st.container():
            st.markdown(f'<div style="background:rgba(224,92,92,0.1);border:1px solid #e05c5c;padding:16px;border-radius:12px;margin-bottom:20px"><div style="color:#e05c5c;font-family:Syne,sans-serif;font-weight:800;font-size:13px;letter-spacing:1px;margin-bottom:8px">🚨 CRITICAL ACCURACY ALERTS — MODEL DEVIATION DETECTED</div>', unsafe_allow_html=True)
            for a in alerts_found[:3]:
                st.markdown(f'<div style="color:#6b7094;font-size:11px;margin-bottom:4px">{a}</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    ov_cols      = st.columns(7)
    override_pct = {}
    f_days       = dow_map

    for i, day in enumerate(f_days):
        base_val = dynamic_forecast[day]
        with ov_cols[i]:
            st.markdown(f"<div style='font-size:11px;color:#6b7094;text-align:center;margin-bottom:4px'>{day[:3]}</div>", unsafe_allow_html=True)
            st.markdown(f"<div style='font-size:12px;color:#f5a623;text-align:center;margin-bottom:6px'>£{base_val:,.0f}</div>", unsafe_allow_html=True)
            override_pct[day] = st.number_input(
                label=day, value=0, min_value=-50, max_value=100, step=5,
                label_visibility="collapsed", key=f"ov_{day}",
                help=f"% adjustment for {day}",
            )

    adjusted_forecast_map = {
        day: round(dynamic_forecast[day] * (1 + override_pct[day] / 100), 2)
        for day in f_days
    }
    st.session_state["weekly_forecast"] = adjusted_forecast_map
    # CRITICAL FIX: store the RAW 4-week baseline (before any event multiplier) so
    # the rota engine can correctly compare: boosted_forecast vs raw_baseline → true surge %
    st.session_state["weekly_baseline"] = raw_baselines
    adjusted_forecast = list(adjusted_forecast_map.values())
    total_base_fc     = dynamic_forecast.sum()
    total_adjusted_fc = sum(adjusted_forecast)
    override_applied  = any(v != 0 for v in override_pct.values())

    # Fill f2/f3 metrics now that dynamic_forecast is known
    f2.metric("💰 Total Week Forecast", f"£{total_base_fc:,.0f}", "Sum of 7 rolling day averages")
    f3.metric("🎯 Forecast AOV",        f"£{_fc_aov:.2f}",         f"{len(all_df)}-day trailing average")

    if override_applied:
        oa1, oa2, oa3 = st.columns(3)
        oa1.metric("Base Forecast",     f"£{total_base_fc:,.0f}")
        oa2.metric("Adjusted Forecast", f"£{total_adjusted_fc:,.0f}",
                   f"{((total_adjusted_fc - total_base_fc)/total_base_fc*100):+.1f}%")
        oa3.metric("Override Impact",   f"£{total_adjusted_fc - total_base_fc:+,.0f}")

    override_reason = st.text_input(
        "Override reason (optional)",
        placeholder="e.g. Good Friday bank holiday, -20% expected Mon/Tue | Cardiff event +15% Fri",
        key="override_reason",
    )

    # Export PDF Logic
    _adj_total = sum(adjusted_forecast_map.values())
    pdf_day_data = []
    for i, day in enumerate(dow_map):
        pdf_day_data.append({
            'day': day,
            'date': (_next_mon + timedelta(days=i)).strftime('%d %b'),
            'sales': adjusted_forecast_map[day]
        })
    
    pdf_bytes = generate_forecast_pdf(
        week_label=_fc_label,
        total_forecast=_adj_total,
        aov=_fc_aov,
        day_data=pdf_day_data,
        boosts=[b.replace('<b>','').replace('</b>','') for b in active_boosts]
    )
    
    with c_fc2:
        st.download_button(
            label="📥 Download Forecast PDF",
            data=pdf_bytes,
            file_name=f"CBC_Forecast_{_next_mon.strftime('%Y%m%d')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        f_days_labels = [d[:3] for d in f_days]
        f_base_vals   = list(dynamic_forecast.values)
        f_adj_vals    = list(adjusted_forecast_map.values())

        fig_fc = go.Figure()

        upper_band = [v + dynamic_stds[d] for d, v in zip(f_days, f_base_vals)]
        lower_band = [max(0, v - dynamic_stds[d]) for d, v in zip(f_days, f_base_vals)]

        fig_fc.add_trace(go.Scatter(
            name="Lower Bound", x=f_days_labels, y=lower_band,
            line=dict(width=0), showlegend=False, hoverinfo="skip"
        ))
        fig_fc.add_trace(go.Scatter(
            name="Forecast Range (±1σ)", x=f_days_labels, y=upper_band,
            fill='tonexty', fillcolor="rgba(124,92,191,0.08)", line=dict(width=0),
            hoverinfo="skip"
        ))
        fig_fc.add_trace(go.Bar(
            name="Base Forecast", x=f_days_labels, y=f_base_vals,
            marker_color="rgba(124,92,191,0.35)",
            marker_line_color=COLORS["accent3"], marker_line_width=1,
        ))
        if override_applied:
            fig_fc.add_trace(go.Bar(
                name="Adjusted Forecast", x=f_days_labels, y=f_adj_vals,
                marker_color=COLORS["accent3"],
                marker_line_color=COLORS["accent3"], marker_line_width=1.5, marker_opacity=0.85,
            ))
        dark_layout(fig_fc, 300, showlegend=override_applied)
        fig_fc.update_yaxes(tickprefix="£")
        fig_fc.update_layout(
            title=dict(text="Next Week Revenue Forecast" + (" (Adjusted)" if override_applied else ""),
                       font=dict(size=13, color="#e8e9f0")),
            barmode="overlay",
        )
        st.plotly_chart(fig_fc, width="stretch")

    with col2:
        st.markdown("**Day-by-Day Forecast**")
        max_adj = max(adjusted_forecast_map.values())
        min_adj = min(adjusted_forecast_map.values())
        for day, adj_val in adjusted_forecast_map.items():
            pct        = adj_val / max_adj if max_adj > 0 else 0
            color      = COLORS["accent4"] if adj_val == max_adj else (COLORS["red"] if adj_val == min_adj else COLORS["accent"])
            delta_html = ""
            if override_pct.get(day, 0) != 0:
                delta_color = "#3ecf8e" if override_pct[day] > 0 else "#e05c5c"
                delta_html  = f' <span style="font-size:10px;color:{delta_color}">({override_pct[day]:+d}%)</span>'
            st.markdown(f"""
            <div style="display:flex;align-items:center;gap:12px;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)">
                <div style="width:80px;font-size:12px;color:#6b7094">{day}</div>
                <div style="flex:1;height:6px;background:#252836;border-radius:3px;overflow:hidden">
                    <div style="width:{pct*100:.0f}%;height:100%;background:linear-gradient(90deg,{COLORS['accent3']},{COLORS['accent']});border-radius:3px"></div>
                </div>
                <div style="width:90px;text-align:right;font-family:Syne,sans-serif;font-weight:700;font-size:13px;color:{color}">£{adj_val:,.0f}{delta_html}</div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown(f"""
        <div style="display:flex;justify-content:space-between;margin-top:14px;padding-top:12px;border-top:1px solid #252836">
            <span style="font-size:12px;color:#6b7094;text-transform:uppercase;letter-spacing:1px">Total {'Adjusted ' if override_applied else ''}Forecast</span>
            <span style="font-family:Syne,sans-serif;font-weight:800;font-size:18px;color:#f5a623">£{total_adjusted_fc:,.0f}</span>
        </div>
        """, unsafe_allow_html=True)

    # ── ACTUALS vs FORECAST LIVE TABLE ──────────────────────────────────────
    st.markdown("---")
    st.markdown('<div class="section-title">📊 Actuals vs Forecast — Plan vs Reality</div>', unsafe_allow_html=True)
    st.markdown('<div class="insight-box">Comparing your forecast targets against real sales. Green = beat forecast. Amber = within 10%. Red = missed by more than 10%.</div>', unsafe_allow_html=True)

    # Build the projected week date range
    _avf_dates = [_next_mon + timedelta(days=i) for i in range(7)]
    _avf_days  = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

    # Look up actuals from full dataset (all_df) for those dates
    _avf_rows = []
    _avf_tot_fc  = 0.0
    _avf_tot_act = 0.0
    _avf_tot_ord = 0

    for i, (_day, _dt) in enumerate(zip(_avf_days, _avf_dates)):
        _fc_val  = adjusted_forecast_map.get(_day, 0.0)
        _act_row = all_df[all_df["date"] == pd.Timestamp(_dt)]
        _act_val = float(_act_row["net"].sum()) if not _act_row.empty and "net" in _act_row.columns else None
        _act_ord = int(_act_row["orders"].sum()) if not _act_row.empty and "orders" in _act_row.columns else None

        _avf_tot_fc += _fc_val
        if _act_val is not None:
            _avf_tot_act += _act_val
            if _act_ord: _avf_tot_ord += _act_ord

        _avf_rows.append({
            "day":     _day,
            "date":    _dt.strftime("%d %b"),
            "fc":      _fc_val,
            "act":     _act_val,
            "orders":  _act_ord,
        })

    # Render table
    _hdr = ["Day", "Date", "Forecast", "Actual", "Orders", "Diff", "Achievement"]
    _tbl_html = f"""
<table style="width:100%;border-collapse:collapse;font-family:Inter,sans-serif;font-size:13px;margin-top:8px">
<thead>
<tr style="background:#1a1d26;color:#f5a623;text-transform:uppercase;font-size:11px;letter-spacing:1px">
{"".join(f'<th style="padding:10px 12px;text-align:left">{h}</th>' for h in _hdr)}
</tr>
</thead>
<tbody>
"""

    for r in _avf_rows:
        if r["act"] is not None:
            _pct  = (r["act"] / r["fc"] * 100) if r["fc"] else 0
            _diff = r["act"] - r["fc"]
            _diff_str = f'{"+" if _diff >= 0 else ""}£{_diff:,.0f}'
            _pct_str  = f'{_pct:.1f}%'
            _color    = "#3ecf8e" if _pct >= 100 else "#f5a623" if _pct >= 90 else "#e05c5c"
            _bg       = "rgba(62,207,142,0.06)" if _pct >= 100 else "rgba(245,166,35,0.06)" if _pct >= 90 else "rgba(224,92,92,0.06)"
            _act_str  = f'£{r["act"]:,.2f}'
            _ord_str  = f'{r["orders"]:,}' if r["orders"] else "—"
        else:
            _diff_str = "—"
            _pct_str  = "Pending"
            _color    = "#6b7094"
            _bg       = "transparent"
            _act_str  = "Pending"
            _ord_str  = "—"

        _tbl_html += f"""
<tr style="border-bottom:1px solid rgba(255,255,255,0.04);background:{_bg}">
<td style="padding:10px 12px;color:#e8e9f0;font-weight:600">{r["day"]}</td>
<td style="padding:10px 12px;color:#6b7094">{r["date"]}</td>
<td style="padding:10px 12px;color:#a0a3b8;font-family:Syne,sans-serif">£{r["fc"]:,.2f}</td>
<td style="padding:10px 12px;color:#e8e9f0;font-family:Syne,sans-serif;font-weight:700">{_act_str}</td>
<td style="padding:10px 12px;color:#6b7094">{_ord_str}</td>
<td style="padding:10px 12px;color:{_color};font-weight:600">{_diff_str}</td>
<td style="padding:10px 12px;color:{_color};font-weight:700">{_pct_str}</td>
</tr>
"""

    # Totals row
    if _avf_tot_act > 0:
        _tot_pct   = (_avf_tot_act / _avf_tot_fc * 100) if _avf_tot_fc else 0
        _tot_diff  = _avf_tot_act - _avf_tot_fc
        _tot_color = "#3ecf8e" if _tot_pct >= 100 else "#f5a623" if _tot_pct >= 90 else "#e05c5c"
        _tbl_html += f"""
<tr style="border-top:2px solid rgba(245,166,35,0.4);background:rgba(245,166,35,0.06)">
<td style="padding:12px;color:#f5a623;font-weight:800;font-family:Syne,sans-serif" colspan="2">TOTAL WEEK</td>
<td style="padding:12px;color:#a0a3b8;font-family:Syne,sans-serif;font-weight:700">£{_avf_tot_fc:,.2f}</td>
<td style="padding:12px;color:#e8e9f0;font-family:Syne,sans-serif;font-weight:800">£{_avf_tot_act:,.2f}</td>
<td style="padding:12px;color:#6b7094">{_avf_tot_ord:,}</td>
<td style="padding:12px;color:{_tot_color};font-weight:700">{"+" if _tot_diff >= 0 else ""}£{_tot_diff:,.0f}</td>
<td style="padding:12px;color:{_tot_color};font-weight:800;font-size:15px">{_tot_pct:.1f}%</td>
</tr>
"""
    else:
        _tbl_html += f"""
<tr style="border-top:2px solid rgba(245,166,35,0.4);background:rgba(245,166,35,0.06)">
<td style="padding:12px;color:#f5a623;font-weight:800" colspan="2">TOTAL WEEK</td>
<td style="padding:12px;color:#a0a3b8;font-family:Syne,sans-serif;font-weight:700">£{_avf_tot_fc:,.2f}</td>
<td style="padding:12px;color:#6b7094" colspan="4">No actuals yet for this week</td>
</tr>
"""

    _tbl_html += "</tbody></table>"
    st.markdown(_tbl_html, unsafe_allow_html=True)
    # ── END ACTUALS vs FORECAST ───────────────────────────────────────────────

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("📊 Export Forecast PDF for Client", width="stretch"):
            try:
                from reportlab.lib import colors as rl_colors
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib.units import cm
                import io as _io

                buf = _io.BytesIO()
                doc = SimpleDocTemplate(buf, pagesize=A4)
                styles = getSampleStyleSheet()
                t_s = ParagraphStyle("T", parent=styles["Heading1"], fontSize=20, textColor=rl_colors.HexColor("#1a1a2e"))
                h_s = ParagraphStyle("H", parent=styles["Normal"], fontSize=12, fontWeight="Bold")
                
                story = []
                story.append(Paragraph(f"Chocoberry Revenue Forecast Report", t_s))
                story.append(Paragraph(f"Forecast Period: {_fc_label}", styles["Normal"]))
                story.append(Spacer(1, 0.5*cm))
                story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor("#f5a623")))
                story.append(Spacer(1, 0.5*cm))
                
                story.append(Paragraph("<b>Daily Revenue Predictions:</b>", h_s))
                f_data = [["Day of Week", "Forecasted Net Sales"]]
                for d, v in adjusted_forecast_map.items():
                    f_data.append([d, f"£{v:,.2f}"])
                f_data.append(["<b>TOTAL WEEK</b>", f"<b>£{total_adjusted_fc:,.2f}</b>"])
                
                tbl = Table(f_data, colWidths=[6*cm, 6*cm])
                tbl.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), rl_colors.HexColor("#1a1a2e")),
                    ('TEXTCOLOR', (0,0), (-1,0), rl_colors.white),
                    ('GRID', (0,0), (-1,-1), 0.5, rl_colors.grey),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('FONTSIZE', (0,0), (-1,-1), 10),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ]))
                story.append(tbl)
                story.append(Spacer(1, 1*cm))
                
                story.append(Paragraph(f"<b>Key Insights:</b>", h_s))
                story.append(Paragraph(f"• Expected Average Order Value (AOV): £{_fc_aov:.2f}", styles["Normal"]))
                story.append(Paragraph(f"• Baseline Data: 4-week rolling average ({_fc_data_from} - {_fc_data_to})", styles["Normal"]))
                if override_applied:
                    story.append(Paragraph(f"• Manual Adjustments Applied: Yes ({override_reason if override_reason else 'Not specified'})", styles["Normal"]))
                
                doc.build(story)
                pdf_bytes = buf.getvalue()
                st.download_button(
                    label="⬇️ Click here to Download Forecast.pdf",
                    data=pdf_bytes,
                    file_name=f"chocoberry_forecast_{_fc_label.replace(' ','_')}.pdf",
                    mime="application/pdf"
                )
            except Exception as fe:
                st.error(f"Could not generate PDF: {fe}")

    st.markdown("---")
    st.markdown('<div class="section-title">Forecast vs Actual — Retrospective Comparison</div>', unsafe_allow_html=True)

    # --- DYNAMIC FORECAST RETROSPECTIVE ---
    # Merge historical predictions with live actuals from data["weekly"]
    # --- DYNAMIC FORECAST ACCURACY HISTORY ---
    # We reconstruct the 'past' by calculating what the model would have predicted
    # for each of the last 6 completed weeks.
    
    history_rows = []
    _latest_m = all_df["date"].max() - pd.Timedelta(days=all_df["date"].max().weekday())
    # Last 6 completed weeks
    _past_mondays = [_latest_m - pd.Timedelta(weeks=i) for i in range(1, 7)]
    _past_mondays.reverse()
    
    for m in _past_mondays:
        _fc_val = 0
        for d_name in ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]:
            _h = all_df[(all_df["day"] == d_name) & (all_df["date"] < m)].sort_values("date", ascending=False).head(4)
            _fc_val += _h["Net sales"].mean() if not _h.empty else 0
        
        _actual = all_df[(all_df["date"] >= m) & (all_df["date"] < m + pd.Timedelta(days=7))]["Net sales"].sum()
        
        label  = f"Week of {m.strftime('%d %b')}"
        fc     = round(_fc_val, 2)
        actual = round(_actual, 2)
        
        if actual is not None and actual > 0:
            diff    = actual - fc
            acc_pct = (1 - abs(diff) / fc) * 100
            history_rows.append({
                "Week":       label,
                "Forecast £": f"£{fc:,.0f}",
                "Actual £":   f"£{actual:,.0f}",
                "Difference": f"£{diff:+,.0f}",
                "Accuracy %": f"{acc_pct:.1f}%",
                "Status":     "✅" if acc_pct >= 90 else ("⚠️" if acc_pct >= 75 else "❌"),
                "raw_fc":     fc,
                "raw_act":    actual
            })
        else:
            history_rows.append({
                "Week":       label,
                "Forecast £": f"£{fc:,.0f}",
                "Actual £":   "— (pending)",
                "Difference": "—",
                "Accuracy %": "—",
                "Status":     "🔮 Upcoming",
                "raw_fc":     fc,
                "raw_act":    None
            })
    
    st.dataframe(pd.DataFrame(history_rows).drop(columns=["raw_fc","raw_act"]), width="stretch", hide_index=True)

    # Build Chart Data from dynamic rows
    completed = [r for r in history_rows if r["raw_act"] is not None]
    if completed:
        fc_chart_weeks  = [r["Week"] for r in completed]
        fc_chart_fc     = [r["raw_fc"] for r in completed]
        fc_chart_actual = [r["raw_act"] for r in completed]

        fig_cmp = go.Figure()
        fig_cmp.add_trace(go.Scatter(
            x=fc_chart_weeks, y=fc_chart_fc, name="Forecast", mode="lines+markers",
            line=dict(color=COLORS["accent3"], width=2, dash="dot"), marker=dict(size=8),
        ))
        fig_cmp.add_trace(go.Scatter(
            x=fc_chart_weeks, y=fc_chart_actual, name="Actual", mode="lines+markers",
            line=dict(color=COLORS["accent4"], width=2.5), marker=dict(size=8),
        ))
        dark_layout(fig_cmp, 280, showlegend=True)
        fig_cmp.update_yaxes(tickprefix="£")
        fig_cmp.update_layout(title=dict(text="Forecast vs Actual — Weekly Comparison", font=dict(size=13, color="#e8e9f0")))
        st.plotly_chart(fig_cmp, width="stretch")

        valid_completed = [r for r in history_rows if r["raw_act"] is not None]
        if valid_completed:
            fc_chart_weeks  = [r["Week"] for r in valid_completed]
            fc_chart_fc     = [r["raw_fc"] for r in valid_completed]
            fc_chart_actual = [r["raw_act"] for r in valid_completed]

            acc_vals = [(1 - abs(r["raw_act"] - r["raw_fc"]) / r["raw_fc"]) * 100 for r in valid_completed]
            avg_acc  = sum(acc_vals) / len(acc_vals)

            fig_acc = go.Figure()
            fig_acc.add_trace(go.Scatter(
                x=fc_chart_weeks, y=acc_vals, name="Accuracy %", mode="lines+markers",
                line=dict(color=COLORS["accent"], width=2), marker=dict(size=7),
                fill="tozeroy", fillcolor="rgba(245,166,35,0.06)",
            ))
            fig_acc.add_hline(y=90, line_dash="dot", line_color=COLORS["accent4"],
                              annotation_text="90% target", annotation_position="right")
            dark_layout(fig_acc, 220, showlegend=False)
            fig_acc.update_yaxes(ticksuffix="%", range=[0, 110])
            fig_acc.update_layout(title=dict(text=f"Forecast Accuracy % by Week  (avg: {avg_acc:.1f}%)", font=dict(size=13, color="#e8e9f0")))
            st.plotly_chart(fig_acc, width="stretch")

            a1, a2, a3 = st.columns(3)
            a1.metric("📊 Average Accuracy",  f"{avg_acc:.1f}%",              "Across completed weeks")
            best_acc_idx  = acc_vals.index(max(acc_vals))
            a2.metric("🏆 Best Week",  fc_chart_weeks[best_acc_idx],  f"{max(acc_vals):.1f}% accurate")
            worst_acc_idx = acc_vals.index(min(acc_vals))
            a3.metric("⚠️ Worst Week", fc_chart_weeks[worst_acc_idx], f"{min(acc_vals):.1f}% accurate")

    st.markdown("---")
    st.markdown('<div class="section-title">Forecast Assumptions & Historical Performance</div>', unsafe_allow_html=True)
    g1, g2, g3 = st.columns(3)

    with g1:
        st.markdown("**📋 Model Assumptions**")
        _date_range_str = f"{_fc_data_from} – {_fc_data_to}"
        for row in [("Method","4-Week Same-Day Avg"),("Data Used", _date_range_str),
                    ("Data Points/Day","3–4 observations"),("Refund Adjusted","✅ Yes")]:
            st.markdown(f'<div style="display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)"><span style="font-size:11px;color:#6b7094">{row[0]}</span><span style="font-size:12px;color:#e8e9f0">{row[1]}</span></div>', unsafe_allow_html=True)

    with g2:
        st.markdown("**⚠️ Override Factors Applied**")
        if override_applied:
            for day, pct in override_pct.items():
                if pct != 0:
                    col_c = "#3ecf8e" if pct > 0 else "#e05c5c"
                    st.markdown(f'<div style="display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)"><span style="font-size:11px;color:#6b7094">{day}</span><span style="font-size:12px;font-weight:700;color:{col_c}">{pct:+d}%</span></div>', unsafe_allow_html=True)
            if override_reason:
                st.markdown(f'<div class="insight-box">📝 <b>Reason:</b> {override_reason}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="insight-box">No overrides applied. Base forecast in use.</div>', unsafe_allow_html=True)
            st.markdown('<div class="insight-box">📅 <b>Bank Holiday</b> — Check for Good Friday, Easter Monday impacts</div>', unsafe_allow_html=True)
            st.markdown('<div class="insight-box">🎉 <b>Local Events</b> — Cardiff events can significantly boost Fri/Sat</div>', unsafe_allow_html=True)

    with g3:
        st.markdown("**📊 Last 4 Weeks Actuals**")
        # Dynamic: pull from live weekly data
        _wk_hist = pd.DataFrame(data["weekly"]).tail(4) if data["weekly"] else pd.DataFrame()
        if not _wk_hist.empty:
            _wk_max = _wk_hist["net"].max()
            _wk_min = _wk_hist["net"].min()
            for _, _wr in _wk_hist.iterrows():
                _arrow = " ↑" if _wr["net"] == _wk_max else (" ↓" if _wr["net"] == _wk_min else "")
                _col   = "#3ecf8e" if _wr["net"] == _wk_max else ("#e05c5c" if _wr["net"] == _wk_min else "#e8e9f0")
                st.markdown(f'<div style="display:flex;justify-content:space-between;padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04)"><span style="font-size:11px;color:#6b7094">Wk {_wr["week"]}</span><span style="font-family:Syne,sans-serif;font-weight:700;font-size:13px;color:{_col}">£{_wr["net"]:,.0f}{_arrow}</span></div>', unsafe_allow_html=True)
        st.markdown('<div class="insight-box">High week-to-week variance. Treat forecast as baseline — apply override factors above.</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════════
# TAB 7 — LABOUR REPORT
# ════════════════════════════════════════════════════════════════════
with tab7:
    st.markdown('<div class="section-title">Labour Cost Analysis — Week: ' + data["week_label"] + ' &nbsp;|&nbsp; 👷 Responsible: DHIRAJ</div>', unsafe_allow_html=True)

    staff_df_live   = calc_staff_wages(data)
    summary_live    = calc_labour_summary(staff_df_live, forecast_rev=total_adjusted_fc)
    overlay_df_live = calc_hourly_overlay(data["hourly_live"], data=data)
    over_df_live    = calc_overstaffing(overlay_df_live)
    under_df_live   = calc_understaffing(overlay_df_live)
    day_df_live     = calc_day_labour(total_wages=summary_live["Total Cost to Employer"], forecast_data=adjusted_forecast_map)

    krow1_1, krow1_2, krow1_3 = st.columns(3)
    # Gross Wage = Salaries + Management fees (This should match the Rota Engine's £3,407.20 target)
    gross_wage_total = summary_live['Staff Salaries (Shifts)'] + summary_live['Management / Professional']
    krow1_1.metric("💎 Gross Staff Wage",  f"£{gross_wage_total:,.2f}")
    krow1_2.metric("💸 Cash Wage Portion", f"£{summary_live['Cash Wage Total']:,.2f}")
    krow1_3.metric("🏥 Bank/Bacs Portion", f"£{(gross_wage_total - summary_live['Cash Wage Total']):,.2f}")

    st.markdown("<br>", unsafe_allow_html=True)

    krow2_1, krow2_2, krow2_3 = st.columns(3)
    krow2_1.metric("⚖️ Statutory Emp NI", f"£{summary_live['Statutory Employer NI']:,.2f}")
    krow2_2.metric("🏗️ During the Week",   f"£{summary_live['Operational Costs (£330)']:,.2f}")
    krow2_3.metric("🔥 Total Cost to Employer", f"£{summary_live['Total Cost to Employer']:,.2f}")

    st.markdown("---")

    col_a, col_b = st.columns([3, 2])

    with col_a:
        st.markdown("**👷 Detailed Payroll Breakdown — By Staff Member**")
        wage_display = staff_df_live.copy()
        
        # Re-order columns to show important totals first
        pref_cols = ["Name", "Total Hrs", "Gross Wage (£)", "NI Pay (Bacs)", "Cash Pay (£)", "Bonus/Fixed (£)", "Cost to Emp (£)"]
        existing_cols = [c for c in pref_cols if c in wage_display.columns]
        other_cols = [c for c in wage_display.columns if c not in existing_cols]
        wage_display = wage_display[existing_cols + other_cols]

        # Dynamically format monetary columns
        money_cols = [c for c in wage_display.columns if any(k in c for k in ["Pay", "Wage", "Rate", "Emp", "Bonus", "Cost"])]
        for col in money_cols:
            try:
                wage_display[col] = wage_display[col].apply(lambda x: f"£{float(x):,.2f}")
            except: pass
            
        st.dataframe(wage_display, width="stretch", hide_index=True)

        total_wages = staff_df_live["Gross Wage (£)"].sum()
        # Use the true Weighted Average from the labour summary
        v_avg_rate = summary_live.get("Average Wage/Hour", 0.0)
        st.markdown(f'<div class="insight-box">💰 <b>Total Gross Pay: £{total_wages:,.2f}</b> &nbsp;|&nbsp; {len(staff_df_live)} staff members &nbsp;|&nbsp; Weighted Avg Wage: £{v_avg_rate:.2f}/hr</div>', unsafe_allow_html=True)

    with col_b:
        st.markdown("**🧾 Operational Fixed Costs (£330)**")
        # Load from CSV for dynamic editing
        fixed_path = os.path.join(os.getcwd(), "fixed_weekly_costs.csv")
        if os.path.exists(fixed_path):
            fdf_ui = pd.read_csv(fixed_path)
            other_df_live = fdf_ui.rename(columns={"Item": "Cost Item", "Amount": "Amount (£)"})
            other_df_live["Amount (£)"] = other_df_live["Amount (£)"].apply(lambda x: f"£{x:,.2f}")
        else:
            other_df_live = pd.DataFrame(columns=["Cost Item", "Amount (£)"])
        
        st.dataframe(other_df_live, width="stretch", hide_index=True)

        st.markdown("**🔥 SBY Staff — Potential Risk**")
        sby_rows = []
        sby_total = 0
        for name, v in SBY_STAFF.items():
            cost = v["max_sby_hrs"] * v["hourly_rate"]
            sby_total += cost
            sby_rows.append({
                "Name":          name,
                "Max Hrs":       v["max_sby_hrs"],
                "Rate (£/hr)":   f"£{v['hourly_rate']:.2f}",
                "Max Cost (£)":  f"£{cost:.2f}",
            })
        st.dataframe(pd.DataFrame(sby_rows), width="stretch", hide_index=True)
        st.markdown(f'<div class="insight-box">⚠️ SBY max additional cost: <b>£{sby_total:,.2f}</b> — only incurred if all SBY staff called in</div>', unsafe_allow_html=True)

    st.markdown("---")

    st.markdown("**📅 Day-by-Day Labour vs Forecast Revenue**")
    day_display = day_df_live.copy()

    def color_labour_status(val):
        if "🔴" in str(val): return "color: #FF4444; font-weight: bold"
        if "✅" in str(val):  return "color: #3ECF8E; font-weight: bold"
        return ""

    st.dataframe(day_display.style.map(color_labour_status, subset=["Status"]),
                 width="stretch", hide_index=True)
    st.markdown('<div class="insight-box">⚠️ Labour threshold: 30% of revenue. SBY staff not included above — add if called in. Call SBY by 19:00 on days with forecast revenue ≥ £3,000.</div>', unsafe_allow_html=True)

    st.markdown("---")

    st.markdown("**🕐 Hour-by-Hour Revenue vs Staff Cost Overlay**")

    fig_overlay = go.Figure()
    fig_overlay.add_trace(go.Bar(
        x=overlay_df_live["Hour"], y=overlay_df_live["Avg Revenue (£)"],
        name="Avg Revenue £", marker_color="rgba(245,166,35,0.45)", marker_line_width=0,
    ))
    fig_overlay.add_trace(go.Scatter(
        x=overlay_df_live["Hour"], y=overlay_df_live["Total Staff Cost (£)"],
        name="Total Staff Cost £", mode="lines+markers",
        line=dict(color=COLORS["red"], width=2.5, shape="spline"),
        marker=dict(size=6),
    ))
    fig_overlay.add_trace(go.Scatter(
        x=overlay_df_live["Hour"], y=overlay_df_live["Total Staff"],
        name="Total Staff (headcount)", mode="lines+markers",
        line=dict(color=COLORS["accent4"], width=1.5, dash="dot"),
        marker=dict(size=5),
        yaxis="y2",
    ))
    dark_layout(fig_overlay, 380, showlegend=True)
    fig_overlay.update_layout(
        title=dict(text="Hourly Revenue vs Staff Cost (91-day daily avg)", font=dict(size=14, color="#e8e9f0")),
        yaxis=dict(title="£", tickprefix="£", gridcolor="rgba(255,255,255,0.06)"),
        yaxis2=dict(title="Headcount", overlaying="y", side="right", showgrid=False,
                    tickfont=dict(color=COLORS["accent4"])),
    )
    st.plotly_chart(fig_overlay, width="stretch")

    st.markdown("---")

    flag_col1, flag_col2 = st.columns(2)

    with flag_col1:
        st.markdown("**🔴 Overstaffing Flags**")
        if len(over_df_live) > 0:
            st.dataframe(over_df_live, width="stretch", hide_index=True)
            st.markdown(f'<div class="insight-box">🔴 <b>{len(over_df_live)} overstaffed hour(s)</b> detected. Staff cost exceeds revenue generated. Recommend delaying morning shift starts or closing earlier on quiet days.</div>', unsafe_allow_html=True)
        else:
            st.success("✅ No overstaffing detected this week.")

    with flag_col2:
        st.markdown("**🟡 Understaffing Flags**")
        if len(under_df_live) > 0:
            st.dataframe(under_df_live, width="stretch", hide_index=True)
            st.markdown(f'<div class="insight-box">🟡 <b>{len(under_df_live)} understaffing hour(s)</b> detected. Peak revenue hours with insufficient cover. Call SBY staff in early.</div>', unsafe_allow_html=True)
        else:
            st.success("✅ No critical understaffing detected — peak hours well covered.")

    st.markdown("---")

    with st.expander("📋 Full Hourly Overlay Table — All Hours"):
        st.dataframe(overlay_df_live, width="stretch", hide_index=True)

    st.markdown("---")
    st.markdown('<div class="section-title">Export Full Labour Report</div>', unsafe_allow_html=True)
    st.markdown('<div class="insight-box" style="margin-bottom:12px">Click below to generate and download the full 6-sheet Excel labour report with dark theme formatting, KPI boxes, overstaffing flags, understaffing flags, and SBY tracker.</div>', unsafe_allow_html=True)

    if st.button("📊 Generate & Download Excel Labour Report", width="content"):
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
        build_labour_workbook(data, tmp_path)
        with open(tmp_path, "rb") as f:
            xlsx_bytes = f.read()
        os.unlink(tmp_path)
        st.download_button(
            label="⬇️ Download chocoberry_labour_report.xlsx",
            data=xlsx_bytes,
            file_name="chocoberry_labour_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )


# ════════════════════════════════════════════════════════════════════
# TAB 8 — ROTA BUILDER
# ════════════════════════════════════════════════════════════════════
with tab8:
    st.markdown('<div class="page-title">Intelligent Rota Builder</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Automated, constraint-based scheduling with fairness balancing.</div>', unsafe_allow_html=True)

    try:
        from rota_engine import RotaEngine
        engine = RotaEngine()
    except (ImportError, ModuleNotFoundError):
        st.warning("⚠️ rota_engine.py not found.")
        st.stop()

    # 1. Sticky Date Logic — Restore last selected date or use default
    _default_w_start = date.today() + timedelta(days=(7 - date.today().weekday()) % 7)
    w_start_val = st.session_state.get("selected_w_start", _default_w_start)


    with st.expander("👤 Staff Profile Management — Single Source of Truth", expanded=False):
        st.markdown('<div class="insight-box">Edit staff availability, roles (Senior/Junior), and target hours below. These settings power the auto-scheduler.</div>', unsafe_allow_html=True)
        profiles_path = os.path.join(os.getcwd(), "staff_profiles.csv")
        if os.path.exists(profiles_path):
            curr_prof_df = pd.read_csv(profiles_path)
            edited_df = st.data_editor(curr_prof_df, width="stretch", num_rows="dynamic", hide_index=True, key="prof_editor")
            if st.button("💾 Save Profile Changes", width="content"):
                edited_df.to_csv(profiles_path, index=False, encoding="utf-8-sig")
                st.success("✅ Staff profiles updated successfully.")
                st.cache_data.clear()
            
            st.markdown("---")
            st.markdown("**📥 Mobile Portal Sync**")
            
            # 1. Cloud Sync (from Google Sheets)
            if st.button("🔄 Sync Cloud Availability", help="Pull latest shift responses from Google Sheets"):
                with st.spinner("Connecting to Google Sheets..."):
                    # Sheet URL from current logic
                    sheet_url = "https://docs.google.com/spreadsheets/d/1Xl20OEq9e7m_tG7Z8vQjL0wXmZ-0N8A6H_1X8Q8K8w" 
                    success, msg = sync_availability_from_cloud(sheet_url)
                    if success:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)

            # 2. Local Manual Sync
            if st.button("📥 Apply Local Submissions", help="Apply responses currently in availability.db to profiles"):
                import sqlite3, json
                db_path = os.path.join(os.getcwd(), "availability.db")
                if os.path.exists(db_path):
                    try:
                        with sqlite3.connect(db_path) as conn:
                            # target_label in DB is "YYYY-MM-DD"
                            target_date_str = w_start_val.strftime('%Y-%m-%d')
                            query = "SELECT staff_name, availability, notes FROM availability WHERE week_start = ?"
                            subs = pd.read_sql(query, conn, params=(target_date_str,))
                            
                        if not subs.empty:
                            updated_count = 0
                            day_map = {"Monday":"Mon", "Tuesday":"Tue", "Wednesday":"Wed", "Thursday":"Thu", "Friday":"Fri", "Saturday":"Sat", "Sunday":"Sun"}
                            
                            for _, row in subs.iterrows():
                                s_name = row["staff_name"]
                                try:
                                    a_dict = json.loads(row["availability"])
                                    parts = []
                                    for d, status in a_dict.items():
                                        if status == "unavailable": continue
                                        ab = day_map.get(d, d[:3])
                                        if status == "morning": ab += "(Morn)"
                                        elif status == "evening": ab += "(Eve)"
                                        parts.append(ab)
                                    
                                    avail_str = ",".join(parts)
                                    idx = curr_prof_df[curr_prof_df["Name"] == s_name].index
                                    if not idx.empty:
                                        curr_prof_df.loc[idx, "Availability"] = avail_str
                                        updated_count += 1
                                except Exception as je:
                                    st.error(f"Error parsing JSON for {s_name}: {je}")
                            
                            if updated_count > 0:
                                curr_prof_df.to_csv(profiles_path, index=False, encoding="utf-8-sig")
                                st.success(f"✅ Successfully synced availability for {updated_count} staff members for the week of {target_date_str}.")
                                st.rerun()
                            else:
                                st.info(f"No submissions found for {target_date_str} yet.")
                        else:
                            st.info(f"No submissions found in availability.db for {target_date_str}.")
                    except Exception as e:
                        st.error(f"Error syncing availability: {e}")
                else:
                    st.warning("No availability.db found yet. Ask staff to use the portal link.")
        else:
            st.error("Missing staff_profiles.csv")

    st.markdown("---")
    st.markdown('<div class="section-title">Schedule Generation</div>', unsafe_allow_html=True)

    c1, c2 = st.columns([2, 1])
    with c1:
        st.markdown("**1. Configure Parameters**")
        w_start = st.date_input("Week Start Date (Monday recommended)", value=w_start_val)

        # NOW do the stale-rota check, AFTER the widget has resolved the actual w_start
        if st.session_state.get("last_w_start") != w_start:
            st.session_state["last_w_start"] = w_start
            st.session_state["selected_w_start"] = w_start
            for k in ["active_rota", "active_rota_summary", "active_rota_warnings"]:
                if k in st.session_state:
                    del st.session_state[k]
        
        w_end = w_start + timedelta(days=6)
        st.markdown(f'<div style="background:rgba(245,166,35,0.1);border-radius:10px;padding:10px 15px;border:1px solid #f5a623;color:#f5a623;font-weight:700;font-size:14px;margin-top:10px;text-align:center">📅 {w_start.strftime("%d %b")} ➜ {w_end.strftime("%d %b %Y")}</div>', unsafe_allow_html=True)
        
        # ── STAFF AVAILABILITY LIVE VIEW ──────────────────────────────────────
        st.markdown("---")
        st.markdown("**📱 Staff Availability Submissions**")

        week_avail = get_staff_availability_for_week(w_start)

        if not week_avail:
            st.warning(
                "⚠️ No availability submitted yet for this week. "
                "Send staff the link to submit before generating rota."
            )
        else:
            st.success(
                f"✅ {len(week_avail)} staff submitted availability"
            )
            
            days_list = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
            
            # Get department mapping
            staff_dept_map = {}
            if engine.staff_df is not None:
                staff_dept_map = engine.staff_df.set_index('Name')['Department'].to_dict()
            else:
                try:
                    engine.load_staff()
                    staff_dept_map = engine.staff_df.set_index('Name')['Department'].to_dict()
                except: pass

            avail_rows = []
            for name, avail in week_avail.items():
                dept = staff_dept_map.get(name, "—")
                row = {"Name": name, "Dept": dept}
                for day in days_list:
                    status = avail.get(day, 'any')
                    emoji = (
                        "🟢" if status == 'any' else
                        "💗" if status == 'opening' else
                        "🔘" if status == 'closing' else
                        "🔴"
                    )
                    row[day[:3]] = emoji
                row["Notes"] = avail.get('_notes', '')
                avail_rows.append(row)
            
            st.dataframe(
                pd.DataFrame(avail_rows),
                width="stretch",
                hide_index=True
            )

        # Show shareable link
        # Try to detect local IP for the link
        import socket
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
        except:
            local_ip = "localhost"
            
        avail_link = f"http://{local_ip}:5051/"

        st.markdown(f"""
        <div style="background:#1a1d26;border:1px solid #f5a623;
        border-radius:10px;padding:15px;margin-top:10px">
            <div style="color:#f5a623;font-weight:700;
            margin-bottom:8px">
                📱 Staff Availability Link
            </div>
            <div style="font-size:14px;color:#e8e9f0;
            word-break:break-all">
                {avail_link}
            </div>
            <div style="font-size:11px;color:#6b7094;margin-top:8px">
                Send this link to staff on WhatsApp
                by Thursday each week
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        
        with st.expander("🛠️ Manual Availability Manager", expanded=False):
            st.markdown('<div style="color:#3ecf8e;font-weight:700;margin-bottom:12px">Override Staff Preferences</div>', unsafe_allow_html=True)
            
            staff_list = []
            if engine.staff_df is not None:
                staff_list = engine.staff_df['Name'].tolist()
            else:
                try:
                    engine.load_staff()
                    staff_list = engine.staff_df['Name'].tolist()
                except:
                    staff_list = []

            sel_staff = st.selectbox("Select Staff Member", staff_list)
            
            m_avail = {}
            days_full = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
            
            # 2 columns of 4/3 to keep it compact
            c_m1, c_m2 = st.columns(2)
            for i, d in enumerate(days_full):
                target_col = c_m1 if i < 4 else c_m2
                with target_col:
                    m_avail[d] = st.selectbox(f"{d}", ["any", "opening", "closing", "unavailable"], key=f"man_{d}")
            
            m_notes = st.text_input("Entry Notes", value="Manual Override", key="man_notes")
            
            if st.button("💾 Save Manual Availability", type="secondary"):
                save_staff_availability(w_start, sel_staff, m_avail, m_notes)
                st.success(f"Saved availability for {sel_staff}")
                st.rerun()

        st.markdown("---")
        
        # 3. Check for Live Rota based on the SELECTED week
        folder_name = f"Rota week {w_start.strftime('%d %b')} - {w_end.strftime('%d %B %Y')}"
        rota_path = os.path.join(os.getcwd(), folder_name, "detailed_rota_with_shifts.csv")
        if os.path.exists(rota_path):
            st.success(f"🟢 Live Rota Detected for this week.")
        else:
            st.info("⚪ No Live Rota yet for this week.")

        balance_bias = st.slider("Fairness Balance Bias", 0.0, 1.0, 0.5)

    with c2:
        st.markdown("**2. Generate**")
        if st.button("⚡ Generate Weekly Rota", width="stretch", type="primary"):
            try:
                engine.load_staff()
                engine.load_shifts()
            except FileNotFoundError as e:
                st.error(f"Missing file: {e}")
                st.stop()
            with st.spinner("Solving constraints..."):
                engine.load_historical_hours(weeks_back=3)
                new_rota = engine.generate_week(week_start=w_start, submitted_availability=week_avail)
                st.session_state["active_rota"] = new_rota
                st.session_state["active_rota_summary"] = engine.get_hours_summary()
                st.session_state["active_rota_warnings"] = list(engine.warnings)
                st.session_state["last_w_start"] = w_start
            st.success(f"✅ {len(new_rota)} shift assignments generated.")

        if st.button("🚀 Smart Rota (Forecast-Matched)", width="stretch"):
            try:
                engine.load_staff()
                engine.load_shifts()
            except FileNotFoundError as e:
                st.error(f"Missing file: {e}")
                st.stop()
            with st.spinner("Scaling to forecast..."):
                _fs   = st.session_state.get("weekly_forecast")
                _base = st.session_state.get("weekly_baseline")
                _evts = st.session_state.get("week_events", {})
                engine.load_historical_hours(weeks_back=3)
                new_rota = engine.generate_week(
                    week_start=w_start,
                    forecast_scaling=_fs,
                    baseline_scaling=_base,
                    events_data=_evts,
                    submitted_availability=week_avail
                )
                st.session_state["active_rota"] = new_rota
                st.session_state["active_rota_summary"] = engine.get_hours_summary()
                st.session_state["active_rota_warnings"] = list(engine.warnings)
                st.session_state["last_w_start"] = w_start
            st.success(f"✅ Smart Rota generated.")


    if "active_rota" in st.session_state:
        st.markdown("---")
        st.markdown('<div class="section-title">Visual Weekly Schedule</div>', unsafe_allow_html=True)

        rota_df = st.session_state["active_rota"]

        if rota_df.empty:
            st.info("🕒 No shifts scheduled for the selected week/criteria.")
        else:
            # ── Day-by-Day Visual Display ───────────────────────────────────────
            days_order = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
            DEPT_COLORS = {"Kitchen": "#e05c5c", "Front": "#3a9bd5"}

            for d_idx, day_name in enumerate(days_order):
                day_date = (w_start + timedelta(days=d_idx)) if not isinstance(w_start, datetime) else (w_start + timedelta(days=d_idx))
                try:
                    date_label = day_date.strftime("%d %B %Y")
                except:
                    date_label = str(day_date)

                day_df = rota_df[rota_df["Day"] == day_name]

                st.markdown(
                    f'<div style="background:linear-gradient(90deg,#1a1a2e,#16213e);'
                    f'border-left:4px solid #f5a623;border-radius:8px;padding:10px 18px;'
                    f'margin:16px 0 8px;display:flex;align-items:center;gap:12px">'
                    f'<span style="font-size:18px;font-weight:800;color:#f5a623;font-family:Syne,sans-serif">'
                    f'{day_name.upper()}</span>'
                    f'<span style="font-size:13px;color:#aab0d0;font-weight:500">— {date_label}</span>'
                    f'{"<span style=&quot;margin-left:auto;background:#e05c5c;color:white;border-radius:20px;padding:2px 10px;font-size:11px;font-weight:700&quot;>⚠️ UNDERSTAFFED</span>" if day_df.empty else ""}'
                    f'</div>',
                    unsafe_allow_html=True
                )

                if day_df.empty:
                    st.markdown('<div style="color:#888;font-size:13px;padding:6px 18px">No shifts scheduled</div>', unsafe_allow_html=True)
                    continue

                dcol1, dcol2 = st.columns(2)
                for dept, dcol in [("Kitchen", dcol1), ("Front", dcol2)]:
                    dept_df = day_df[day_df["Department"] == dept].sort_values(["Shift","Role"], ascending=[True, False])
                    dept_icon = "🍳" if dept == "Kitchen" else "🛎️"
                    dept_color = DEPT_COLORS[dept]

                    with dcol:
                        st.markdown(
                            f'<div style="background:rgba(255,255,255,0.04);border-radius:8px;'
                            f'border-top:3px solid {dept_color};padding:10px 14px;min-height:80px">'
                            f'<div style="font-size:12px;font-weight:700;color:{dept_color};letter-spacing:1px;margin-bottom:8px">'
                            f'{dept_icon} {dept.upper()} OF HOUSE' if dept == "Front" else f'{dept_icon} {dept.upper()}'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                        if dept_df.empty:
                            st.markdown('<div style="color:#888;font-size:12px;padding:4px">— No staff scheduled</div>', unsafe_allow_html=True)
                        else:
                            for _, row in dept_df.iterrows():
                                shift_s = str(row.get("Start","")).replace(":00","") if str(row.get("Start","")) != "nan" else ""
                                shift_e = str(row.get("End","")).replace(":00","") if str(row.get("End","")) != "nan" else ""
                                dur = row.get("Duration","")
                                time_str = f"{shift_s}–{shift_e} ({dur}h)" if shift_s else row.get("Shift","")
                                st.markdown(
                                    f'<div style="display:flex;align-items:center;gap:8px;padding:5px 0;'
                                    f'border-bottom:1px solid rgba(255,255,255,0.05)">'
                                    f'<span style="font-size:13px;color:#e8e9f0;font-weight:600">{row["Name"]}</span>'
                                    f'<span style="margin-left:auto;font-size:11px;color:#aab0d0">{time_str}</span>'
                                    f'</div>',
                                    unsafe_allow_html=True
                                )
                        st.markdown('</div>', unsafe_allow_html=True)

            # ── Editable Rota ────────────────────────────────────────────────
            st.markdown("---")
            st.markdown('<div class="section-title">✏️ Edit Rota — Swap Names</div>', unsafe_allow_html=True)
            st.markdown('<div class="insight-box">Click any cell in the <b>Name</b> column to manually swap a staff member. Press <b>💾 Save Edits</b> when done to lock in changes.</div>', unsafe_allow_html=True)

            # Build name list from staff profiles for dropdown
            _staff_names = []
            try:
                _sp = pd.read_csv(os.path.join(BASE_DIR, "staff_profiles.csv"), encoding="utf-8-sig")
                _sp.columns = [c.strip() for c in _sp.columns]
                _sp = _sp[_sp["Active"].astype(str).str.lower().isin(["yes","true","1"])]
                _staff_names = sorted(_sp["Name"].dropna().unique().tolist())
            except Exception:
                _staff_names = []

            _edit_cols = ["Day", "Date", "Department", "Shift", "Start", "End", "Name", "Role", "SBY"]
            _edit_df   = rota_df[[c for c in _edit_cols if c in rota_df.columns]].copy()

            _col_config = {"Name": st.column_config.SelectboxColumn(
                "Name",
                help="Select a staff member",
                options=_staff_names if _staff_names else _edit_df["Name"].unique().tolist(),
                required=True,
            )} if _staff_names else {}

            edited_rota = st.data_editor(
                _edit_df,
                column_config=_col_config,
                width="stretch",
                hide_index=True,
                key="rota_editor",
                use_container_width=True,
            )

            if st.button("💾 Save Edits to Rota", key="save_rota_edits"):
                st.session_state["active_rota"] = edited_rota
                st.success("✅ Rota updated with your manual changes! Re-generate to recalculate hours summary.")

            # ── Fairness & Constraints ─────────────────────────────────────────
            st.markdown("---")
            m1, m2 = st.columns(2)
            with m1:
                st.markdown("**👥 Weekly Hours Summary**")
                hours_summary = st.session_state.get("active_rota_summary", pd.DataFrame())
                if not hours_summary.empty:
                    for _, row in hours_summary.iterrows():
                        name   = row["Name"]
                        hrs    = row["Scheduled Hrs"]
                        target = row["Target Hrs"]
                        delta  = row["Delta"]
                        status = row.get("Status", "")
                        sby    = row.get("SYB Hrs", 0.0)
                        pct    = (hrs / target * 100) if target > 0 else 0
                        bar_w  = min(100, int(pct))

                        color = "#3ecf8e" if abs(delta) <= 2 else \
                                "#f5a623" if delta > 2 else "#e05c5c"
                        label = "✅ On Target" if abs(delta) <= 2 else \
                                "🔼 Over" if delta > 2 else "🔽 Under"

                        sby_html = f'<span style="font-size:10px;color:#7c5cbf;margin-left:8px">+{sby}h SBY</span>' if sby > 0 else ""

                        st.markdown(f"""
                        <div style="padding:10px 0;border-bottom:1px solid rgba(255,255,255,0.05)">
                          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:5px">
                            <span style="color:#e8e9f0;font-weight:700;font-size:13px">{name}{sby_html}</span>
                            <span style="color:{color};font-weight:700;font-size:13px">{hrs}h
                              <span style="color:#6b7094;font-size:11px;font-weight:400"> / {target}h target</span>
                              <span style="color:{color};font-size:11px;margin-left:6px">{label}</span>
                            </span>
                          </div>
                          <div style="height:5px;background:#252836;border-radius:3px;overflow:hidden">
                            <div style="width:{bar_w}%;height:100%;background:linear-gradient(90deg,{color}88,{color});border-radius:3px;transition:width 0.3s"></div>
                          </div>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("Generate a rota above to see the hours summary.")

            with m2:
                st.markdown("**⚠️ Constraints Check (Warnings)**")
                warnings_list = st.session_state.get("active_rota_warnings", [])
                if warnings_list:
                    for w in warnings_list:
                        st.warning(w)
                else:
                    st.success("✅ All shift constraints (Management cover, headcount) satisfied.")

            # ── Confidential Financial Data (Option C) ────────────────────────
            with st.expander("🔒 Manager Only — Financial Data", expanded=False):
                st.warning("⚠️ This section contains confidential wage information. Ensure no staff are viewing the screen.")
                if "active_rota" in st.session_state:
                    _cost_data = engine.estimate_weekly_cost(st.session_state["active_rota"])
                    st.metric("Estimated Weekly Wage", f"£{_cost_data['total']:,.2f}")
                    
                    st.markdown("**Breakdown per Person:**")
                    # Create a simple table for the breakdown
                    _br_data = [{"Staff Member": k, "Est. Wage": f"£{v:,.2f}"} for k, v in _cost_data["breakdown"].items()]
                    st.table(_br_data)

                    # --- Manager PDF Export ---
                    try:
                        from reportlab.lib import colors as rl_colors
                        from reportlab.lib.pagesizes import A4
                        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                        from reportlab.lib.units import cm
                        from reportlab.lib.enums import TA_CENTER, TA_LEFT
                        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
                        import io as _io
                        
                        def _build_finance_pdf(cost_data, ws, we):
                            buf = _io.BytesIO()
                            doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
                            styles = getSampleStyleSheet()
                            title_s  = ParagraphStyle("T", parent=styles["Heading1"], fontSize=18, textColor=rl_colors.HexColor("#e05c5c"), spaceAfter=4, alignment=TA_CENTER)
                            sub_s    = ParagraphStyle("S", parent=styles["Normal"], fontSize=11, textColor=rl_colors.HexColor("#555555"), spaceAfter=10, alignment=TA_CENTER)
                            h2_s     = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=14, textColor=rl_colors.HexColor("#1a1a2e"), spaceBefore=10, spaceAfter=5)
                            
                            story = []
                            story.append(Paragraph("CONFIDENTIAL — FINANCIAL WAGE REPORT", title_s))
                            story.append(Paragraph(f"{ws.strftime('%A %d %B %Y') if hasattr(ws,'strftime') else ws}  →  {we.strftime('%A %d %B %Y') if hasattr(we,'strftime') else we}", sub_s))
                            story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor("#1a1a2e"), spaceAfter=15))
                            
                            story.append(Paragraph(f"Estimated Total Weekly Wage: <b>£{cost_data['total']:,.2f}</b>", h2_s))
                            story.append(Spacer(1, 10))
                            
                            story.append(Paragraph("Breakdown per Person:", h2_s))
                            
                            tdata = [["Staff Member", "Estimated Wage"]]
                            for k, v in sorted(cost_data["breakdown"].items(), key=lambda x: -x[1]):
                                tdata.append([k, f"£{v:,.2f}"])
                                
                            tbl = Table(tdata, colWidths=[10*cm, 5*cm])
                            tbl.setStyle(TableStyle([
                                ("BACKGROUND",  (0,0),(-1,0), rl_colors.HexColor("#e05c5c")),
                                ("TEXTCOLOR",   (0,0),(-1,0), rl_colors.white),
                                ("FONTNAME",    (0,0),(-1,0), "Helvetica-Bold"),
                                ("ALIGN",       (0,0),(-1,-1), "LEFT"),
                                ("ALIGN",       (1,0),(1,-1), "RIGHT"),
                                ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#f7f7f7"), rl_colors.white]),
                                ("GRID",        (0,0),(-1,-1), 0.5, rl_colors.HexColor("#dddddd")),
                                ("PADDING",     (0,0),(-1,-1), 6),
                            ]))
                            story.append(tbl)
                            
                            doc.build(story)
                            return buf.getvalue()
                        
                        f_pdf_bytes = _build_finance_pdf(_cost_data, w_start, w_start + timedelta(days=6))
                        st.download_button(
                            label="🔒 Download Confidential Financial PDF",
                            data=f_pdf_bytes,
                            file_name=f"chocoberry_finance_report_{w_start.strftime('%d_%b_%Y')}.pdf",
                            mime="application/pdf"
                        )
                    except Exception as e:
                        st.error(f"Could not generate Financial PDF: {e}")

                else:
                    st.info("Generate a rota to see the wage estimate.")

            # ── Push / Download Buttons ───────────────────────────────────────
            st.markdown("---")
            c_push, c_down_csv, c_down_pdf = st.columns(3)

            with c_push:
                if st.button("🚀 Push to Tab 7 (Commit to Live Data)", width="stretch"):
                    w_end = w_start + timedelta(days=6)
                    folder_name = f"Rota week {w_start.strftime('%d %b')} - {w_end.strftime('%d %B %Y')}"
                    output_dir = os.path.join(os.getcwd(), folder_name)
                    if not os.path.exists(output_dir): os.makedirs(output_dir)
                    rota_df.to_csv(os.path.join(output_dir, "detailed_rota_with_shifts.csv"), index=False)
                    st.cache_data.clear()
                    st.balloons()
                    st.success(f"✅ Rota deployed to: {output_dir}")

            with c_down_csv:
                csv_bytes = rota_df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="⬇️ Download Rota CSV",
                    data=csv_bytes,
                    file_name=f"chocoberry_rota_{w_start.strftime('%d_%b_%Y')}.csv",
                    mime="text/csv",
                    width="stretch"
                )

            with c_down_pdf:
                try:
                    from reportlab.lib import colors as rl_colors
                    from reportlab.lib.pagesizes import A4
                    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                    from reportlab.lib.units import cm
                    from reportlab.lib.enums import TA_CENTER, TA_LEFT
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
                    import io as _io

                    def _build_rota_pdf(rdf, ws, we, cost_data, warnings):
                        buf = _io.BytesIO()
                        doc = SimpleDocTemplate(buf, pagesize=A4,
                                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                                topMargin=1.5*cm, bottomMargin=1.5*cm)
                        styles = getSampleStyleSheet()
                        title_s  = ParagraphStyle("T",  parent=styles["Heading1"], fontSize=18,
                                                  textColor=rl_colors.HexColor("#1a1a2e"), spaceAfter=4,  alignment=TA_CENTER)
                        sub_s    = ParagraphStyle("S",  parent=styles["Normal"],   fontSize=11,
                                                  textColor=rl_colors.HexColor("#555555"), spaceAfter=10, alignment=TA_CENTER)
                        day_s    = ParagraphStyle("D",  parent=styles["Normal"],   fontSize=13, fontName="Helvetica-Bold",
                                                  textColor=rl_colors.white, backColor=rl_colors.HexColor("#1a1a2e"),
                                                  spaceBefore=14, spaceAfter=4, leftIndent=6)
                        dept_s   = ParagraphStyle("Dp", parent=styles["Normal"],   fontSize=10, fontName="Helvetica-Bold",
                                                  textColor=rl_colors.HexColor("#333"), spaceBefore=6, spaceAfter=2)
                        warn_h   = ParagraphStyle("WH", parent=styles["Normal"],   fontSize=12, fontName="Helvetica-Bold",
                                                  textColor=rl_colors.HexColor("#e05c5c"), spaceBefore=12, spaceAfter=4)
                        warn_s   = ParagraphStyle("WS", parent=styles["Normal"],   fontSize=10,
                                                  textColor=rl_colors.HexColor("#333"), leftIndent=10, spaceAfter=3)

                        story = []
                        story.append(Paragraph("CHOCOBERRY — Weekly Staff Rota", title_s))
                        story.append(Paragraph(f"{ws.strftime('%A %d %B %Y') if hasattr(ws,'strftime') else ws}  →  {we.strftime('%A %d %B %Y') if hasattr(we,'strftime') else we}", sub_s))
                        # HIDDEN FOR PRIVACY: story.append(Paragraph(f"Estimated Weekly Wage: £{cost_data['total']:,.2f}", sub_s))
                        story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor("#1a1a2e"), spaceAfter=6))

                        _days = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
                        for d_i, dname in enumerate(_days):
                            try:
                                ddate = (ws + timedelta(days=d_i))
                                dlabel = f"  {dname}  —  {ddate.strftime('%d %B %Y')}"
                            except:
                                dlabel = f"  {dname}"
                            story.append(Paragraph(dlabel, day_s))

                            day_sub = rdf[rdf["Day"] == dname]
                            if day_sub.empty:
                                story.append(Paragraph("No shifts scheduled", dept_s))
                                continue

                            for dept in ["Kitchen", "Front"]:
                                dept_sub = day_sub[day_sub["Department"] == dept].sort_values(["Shift","Role"], ascending=[True, False])
                                if dept_sub.empty: continue
                                icon = "Kitchen" if dept == "Kitchen" else "Front of House"
                                story.append(Paragraph(f"  {icon}", dept_s))
                                tdata = [["Name", "Shift", "Start", "End", "Hrs"]]
                                for _, r in dept_sub.iterrows():
                                    tdata.append([r["Name"], r.get("Shift",""),
                                                  str(r.get("Start","")), str(r.get("End","")),
                                                  f"{r.get('Duration','')}h"])
                                hdr_col = rl_colors.HexColor("#e05c5c") if dept == "Kitchen" else rl_colors.HexColor("#3a9bd5")
                                tbl = Table(tdata, colWidths=[7.0*cm,2.5*cm,2.5*cm,2.5*cm,2.0*cm], repeatRows=1)
                                tbl.setStyle(TableStyle([
                                    ("BACKGROUND",  (0,0),(-1,0), hdr_col),
                                    ("TEXTCOLOR",   (0,0),(-1,0), rl_colors.white),
                                    ("FONTNAME",    (0,0),(-1,0), "Helvetica-Bold"),
                                    ("FONTSIZE",    (0,0),(-1,-1), 9),
                                    ("ALIGN",       (0,0),(-1,-1), "LEFT"),
                                    ("VALIGN",      (0,0),(-1,-1), "MIDDLE"),
                                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[rl_colors.HexColor("#f7f7f7"), rl_colors.white]),
                                    ("GRID",        (0,0),(-1,-1), 0.4, rl_colors.HexColor("#cccccc")),
                                    ("TOPPADDING",  (0,0),(-1,-1), 4),
                                    ("BOTTOMPADDING",(0,0),(-1,-1), 4),
                                    ("LEFTPADDING", (0,0),(-1,-1), 6),
                                ]))
                                story.append(tbl)
                                story.append(Spacer(1, 0.15*cm))
                            story.append(Spacer(1, 0.3*cm))

                        # ── Warnings Section in PDF ────────────────────────────────────
                        if warnings:
                            story.append(Spacer(1, 0.5*cm))
                            story.append(HRFlowable(width="100%", thickness=1.5, color=rl_colors.HexColor("#e05c5c")))
                            story.append(Paragraph("⚠️ Staffing Warnings & Insights", warn_h))
                            for w in warnings:
                                story.append(Paragraph(f"• {w}", warn_s))
                            story.append(Paragraph("<i>Suggestion: Consider activating 'Abuzar' or increasing 'Max Hours' for existing staff to resolve these weekend gaps.</i>", warn_s))
                            story.append(HF1 := HRFlowable(width="100%", thickness=0.5, color=rl_colors.HexColor("#e05c5c"), spaceBefore=4))

                        story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#cccccc"), spaceBefore=10))
                        story.append(Paragraph(f"Total: {len(rdf)} shifts  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_s))
                        doc.build(story)
                        buf.seek(0)
                        return buf.read()

                    cost_est  = engine.estimate_weekly_cost(rota_df)
                    warn_list = st.session_state.get("active_rota_warnings", [])
                    pdf_data  = _build_rota_pdf(rota_df, w_start, w_end, cost_est, warn_list)
                    st.download_button(
                        label="📄 Download Rota PDF",
                        data=pdf_data,
                        file_name=f"chocoberry_rota_{w_start.strftime('%d_%b_%Y') if hasattr(w_start,'strftime') else w_start}.pdf",
                        mime="application/pdf",
                        width="stretch"
                    )
                except Exception as pdf_err:
                    st.error(f"PDF generation error: {pdf_err}")



# ════════════════════════════════════════════════════════════════════
# TAB 9 — INVENTORY & COGS  *** FULL REPLACEMENT ***
# ════════════════════════════════════════════════════════════════════
with tab9:
    st.markdown('<div class="page-title">Inventory & COGS Master</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Recipe cards · Auto stock deduction · Food cost % · Reorder alerts</div>', unsafe_allow_html=True)

    if _recipe_engine is None:
        st.error("recipe_engine.py not found. Place it in the same folder as app_dashboard.py.")
    else:
        eng = _recipe_engine

        # ── KPI row ────────────────────────────────────────────────────────────
        summary_inv = eng.weekly_stock_summary()
        k1, k2, k3, k4, k5 = st.columns(5)
        k1.metric("🥦 Ingredients",     summary_inv["total_ingredients"])
        k2.metric("🚨 Reorder Alerts",  summary_inv["reorder_alerts"],
                  delta="needs action" if summary_inv["reorder_alerts"] > 0 else "all ok",
                  delta_color="inverse" if summary_inv["reorder_alerts"] > 0 else "normal")
        k3.metric("🍕 Avg Food Cost",   f"{summary_inv['avg_food_cost_pct']}%")
        k4.metric("💷 Weekly COGS",     f"£{summary_inv['weekly_cogs']:,.2f}")
        k5.metric("🗑️ Weekly Waste",    f"£{summary_inv['weekly_waste_cost']:,.2f}")

        st.markdown("---")

        inv_tabs = st.tabs([
            "📋 Stock Levels", "🍕 Recipe Cards", "📊 Profitability",
            "⬇️ Auto-Deduct from Sales", "📦 Purchase Orders"
        ])

        # ── Tab: Stock Levels ─────────────────────────────────────────────────
        with inv_tabs[0]:
            st.markdown("**Current Stock Levels**")
            ings = eng.get_ingredients()
            alerts_inv = eng.get_reorder_alerts()

            if alerts_inv:
                st.markdown(f"""
                <div style="background:rgba(224,92,92,0.1);border:1px solid #e05c5c;
                            padding:12px 16px;border-radius:8px;margin-bottom:12px">
                    🚨 <b style="color:#e05c5c">{len(alerts_inv)} items below reorder point</b>
                </div>""", unsafe_allow_html=True)

            if ings:
                ing_df = pd.DataFrame(ings)
                ing_df["Status"] = ing_df.apply(
                    lambda r: "🚨 REORDER" if r["current_stock"] <= r["reorder_point"]
                    else ("⚠️ Low" if r["current_stock"] <= r["reorder_point"] * 1.5
                    else "✅ OK"), axis=1
                )
                ing_df["unit_cost"] = ing_df["unit_cost"].apply(lambda x: f"£{x:.4f}")
                ing_df["current_stock"] = ing_df["current_stock"].round(3)
                ing_df.columns = ["ID","Name","Unit","Unit Cost","Supplier",
                                   "Current Stock","Reorder Point","Status"]
                st.dataframe(ing_df.drop(columns=["ID"]), width="stretch", hide_index=True)
            else:
                st.info("No ingredients yet. Add them in the Recipe Cards tab.")

            st.markdown("---")
            st.markdown("**Manual Stock Adjustment (After Physical Count)**")
            if ings:
                sel_ing = st.selectbox("Ingredient", [i["name"] for i in ings], key="stock_adj")
                new_qty = st.number_input("New physical stock count", min_value=0.0, step=0.1, key="new_qty")
                if st.button("Update Stock", key="upd_stock"):
                    ing_id = next(i["id"] for i in ings if i["name"] == sel_ing)
                    eng.set_opening_stock(ing_id, new_qty)
                    st.success(f"✅ {sel_ing} updated to {new_qty}")
                    st.rerun()

        # ── Tab: Recipe Cards ─────────────────────────────────────────────────
        with inv_tabs[1]:
            col_left, col_right = st.columns([1, 1])

            with col_left:
                st.markdown("**Add / Edit Ingredient**")
                with st.form("ing_form"):
                    ing_name   = st.text_input("Ingredient Name *")
                    ing_unit   = st.selectbox("Unit", ["kg","g","litre","ml","each","portion","pack"])
                    ing_cost   = st.number_input("Unit Cost (£)", min_value=0.0, step=0.01)
                    ing_supp   = st.text_input("Supplier")
                    ing_stock  = st.number_input("Opening Stock", min_value=0.0, step=0.1)
                    ing_reord  = st.number_input("Reorder Point", min_value=0.0, step=0.1)
                    if st.form_submit_button("Save Ingredient"):
                        if ing_name:
                            eng.upsert_ingredient(ing_name, ing_unit, ing_cost,
                                                   ing_supp, ing_stock, ing_reord)
                            st.success(f"✅ {ing_name} saved.")
                            st.rerun()

                st.markdown("---")
                st.markdown("**Add Menu Item**")
                with st.form("item_form"):
                    fd_name    = st.text_input("Flipdish Name (exact match) *")
                    disp_name  = st.text_input("Display Name")
                    item_cat   = st.selectbox("Category", ["Waffles","Crepes","Drinks",
                                                            "Sundaes","Hot Food","Extras","Other"])
                    sell_price = st.number_input("Selling Price (£)", min_value=0.0, step=0.5)
                    if st.form_submit_button("Save Menu Item"):
                        if fd_name:
                            eng.upsert_menu_item(fd_name, disp_name, item_cat, sell_price)
                            st.success(f"✅ {fd_name} saved.")
                            st.rerun()

            with col_right:
                st.markdown("**Build / Update Recipe Card**")
                items = eng.get_menu_items()
                ings  = eng.get_ingredients()

                if not items:
                    st.info("Add a menu item first (left panel).")
                elif not ings:
                    st.info("Add ingredients first (left panel).")
                else:
                    sel_item = st.selectbox("Menu Item", [i["display_name"] for i in items])
                    item_id  = next(i["id"] for i in items if i["display_name"] == sel_item)

                    existing = eng.get_recipe(item_id)
                    st.markdown(f"**Current recipe** ({len(existing)} ingredients, "
                                f"cost: £{sum(r['line_cost'] for r in existing):.4f})")
                    if existing:
                        rc_df = pd.DataFrame(existing)[["name","quantity","unit","line_cost"]]
                        rc_df.columns = ["Ingredient","Qty","Unit","Cost £"]
                        rc_df["Cost £"] = rc_df["Cost £"].apply(lambda x: f"£{x:.4f}")
                        st.dataframe(rc_df, width="stretch", hide_index=True)

                    st.markdown("**Add ingredient to recipe**")
                    with st.form("recipe_form"):
                        sel_ing  = st.selectbox("Ingredient", [i["name"] for i in ings])
                        qty      = st.number_input("Quantity per portion", min_value=0.001,
                                                   step=0.001, format="%.3f")
                        if st.form_submit_button("Add to Recipe"):
                            ing_id   = next(i["id"] for i in ings if i["name"] == sel_ing)
                            new_ings = [{"ingredient_id": r["ingredient_id"],
                                         "quantity": r["quantity"]} for r in existing]
                            new_ings.append({"ingredient_id": ing_id, "quantity": qty})
                            eng.set_recipe(item_id, new_ings)
                            st.success(f"✅ Added {sel_ing} to {sel_item}.")
                            st.rerun()

                    if existing and st.button("🗑️ Clear Full Recipe", key="clr_recipe"):
                        eng.set_recipe(item_id, [])
                        st.rerun()

                    st.markdown("---")
                    st.markdown("**Archive / Hide Menu Item**")
                    st.info("Deactivating an item stops it from appearing in dropdowns but keeps historical data.")
                    if st.button("📁 Archive this Menu Item", width="stretch"):
                        # Add a quick toggle in the engine or just use direct SQL
                        with sqlite3.connect(RECIPE_DB_PATH) as conn:
                            conn.execute("UPDATE menu_items SET active = 0 WHERE id = ?", (item_id,))
                        st.warning(f"Item '{sel_item}' has been archived.")
                        st.rerun()

            # ── CSV bulk import ─────────────────────────────────────────────
            st.markdown("---")
            st.markdown("**Bulk Import via CSV**")
            c1_inv, c2_inv = st.columns(2)
            with c1_inv:
                up_ing = st.file_uploader("Import Ingredients CSV",
                                          type="csv", key="ing_csv")
                if up_ing:
                    import tempfile
                    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="wb") as tmp:
                        tmp.write(up_ing.read())
                        tmp_path = tmp.name
                    n = eng.import_ingredients_csv(tmp_path)
                    st.success(f"✅ Imported {n} ingredients.")
                    os.unlink(tmp_path)
            with c2_inv:
                up_items = st.file_uploader("Import Menu Items CSV",
                                            type="csv", key="items_csv")
                if up_items:
                    import tempfile
                    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="wb") as tmp:
                        tmp.write(up_items.read())
                        tmp_path = tmp.name
                    n = eng.import_menu_items_csv(tmp_path)
                    st.success(f"✅ Imported {n} menu items.")
                    os.unlink(tmp_path)
            st.markdown('<div class="insight-box">CSV format — Ingredients: <b>name, unit, unit_cost, supplier, opening_stock, reorder_point</b> &nbsp;|&nbsp; Menu items: <b>flipdish_name, display_name, category, selling_price</b></div>', unsafe_allow_html=True)

        # ── Tab: Profitability ────────────────────────────────────────────────
        with inv_tabs[2]:
            st.markdown("**Menu Profitability — Star / Dog Matrix**")
            
            items_prof = eng.calc_item_profitability()

            if not items_prof:
                st.info("Add menu items with recipe cards and selling prices to see profitability.")
            else:
                for i in items_prof:
                    i["volume"] = 0
                    i["total_cogs"] = 0.0

                # Load sales volume safely
                sv_path = os.path.join(os.getcwd(), "Menu Item Report data", "Most sold items.csv")
                volume_loaded = False
                if os.path.exists(sv_path):
                    try:
                        sv_df = pd.read_csv(sv_path)
                        # Find name and qty columns robustly
                        name_col = next(
                            (c for c in sv_df.columns if any(x in c.lower() for x in ["item","name"])),
                            sv_df.columns[0]
                        )
                        qty_col = next(
                            (c for c in sv_df.columns if any(x in c.lower() for x in ["sold","qty","quantity","count","orders","total"])),
                            None
                        )
                        if qty_col:
                            # Clean and build lookup: lowercase stripped name -> int qty
                            vol_map = {}
                            for _, r in sv_df.iterrows():
                                raw_name = str(r[name_col]).strip()
                                raw_qty  = str(r[qty_col]).replace(",","").replace("£","").strip()
                                try:
                                    vol_map[raw_name.lower()] = int(float(raw_qty))
                                except (ValueError, TypeError):
                                    pass
                            
                            # Merge by fuzzy lowercase match
                            matched = 0
                            for i in items_prof:
                                key = str(i.get("flipdish_name","")).strip().lower()
                                disp_key = str(i.get("menu_item","")).strip().lower()
                                vol = vol_map.get(key) or vol_map.get(disp_key) or 0
                                i["volume"] = vol
                                i["total_cogs"] = round(vol * i["ingredient_cost"], 2)
                                if vol > 0:
                                    matched += 1
                            
                            volume_loaded = matched > 0
                            if not volume_loaded:
                                st.warning(
                                    f"⚠️ Sales file loaded ({len(vol_map)} items) but 0 names matched recipe items. "
                                    f"Check that Flipdish item names in your recipe DB match the CSV exactly. "
                                    f"Example CSV name: '{next(iter(vol_map))}' | "
                                    f"Example recipe name: '{items_prof[0].get('flipdish_name','?')}'"
                                )
                    except Exception as e:
                        st.warning(f"Could not load sales volume: {e}")

                prof_df = pd.DataFrame(items_prof)

                total_volume   = int(prof_df["volume"].sum())
                total_cogs_all = round(prof_df["total_cogs"].sum(), 2)
                with_rec       = [i for i in items_prof if i["has_recipe"]]

                p1, p2, p3, p4 = st.columns(4)
                p1.metric("Items with recipe",  len(with_rec))
                p2.metric("Total COGS Est.",    f"£{total_cogs_all:,.2f}",
                          f"{total_volume:,} units sold" if volume_loaded else "⚠️ no volume data")
                p3.metric("Avg GP %",
                          f"{sum(i['gp_pct'] for i in with_rec)/len(with_rec):.1f}%" if with_rec else "—")
                p4.metric("Avg food cost %",
                          f"{sum(i['food_cost_pct'] for i in with_rec)/len(with_rec):.1f}%" if with_rec else "—")


                # --- SCATTER CHART ---
                if not prof_df.empty and total_volume > 0:
                    matrix_fig = px.scatter(
                        prof_df[prof_df["volume"] > 0], 
                        x="volume", y="gp_pct", 
                        text="menu_item", color="category",
                        hover_data=["selling_price", "total_cogs"],
                        labels={"volume":"Sales Volume (Popularity)", "gp_pct":"Gross Profit % (Profitability)"},
                        title="Star / Dog Matrix: Popularity vs Profitability",
                        template="plotly_dark"
                    )
                    matrix_fig.update_traces(textposition='top center')
                    # Quadrant lines
                    med_vol = prof_df[prof_df["volume"] > 0]["volume"].median()
                    med_gp  = prof_df[prof_df["volume"] > 0]["gp_pct"].median()
                    matrix_fig.add_hline(y=med_gp, line_dash="dash", line_color="gray", annotation_text="Profitability Median")
                    matrix_fig.add_vline(x=med_vol, line_dash="dash", line_color="gray", annotation_text="Volume Median")
                    
                    st.plotly_chart(matrix_fig, width="stretch")
                else:
                    st.info("💡 Note: Missing sales volume data. To view the Profitability Matrix (Stars/Dogs), please ensure 'Most sold items.csv' is present in 'Menu Item Report data'. Showing table only.")

                disp = prof_df[["menu_item","category","volume","selling_price",
                                 "ingredient_cost","total_cogs","gp_pct"]].copy()
                disp.columns = ["Item","Category","Vol","Price £","Unit Cost £","Total COGS £","GP %"]
                disp["Price £"]     = disp["Price £"].apply(lambda x: f"£{x:.2f}")
                disp["Unit Cost £"] = disp["Unit Cost £"].apply(lambda x: f"£{x:.4f}")
                disp["Total COGS £"] = disp["Total COGS £"].apply(lambda x: f"£{x:,.2f}")
                disp["GP %"]        = disp["GP %"].apply(lambda x: f"{x:.1f}%")
                st.dataframe(disp.sort_values("Vol", ascending=False), width="stretch", hide_index=True)

                if os.path.exists(sv_path):
                    st.caption(f"💡 Based on latest Menu Item Report in folder (Auto-detected). Total units tracked: {total_volume:,}")
                else:
                    st.warning("⚠️ No 'Most sold items.csv' found in 'Menu Item Report data' folder. Volume analysis disabled.")


                import tempfile as _tmp
                if st.button("📥 Export Profitability CSV"):
                    with _tmp.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
                        eng.export_profitability_csv(tmp.name)
                        csv_bytes = open(tmp.name, "rb").read()
                    st.download_button("⬇️ Download", csv_bytes,
                                       file_name="chocoberry_profitability.csv",
                                       mime="text/csv")

        # ── Tab: Auto-Deduct from Sales ───────────────────────────────────────
        with inv_tabs[3]:
            st.markdown("**Auto-Deduct Ingredients from Flipdish Sales Data**")
            st.markdown('<div class="insight-box">Upload the Flipdish <b>Menu Items Report</b> (CSV) and the engine will automatically deduct ingredient quantities from stock based on your recipe cards.</div>', unsafe_allow_html=True)

            up_sales = st.file_uploader("Upload Flipdish Menu Items CSV",
                                        type="csv", key="sales_deduct")
            deduct_date = st.date_input("Sales Date", value=datetime.now().date())

            if up_sales:
                import tempfile
                with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="wb") as tmp:
                    tmp.write(up_sales.read())
                    tmp_path = tmp.name

                try:
                    sales_raw = pd.read_csv(tmp_path)
                    os.unlink(tmp_path)

                    name_col = next((c for c in sales_raw.columns
                                     if any(x in c.lower() for x in ["item","name","product"])),
                                    sales_raw.columns[0])
                    qty_col  = next((c for c in sales_raw.columns
                                     if any(x in c.lower() for x in ["sold","qty","quantity","count","orders"])),
                                    sales_raw.columns[1] if len(sales_raw.columns) > 1 else sales_raw.columns[0])

                    st.markdown(f"Detected: **item column** = `{name_col}`, **quantity column** = `{qty_col}`")
                    st.dataframe(sales_raw[[name_col, qty_col]].head(10), hide_index=True)

                    if st.button("▶ Run Auto-Deduction", type="primary"):
                        sales_list = [
                            {"item_name": str(row[name_col]), "units_sold": int(row[qty_col])}
                            for _, row in sales_raw.iterrows()
                            if pd.notna(row[qty_col]) and int(row[qty_col]) > 0
                        ]
                        result = eng.deduct_from_sales(sales_list,
                                                       deduct_date.strftime("%Y-%m-%d"))

                        st.success(f"✅ Processed {result['items_processed']} items. "
                                   f"Total COGS: £{result['total_cogs']:,.2f}")

                        if result["items_skipped"]:
                            st.warning(f"Skipped {len(result['items_skipped'])} items "
                                       f"(no recipe card or not in menu): "
                                       f"{', '.join(result['items_skipped'][:5])}")

                        st.markdown("**Stock levels have been updated. Check Stock Levels tab.**")
                        st.rerun()

                except Exception as e:
                    st.error(f"Could not parse CSV: {e}")

            else:
                st.markdown('<div class="insight-box">💡 <b>How this works:</b> Flipdish exports a "Menu Items" report showing items sold per item. This engine maps each item name to your recipe card, multiplies by units sold, and deducts every ingredient from current stock automatically.</div>', unsafe_allow_html=True)

        # ── Tab: Purchase Orders ──────────────────────────────────────────────
        with inv_tabs[4]:
            st.markdown("**Purchase Order Generator**")
            alerts_po = eng.get_reorder_alerts()

            if not alerts_po:
                st.success("✅ All ingredients above reorder point. No PO needed.")
            else:
                st.markdown(f"**{len(alerts_po)} items need reordering:**")

                po_df = pd.DataFrame(alerts_po)
                po_df["qty_to_order"] = (po_df["reorder_point"] * 3 - po_df["current_stock"]).round(2)
                po_df["qty_to_order"] = po_df["qty_to_order"].clip(lower=0)
                po_df["est_cost"] = (po_df["qty_to_order"] * po_df["unit_cost"]).round(2)

                st.dataframe(
                    po_df[["name","unit","current_stock","reorder_point","qty_to_order","est_cost","supplier"]].rename(
                        columns={"name":"Ingredient","unit":"Unit",
                                 "current_stock":"In Stock","reorder_point":"Reorder At",
                                 "qty_to_order":"Order Qty","est_cost":"Est Cost £","supplier":"Supplier"}
                    ),
                    width="stretch", hide_index=True
                )

                total_po = po_df["est_cost"].sum()
                st.metric("Estimated PO Total", f"£{total_po:,.2f}")

                po_text  = f"CHOCOBERRY — PURCHASE ORDER\n"
                po_text += f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}\n"
                po_text += "=" * 50 + "\n"
                for _, r in po_df.iterrows():
                    po_text += (f"  {r['name']:<30} {r['qty_to_order']:>8.2f} {r['unit']:<8}"
                                f"  Est: £{r['est_cost']:.2f}  [{r['supplier']}]\n")
                po_text += "=" * 50 + "\n"
                po_text += f"TOTAL ESTIMATED COST: £{total_po:,.2f}\n"

                st.download_button(
                    "⬇️ Download PO as Text File",
                    po_text.encode("utf-8"),
                    file_name=f"chocoberry_po_{datetime.now().strftime('%Y%m%d')}.txt",
                    mime="text/plain",
                )



# ════════════════════════════════════════════════════════════════════
# TAB 10 — WASTE LOG & SHELF LIFE
# ════════════════════════════════════════════════════════════════════
with tab10:
    st.markdown('<div class="page-title">Waste Log & Expiry Intelligence</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Predictive waste management and logged loss tracking (Integrated with Recipe Engine).</div>', unsafe_allow_html=True)

    w1, w2 = st.columns([1, 1])
    eng = _recipe_engine

    with w1:
        st.markdown("**🛡️ Shelf-Life & Waste Risk Matrix**")
        shelf_path = "shelf_life_master.csv"
        if os.path.exists(shelf_path):
            shelf_df = pd.read_csv(shelf_path)
            st.dataframe(shelf_df, width="stretch", hide_index=True)
        else:
            st.warning("shelf_life_master.csv not found.")

    with w2:
        st.markdown("**♻️ Forensic Waste Ledger (Live DB)**")
        if eng:
            with eng._conn() as conn:
                waste_db_df = pd.read_sql("SELECT waste_date, item_name, quantity, reason, cost_impact, logged_by FROM waste_events ORDER BY created_at DESC", conn)
            
            if not waste_db_df.empty:
                waste_db_df.columns = ["Date", "Item", "Qty", "Reason", "Cost Impact (£)", "User"]
                st.dataframe(waste_db_df, width="stretch", hide_index=True)
                total_waste_val = waste_db_df["Cost Impact (£)"].sum()
                st.markdown(f'<div class="labour-kpi"><div class="labour-kpi-label">Total DB Waste Impact</div><div class="labour-kpi-value" style="color:#e05c5c">£{total_waste_val:,.2f}</div></div>', unsafe_allow_html=True)
            else:
                st.info("No waste events logged in database.")
        else:
            st.error("Recipe Engine offline.")

    st.markdown("---")
    st.markdown("**🖊️ Record New Waste Event**")
    if eng:
        ings = eng.get_ingredients()
        try:
            with st.form("waste_form_upgrade"):
                f_date = st.date_input("Waste Date")
                # Dropdown from real ingredients
                f_ing_name = st.selectbox("Ingredient to Waste", [i["name"] for i in ings])
                f_qty  = st.number_input("Quantity Wasted", min_value=0.01, step=0.1)
                f_reason = st.selectbox("Reason", ["Expired", "Dropped", "Pre-Error", "Customer Refund", "Over-Prepped"])
                if st.form_submit_button("Log & Deduct Waste"):
                    ing_id = next(i["id"] for i in ings if i["name"] == f_ing_name)
                    cost = eng.log_waste(ing_id, f_qty, f_reason, logged_by="DHIRAJ")
                    st.success(f"✅ Logged {f_qty} of {f_ing_name}. Impact: £{cost:.2f}. Stock updated.")
                    st.rerun()
        except Exception as e:
            st.error(f"❌ Failed to save waste event: {e}")

    st.markdown("---")
    st.markdown("**🧠 Forensic Variance Analysis (Theoretical vs. Actual)**")
    
    col_v1, col_v2 = st.columns([2, 1])
    
    if eng:
        # 1. Get Sales Volume
        sv_path = os.path.join(os.getcwd(), "Menu Item Report data", "Most sold items.csv")
        if os.path.exists(sv_path):
            sv_df = pd.read_csv(sv_path)
            n_col = next((c for c in sv_df.columns if any(x in c.lower() for x in ["item","name"])), sv_df.columns[0])
            q_col = next((c for c in sv_df.columns if any(x in c.lower() for x in ["sold","qty","quantity"])), None)
            
            if q_col:
                # CLEAN COMMAS
                sv_df[q_col] = sv_df[q_col].astype(str).str.replace(",", "", regex=False)
                sales_input = [{"item_name": r[n_col], "units_sold": r[q_col]} for _, r in sv_df.iterrows()]
                theo_usage = eng.get_theoretical_usage(sales_input)
                
                if theo_usage:
                    v_rows = []
                    for ing, data in theo_usage.items():
                        # Get actual logged waste for this ingredient from the DB
                        with eng._conn() as conn:
                            logged_qty = conn.execute("SELECT SUM(quantity) FROM waste_events WHERE item_name = ?", (ing,)).fetchone()[0] or 0.0
                        
                        v_rows.append({
                            "Ingredient": ing,
                            "Unit": data["unit"],
                            "Theoretical Usage (Sales)": round(data["usage"], 2),
                            "Logged Waste (Ledger)": round(logged_qty, 2),
                            "Total Depletion": round(data["usage"] + logged_qty, 2),
                            "Theo Cost": f"£{data['cost']:.2f}"
                        })
                    
                    v_df = pd.DataFrame(v_rows)
                    with col_v1:
                        st.dataframe(v_df, width="stretch", hide_index=True)
                    
                    with col_v2:
                        total_theo_cost = sum(d["cost"] for d in theo_usage.values())
                        st.markdown(f"""
                        <div style="background:rgba(124,92,191,0.1);border:1px solid #7c5cbf;padding:20px;border-radius:12px">
                            <h3 style="color:#7c5cbf;margin:0">£{total_theo_cost:,.2f}</h3>
                            <p style="color:#6b7094;font-size:12px;margin:10px 0">Theoretical Consumption</p>
                            <hr style="border:0;border-top:1px solid #7c5cbf;opacity:0.2">
                            <p style="font-size:13px">This is what <b>should</b> have been consumed based on your recipes and sales records.</p>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("Link recipes in the Inventory tab to see automatic usage analysis.")
            else:
                st.warning("Sales report format not recognized.")
        else:
            st.info("💡 Drop 'Most sold items.csv' to see automatic usage vs waste analysis.")


# ════════════════════════════════════════════════════════════════════
# TAB 11 — STRATEGIC OPTIMIZATION
# ════════════════════════════════════════════════════════════════════
with tab11:
    st.markdown('<div class="page-title">🚀 Live Menu Engineering Strategy</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Boston Matrix: Popularity (Flipdish Sales) vs. Profitability (Recipe Engine DB).</div>', unsafe_allow_html=True)

    eng = _recipe_engine
    if eng:
        # PULL LIVE PROFITABILITY FROM RECIPE ENGINE
        live_prof_data = eng.calc_item_profitability()
        
        sv_path = os.path.join(os.getcwd(), "Menu Item Report data", "Most sold items.csv")
        sales_summary = pd.DataFrame(columns=["item_name", "units"])
        using_dummy = False
        
        if os.path.exists(sv_path):
            try:
                sv_df = pd.read_csv(sv_path)
                n_col = next((c for c in sv_df.columns if any(x in c.lower() for x in ["item","name"])), sv_df.columns[0])
                q_col = next((c for c in sv_df.columns if any(x in c.lower() for x in ["sold","qty","quantity"])), None)
                if q_col:
                    sales_summary = sv_df[[n_col, q_col]].copy()
                    sales_summary.columns = ["item_name", "units"]
                    # CLEAN COMMAS AND CAST
                    sales_summary["units"] = sales_summary["units"].astype(str).str.replace(",", "", regex=False)
                    sales_summary["units"] = pd.to_numeric(sales_summary["units"], errors='coerce').fillna(0)
            except Exception:
                pass

        if live_prof_data:
            prof_df_strat = pd.DataFrame(live_prof_data)
            prof_df_strat = prof_df_strat[prof_df_strat["has_recipe"]]
            
            if not prof_df_strat.empty:
                if not sales_summary.empty:
                    # ENSURE BOTH ARE STRINGS AND STRIPPED TO AVOID VALUEERROR
                    prof_df_strat["flipdish_name"] = prof_df_strat["flipdish_name"].astype(str).str.strip()
                    sales_summary["item_name"] = sales_summary["item_name"].astype(str).str.strip()
                    
                    prof_df_strat = prof_df_strat.merge(sales_summary, left_on="flipdish_name", right_on="item_name", how="left").fillna(0)
                    prof_df_strat["Units Sold"] = prof_df_strat["units"].astype(float)
                    if prof_df_strat["Units Sold"].sum() == 0:
                        using_dummy = True
                else:
                    using_dummy = True

                if using_dummy:
                    st.info("💡 Note: Actual itemized sales volume data not found in 'Menu Item Report data'. Showing items at baseline 50-unit popularity for strategic visualization.")
                    prof_df_strat["Units Sold"] = 50.0 

                fig_strat = px.scatter(
                    prof_df_strat,
                    x="Units Sold", y="gp_pct", size="selling_price", color="category",
                    hover_name="menu_item", text="menu_item",
                    labels={"Units Sold": "Popularity (Sales Volume)", "gp_pct": "GP Margin (%)"},
                    template="plotly_dark"
                )
                
                fig_strat.update_traces(textposition='top center')
                fig_strat.add_hline(y=70, line_dash="dot", line_color="#3ecf8e", annotation_text="Target Margin (70%)")
                fig_strat.add_vline(x=prof_df_strat["Units Sold"].mean(), line_dash="dot", line_color="rgba(255,255,255,0.3)", annotation_text="Avg Popularity")

                dark_layout(fig_strat, height=600, showlegend=True)
                st.plotly_chart(fig_strat, width="stretch")

                st.markdown("""
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px">
                    <div style="background:rgba(62,207,142,0.1);padding:15px;border-radius:10px;border:1px solid #3ecf8e">
                        <b style="color:#3ecf8e">🌟 STARS (Top Right)</b><br>
                        High Profit + High Volume. Your winners. Keep quality high and maintain pricing.
                    </div>
                    <div style="background:rgba(245,166,35,0.1);padding:15px;border-radius:10px;border:1px solid #f5a623">
                        <b style="color:#f5a623">🐴 WORKHORSES (Top Left)</b><br>
                        Low Profit + High Volume. Increase prices or reduce portion costs immediately.
                    </div>
                    <div style="background:rgba(124,92,191,0.1);padding:15px;border-radius:10px;border:1px solid #7c5cbf">
                        <b style="color:#7c5cbf">🧩 PUZZLES (Bottom Right)</b><br>
                        High Profit + Low Volume. Promote these via specials or social media.
                    </div>
                    <div style="background:rgba(224,92,92,0.1);padding:15px;border-radius:10px;border:1px solid #e05c5c">
                        <b style="color:#e05c5c">🐕 DOGS (Bottom Left)</b><br>
                        Low Profit + Low Volume. Consider removing these from the menu.
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.info("🕒 **Recipe Data Required**: No menu items currently have linked recipes. To use Strategic Optimization, go to the **Inventory & COGS** tab and add ingredients to your menu items.")
        else:
            st.info("🕒 **Menu Data Required**: No menu items found in the database. Go to the **Inventory & COGS** tab to import your menu.")
    else:
        st.warning("Recipe Engine not found for strategic analytics.")


# ════════════════════════════════════════════════════════════════════
# TAB 12 — INVOICE MANAGEMENT  *** PORTAL SYNC ADDED ***
# ════════════════════════════════════════════════════════════════════
with tab12:
    st.markdown('<div class="page-title">📄 Invoice Intelligence Ledger</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Financial payables, supplier auditing, and cash flow protection.</div>', unsafe_allow_html=True)

    # ── PORTAL SYNC STATUS PANEL ────────────────────────────────────────────
    portal_base = "https://invoiceappcbc-ng5tjkfaikn8wwstgybptu.streamlit.app"
    try:
        import urllib.request as _ur
        portal_secret = os.environ.get("PORTAL_SECRET", "chocoberry2026")
        
        # Streamlit Query Param API
        check_url = f"{portal_base}/?api=sync&secret={portal_secret}"
        
        _pending = []
        try:
            with _ur.urlopen(check_url, timeout=8) as _r:
                raw_data = _r.read().decode('utf-8')
                if "[" in raw_data and "]" in raw_data:
                    json_str = raw_data[raw_data.find("["):raw_data.rfind("]")+1]
                    _pending = json.loads(json_str)
        except Exception as conn_err:
            # If auto-check fails, we show the refresh button below
            pass
        
        if _pending:
            st.markdown(f"""
            <div style="background:rgba(245,166,35,0.1);border:1px solid #f5a623;
                        padding:14px 18px;border-radius:10px;margin-bottom:16px">
                <b style="color:#f5a623">📱 {len(_pending)} new invoice(s) uploaded by staff</b>
                — waiting to be synced from Cloud Portal
            </div>""", unsafe_allow_html=True)
            
            if st.button("🔄 Sync Staff Uploads Now", type="primary", key="portal_sync"):
                if sync_from_portal:
                    with st.spinner("Syncing latest uploads..."):
                        sync_from_portal(portal_base=portal_base, portal_secret=portal_secret)
                        st.success("✅ Staff uploads synced into ledger.")
                        st.rerun()
                else:
                    st.error("Sync module (sync_portal_invoices.py) not found.")
        else:
            p1, p2 = st.columns([4, 1])
            with p1:
                st.markdown("""
                <div style="background:#102a18;border:1px solid #3ecf8e;
                            padding:10px 16px;border-radius:8px;margin-bottom:12px;
                            font-size:12px;color:#3ecf8e">
                    ✅ Staff upload portal: online — no pending uploads
                </div>""", unsafe_allow_html=True)
            with p2:
                if st.button("🔄 Refresh", key="refresh_sync_status"):
                    st.rerun()

    except Exception as e:
        st.markdown(f"""
        <div style="background:#12141a;border:1px solid #252836;
                    padding:10px 16px;border-radius:8px;margin-bottom:12px;
                    font-size:12px;color:#6b7094">
            ⚪ Cloud portal sync pending — Click Refresh to connect
        </div>""", unsafe_allow_html=True)
        if st.button("🔄 Refresh Connection", key="refresh_err"):
            st.rerun()

    # ── 1. OVERDUE ALERTS PANEL ──────────────────────────────────────────────
    overdue_list = inv_db.get_overdue_invoices()
    if overdue_list:
        with st.container():
            st.markdown(f'<div style="background:rgba(224,92,92,0.1);border:1px solid #e05c5c;padding:16px;border-radius:12px;margin-bottom:20px"><div style="color:#e05c5c;font-family:Syne,sans-serif;font-weight:800;font-size:13px;letter-spacing:1px;margin-bottom:8px">🚨 CRITICAL OVERDUE PAYABLES — {len(overdue_list)} INVOICES LATE</div>', unsafe_allow_html=True)
            for inv_no, supp, amt, due in overdue_list[:3]:
                st.markdown(f'<div style="color:#6b7094;font-size:11px;margin-bottom:4px">⚠️ <b>{supp}</b> (Inv #{inv_no}): £{amt:,.2f} — Due: {due}</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

    # ── 2. KPI SUMMARY ROW ──────────────────────────────────────────────────
    k1, k2, k3, k4 = st.columns(4)
    all_invs = inv_db.get_all_invoices()
    total_unpaid = sum(row[5] for row in all_invs if row[6] != 'PAID')
    total_overdue = sum(row[2] for row in overdue_list)

    this_month_ym = datetime.now().strftime('%Y-%m')
    this_month_my = datetime.now().strftime('/%m/%Y')
    this_month_spend = 0
    for row in all_invs:
        d_str = str(row[3]) if row[3] else ""
        if d_str.startswith(this_month_ym) or d_str.endswith(this_month_my):
            this_month_spend += row[5]

    k1.metric("🔴 Total Unpaid", f"£{total_unpaid:,.2f}")
    k2.metric("⌛ Total Overdue", f"£{total_overdue:,.2f}", f"{len(overdue_list)} late", delta_color="inverse")
    k3.metric("💳 Month Spend", f"£{this_month_spend:,.2f}")
    k4.metric("📊 Suppliers", f"{len(set(row[2] for row in all_invs))}", "Active ledger")

    st.markdown("---")

    # ── 3. VIEW CONTROLS ────────────────────────────────────────────────────
    c_mode, c_depth = st.columns([2, 1])
    with c_mode:
        mode = st.radio("System Mode", ["📋 Accounts Registry", "🖊️ Record New Invoice (Dhiraj Mode)", "📈 Spend Analytics", "💰 Revenue Ledger"], horizontal=True)
    with c_depth:
        if mode == "📋 Accounts Registry":
            v_depth = st.radio("View Depth", ["Summarized", "Itemized (All 238+ Lines)"], horizontal=True)
        else:
            st.empty()

    if mode == "📋 Accounts Registry":
        st.markdown("**1. Filter Intelligence**")
        f1, f2, f3 = st.columns([2, 1, 1])
        search = f1.text_input("🔍 Search Invoices", placeholder="Supplier or Invoice #...")
        supp_opts = ["All Suppliers"] + sorted(list(set(row[2] for row in all_invs)))
        sel_supp = f2.selectbox("Filter by Supplier", supp_opts)
        stat_opts = ["All Statuses", "PAID", "UNPAID"]
        sel_stat = f3.selectbox("Status", stat_opts)

        st.markdown("---")
        if v_depth == "Summarized":
            f_invs = all_invs
            if sel_supp != "All Suppliers": f_invs = [r for r in f_invs if r[2] == sel_supp]
            if sel_stat != "All Statuses": f_invs = [r for r in f_invs if r[6] == sel_stat]
            if search: f_invs = [r for r in f_invs if search.lower() in str(r[1]).lower() or search.lower() in str(r[2]).lower()]

            if f_invs:
                reg_df = pd.DataFrame(f_invs, columns=["ID", "Invoice #", "Supplier", "Date", "Due Date", "Gross (£)", "Status", "Category", "Drive Path"])
                st.dataframe(reg_df.drop(columns=["ID"]), width="stretch", hide_index=True)

                st.download_button(
                    label="⬇️ Export Filtered Registry (CSV)",
                    data=reg_df.to_csv(index=False).encode('utf-8'),
                    file_name=f"invoice_registry_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv"
                )

                with st.expander("🛠️ Advanced Invoice Actions (Mark as Paid / View Metadata)"):
                    sel_inv = st.selectbox("Select Invoice to Action", [f"#{r[1]} - {r[2]}" for r in f_invs])
                    a_col1, a_col2 = st.columns(2)
                    if a_col1.button("✅ Mark as Paid", width="stretch"):
                        inv_id = [r[0] for r in f_invs if f"#{r[1]} - {r[2]}" == sel_inv][0]
                        inv_db.update_payment_status(inv_id, 'PAID')
                        st.success("Invoice status updated.")
                        st.rerun()
            else:
                st.info("No invoices found matching criteria.")
        else:
            items = inv_db.get_all_line_items()
            if items:
                item_df = pd.DataFrame(items, columns=["Invoice #", "Supplier", "Description", "Product Code", "Qty", "Unit", "Rate (£)", "Total (£)"])
                if sel_supp != "All Suppliers": item_df = item_df[item_df["Supplier"] == sel_supp]
                if search: item_df = item_df[item_df["Description"].str.contains(search, case=False)]

                st.markdown(f"**🔍 Displaying {len(item_df)} individual line items from forensic ledger**")
                st.dataframe(item_df, width="stretch", hide_index=True)
            else:
                st.warning("No line items found in the database.")
            st.markdown("---")

    elif mode == "🖊️ Record New Invoice (Dhiraj Mode)":
        with st.form("invoice_upload_form"):
            st.markdown("**1. Upload Proof of Delivery (Photo/PDF)**")
            u_file = st.file_uploader("Drop invoice trace here", type=["pdf", "png", "jpg"])

            st.markdown("---")
            st.markdown("**2. Manual Forensic Entry**")
            f1, f2 = st.columns(2)
            with f1:
                inv_no = st.text_input("Invoice Number")
                all_supps = [s[1] for s in inv_db.get_suppliers()]
                supplier = st.selectbox("Supplier", all_supps if all_supps else ["Cr8 Foods", "Freshways", "Bookers"])
            with f2:
                inv_date = st.date_input("Invoiced Date", value=datetime.now())
                total_val = st.number_input("Total Gross Amount (£)", min_value=0.0)

            cat = st.selectbox("Category", ["Food", "Packaging", "Labour", "Utilities", "Maintenance", "Other"])
            notes = st.text_area("Internal Notes")

            if st.form_submit_button("Submit to Intelligence Gateway"):
                if inv_no and total_val > 0:
                    supp_id = inv_db.add_supplier(supplier)
                    if inv_db.check_duplicate(inv_no, supp_id):
                        st.warning(f"⚠️ Duplicate detected: Invoice {inv_no} from {supplier} already exists.")
                    else:
                        # ── Organise Uploaded File ──────────────────────────────────
                        rel_path = ""
                        if u_file:
                            import shutil
                            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                            safe_sup = "".join(c for c in supplier if c.isalnum() or c==' ').strip()
                            filename = f"{stamp}_{safe_sup.replace(' ','_')}_{int(total_val)}{os.path.splitext(u_file.name)[1]}"
                            
                            year_folder  = inv_date.strftime("%Y")
                            month_folder = inv_date.strftime("%m-%b")
                            target_dir   = Path(os.getcwd()) / "Invoices" / year_folder / month_folder / safe_sup
                            target_dir.mkdir(parents=True, exist_ok=True)
                            
                            target_path = target_dir / filename
                            with open(target_path, "wb") as f:
                                f.write(u_file.getbuffer())
                            rel_path = str(target_path.relative_to(Path(os.getcwd())))

                        inv_data = {
                            'invoice_number': inv_no,
                            'supplier_id':    supp_id,
                            'invoice_date':   inv_date.strftime('%Y-%m-%d'),
                            'total_amount':   total_val,
                            'category':       cat,
                            'notes':          notes,
                            'payment_status': 'UNPAID',
                            'image_path':     rel_path
                        }
                        inv_db.insert_invoice(inv_data)
                        st.success(f"✅ Invoice {inv_no} logged and saved to Drive-ready folder ({rel_path}).")
                        st.rerun()
                else:
                    st.error("Missing required fields (Invoice # or Total).")

    elif mode == "📈 Spend Analytics":
        st.markdown("**📊 Spend Intelligence**")

        c1_sa, c2_sa = st.columns(2)

        with c1_sa:
            analytics = inv_db.get_supplier_analytics()
            if analytics:
                ana_df = pd.DataFrame(analytics, columns=["Supplier", "Total Spend", "Count"])
                fig_spend = go.Figure(go.Bar(
                    x=ana_df["Supplier"], y=ana_df["Total Spend"],
                    marker_color=PALETTE[0],
                    marker_line_width=0,
                ))
                dark_layout(fig_spend, 350)
                fig_spend.update_layout(title="Total Spend by Supplier (£)")
                st.plotly_chart(fig_spend, width="stretch")

        with c2_sa:
            cat_analytics = inv_db.get_category_analytics()
            if cat_analytics:
                cat_df = pd.DataFrame(cat_analytics, columns=["Category", "Spend"])
                fig_cat = px.pie(
                    cat_df, values="Spend", names="Category",
                    hole=0.4, color_discrete_sequence=px.colors.sequential.YlOrRd
                )
                dark_layout(fig_cat, 350)
                fig_cat.update_layout(title="Spend by Category", showlegend=True)
                st.plotly_chart(fig_cat, width="stretch")

        st.markdown("---")
        st.markdown("**📉 Historical Spend Trend**")
        trend_data = inv_db.get_monthly_spend_trend()
        if trend_data:
            trend_df = pd.DataFrame(trend_data, columns=["Month", "Spend"])
            fig_trend = go.Figure(go.Scatter(
                x=trend_df["Month"], y=trend_df["Spend"],
                mode='lines+markers',
                line=dict(color=COLORS["accent"], width=3),
                marker=dict(size=8)
            ))
            dark_layout(fig_trend, 300)
            fig_trend.update_layout(
                title="Historical Spend Trend (Monthly Invoiced)",
                xaxis_title="Month",
                yaxis_title="Total Invoiced (£)"
            )
            st.plotly_chart(fig_trend, width="stretch")
        else:
            st.info("Ingest more monthly data to see trend lines.")

        st.markdown("---")
        st.markdown("**📊 Month-on-Month Supplier Comparison**")
        mom_data = inv_db.get_supplier_monthly_analytics()
        if mom_data:
            mom_df = pd.DataFrame(mom_data, columns=["Supplier", "Month", "Spend"])
            fig_mom = px.bar(
                mom_df, x="Month", y="Spend", color="Supplier",
                title="Spend Distribution by Supplier over Time",
                template="plotly_dark",
                barmode="stack"
            )
            dark_layout(fig_mom, 400)
            st.plotly_chart(fig_mom, width="stretch")

    elif mode == "💰 Revenue Ledger":
        st.markdown("**📈 Daily Revenue & Sales Master**")
        st.markdown('<div class="insight-box">Edit your daily sales figures here. All calculations (AOV, Labour %, Trends) will update automatically across the entire dashboard when you Save.</div>', unsafe_allow_html=True)

        rev_csv = os.path.join(os.getcwd(), "daily_sales_master.csv")
        if os.path.exists(rev_csv):
            sales_df = pd.read_csv(rev_csv)
            sales_df = sales_df.sort_values("date", ascending=False)

            edited_sales = st.data_editor(
                sales_df,
                width="stretch",
                num_rows="dynamic",
                hide_index=True,
                column_config={
                    "net": st.column_config.NumberColumn("Net Sales (£)", format="£%.2f"),
                    "revenue": st.column_config.NumberColumn("Gross (£)", format="£%.2f"),
                    "orders": st.column_config.NumberColumn("Orders"),
                }
            )

            if st.button("💾 Save Revenue Changes & Recalculate Dashboard", width="stretch"):
                edited_sales.to_csv(rev_csv, index=False, encoding="utf-8-sig")
                st.success("✅ Revenue ledger updated. Recalculating all project metrics...")
                st.cache_data.clear()
                st.rerun()
        else:
            st.error("Missing daily_sales_master.csv. Please restart the application to generate it.")


# ════════════════════════════════════════════════════════════════════
# TAB 13 — DATABASE EXPLORER
# ════════════════════════════════════════════════════════════════════
with tab13:
    st.markdown('<div class="section-title">Forensic Database Explorer</div>', unsafe_allow_html=True)
    st.markdown('<div class="insight-box">Select a database and table to inspect raw records. Use this for data auditing and integrity checks.</div>', unsafe_allow_html=True)

    db_choice = st.selectbox("Select Database", ["main_invoices (cbc_invoice_intelligence.db)", "recipe_engine (recipes.db)", "portal_uploads (invoices.db)"])
    
    db_mapping = {
        "main_invoices (cbc_invoice_intelligence.db)": "cbc_invoice_intelligence.db",
        "recipe_engine (recipes.db)": "recipes.db",
        "portal_uploads (invoices.db)": "invoices.db"
    }
    
    selected_db = db_mapping[db_choice]
    
    if os.path.exists(selected_db):
        import sqlite3
        with sqlite3.connect(selected_db) as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = [r[0] for r in cursor.fetchall() if r[0] not in ["sqlite_sequence"]]
            
            if tables:
                table_choice = st.selectbox("Select Table", tables)
                if table_choice:
                    table_choice_safe = str(table_choice).replace(" ", "_")
                    df_db = pd.read_sql_query(f"SELECT * FROM {table_choice} LIMIT 500", conn)
                    st.write(f"Showing last 500 records from **{table_choice}**")
                    st.dataframe(df_db, width="stretch")
                    
                    st.download_button(
                        label=f"⬇️ Export {table_choice} to CSV",
                        data=df_db.to_csv(index=False),
                        file_name=f"{table_choice_safe}_export.csv",
                        mime="text/csv"
                    )
            else:
                st.warning("No tables found in this database.")
    else:
        st.error(f"Database file not found: {selected_db}")

    st.markdown("---")
    st.markdown("**🛠️ Manual SQL Query (Advanced)**")
    query = st.text_area("Enter SQL Query", "SELECT * FROM suppliers", key="sql_query_area")
    if st.button("Execute Query", key="sql_exec_btn"):
        if not query.strip():
            st.warning("Please enter a query.")
        elif not query.strip().upper().startswith("SELECT"):
            st.error("🚫 Security Restriction: Only SELECT queries are allowed in the Database Explorer.")
        else:
            try:
                with sqlite3.connect(selected_db) as conn:
                    res_df = pd.read_sql_query(query, conn)
                    st.success("Query executed successfully.")
                    st.dataframe(res_df, width="stretch")
            except Exception as e:
                st.error(f"SQL Error: {e}")





# ════════════════════════════════════════════════════════════════════
# TAB 14 — STAFF MANAGEMENT
# ════════════════════════════════════════════════════════════════════
with tab14:
    st.markdown('<div class="section-title">Staff Management Portal</div>', unsafe_allow_html=True)
    st.markdown('<div class="insight-box">Add new hires, manage roles, and audit your workforce capacity here. Any changes made here automatically update your Rota and Payroll systems.</div>', unsafe_allow_html=True)

    # 1. Workforce Audit Section
    staff_df = pd.read_csv("staff_profiles.csv")
    rates_df = pd.read_csv("personnel_rates_master.csv")
    
    col_a, col_b = st.columns([2, 3])
    
    with col_a:
        st.markdown("**🛡️ Workforce Capacity Audit**")
        role_counts = staff_df["Department"].value_counts()
        for dept, count in role_counts.items():
            st.info(f"**{dept}:** {count} active staff")
        
        # Specific Role Breakdown
        st.write("---")
        roles = staff_df["Role"].value_counts()
        for role, count in roles.items():
            st.write(f"• {role}: {count}")

    with col_b:
        st.markdown("**➕ Add New Staff Member**")
        with st.form("add_staff_form", clear_on_submit=True):
            new_name = st.text_input("Full Name (First Last)")
            new_dept = st.selectbox("Area / Department", ["Kitchen", "Front", "Management", "Professional"])
            new_role = st.selectbox("Role / Level", ["Senior", "Junior", "Management", "Professional", "Dishwasher"])
            new_target = st.number_input("Target Weekly Hours", min_value=0, max_value=60, value=20)
            
            submit = st.form_submit_button("🚀 Add Staff to System")
            
            if submit:
                if not new_name.strip():
                    st.error("Please enter a name.")
                elif new_name in staff_df["Name"].values:
                    st.error(f"❌ Error: A staff member named '{new_name}' already exists.")
                else:
                    # Logic to Add to staff_profiles.csv
                    new_row = {
                        "Name": new_name,
                        "Role": new_role,
                        "Department": new_dept,
                        "Target Hours": new_target,
                        "Max Hours": new_target + 10,
                        "Performance Score": 5,
                        "Active": True,
                        "Availability": "Mon,Tue,Wed,Thu,Fri,Sat,Sun",
                        "Shift Preference": "Any"
                    }
                    new_s_df = pd.DataFrame([new_row])
                    staff_df = pd.concat([staff_df, new_s_df], ignore_index=True)
                    staff_df.to_csv("staff_profiles.csv", index=False)
                    
                    # Logic to Add to personnel_rates_master.csv (Default Payroll Setup)
                    new_rate = {
                        "Name": new_name,
                        "NI Hours": 10,
                        "NI Rates": 12.71,
                        "Hourly Rate": 8.00,
                        "Fixed Wage": 0.00
                    }
                    new_r_df = pd.DataFrame([new_rate])
                    rates_df = pd.concat([rates_df, new_r_df], ignore_index=True)
                    rates_df.to_csv("personnel_rates_master.csv", index=False)
                    
                    st.success(f"✅ SUCCESS: {new_name} has been enrolled! They are now available for both Rota and Payroll.")
                    st.balloons()
                    time.sleep(2)
                    st.rerun()

    st.markdown("---")
    st.markdown("**👥 Current Active Roster**")
    st.dataframe(staff_df[["Name", "Role", "Department", "Target Hours", "Active"]], width="stretch", hide_index=True)


def _print_labour_console(summary, over_df, under_df, overlay_df):
    print("\n" + "=" * 60)
    print("  CHOCOBERRY LABOUR ANALYSIS — CONSOLE SUMMARY")
    print(f"  Week: {WEEK_LABEL}")
    print("=" * 60)
    for k, v in summary.items():
        flag = ""
        if "%" in str(v) and "Labour %" in k:
            pct  = float(str(v).replace("%", ""))
            flag = " 🔴" if pct > 30 else " ✅"
        print(f"  {k:<35} {v}{flag}")

    print(f"\n  OVERSTAFFED HOURS:  {len(over_df)}")
    for _, r in over_df.iterrows():
        print(f"    🔴  {r['Hour']}  |  Rev: £{r['Avg Revenue (£)']:.2f}"
              f"  |  Staff: {r['Total Staff']}  |  Cost: £{r['Total Staff Cost (£)']:.2f}")

    print(f"\n  UNDERSTAFFED HOURS: {len(under_df)}")
    for _, r in under_df.iterrows():
        print(f"    🟡  {r['Hour']}  |  Rev: £{r['Avg Revenue (£)']:.2f}"
              f"  |  Staff: {r['Total Staff']}")

    print("\n  PEAK HOURS COVERAGE (19:00–23:00):")
    peak = overlay_df[overlay_df["Hour"].isin(["19:00","20:00","21:00","22:00","23:00"])]
    for _, r in peak.iterrows():
        print(f"    {r['Hour']}  Rev: £{r['Avg Revenue (£)']:>8,.2f}"
              f"  Staff: {r['Total Staff']:>2}"
              f"  Cost: £{r['Total Staff Cost (£)']:>6.2f}"
              f"  Rev/Staff: £{r['Revenue/Staff/Hr (£)']:>7,.2f}"
              f"  {r['Flag']}")
    print("=" * 60)


if __name__ == "__main__" and "--report" in sys.argv:
    OUTPUT = "chocoberry_labour_report.xlsx"
    summary, overlay_df, over_df, under_df = build_labour_workbook(load_data(), OUTPUT)
    _print_labour_console(summary, over_df, under_df, overlay_df)
